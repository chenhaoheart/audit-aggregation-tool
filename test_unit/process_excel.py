
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter

file_path = r'D:\山洪\项目开发\青海2025\青海25外业成果提交\景文团队数据\互助县\隐患点计算\横断面数据_跨沟道路和桥涵(1).xlsx'

# 读取Excel文件
wb = openpyxl.load_workbook(file_path)
ws = wb.active

print('数据行数:', ws.max_row)
print('数据列数:', ws.max_column)

print('\n列名:')
for col_idx in range(1, ws.max_column + 1):
    col_letter = get_column_letter(col_idx)
    cell_value = ws[f'{col_letter}1'].value
    print(f'{col_idx}. {col_letter}: {cell_value}')

# 读取F列（第6列）、AA列（第27列）、AG列（第33列）的数据
data = []
for row_idx in range(2, ws.max_row + 1):
    f_value = ws[f'F{row_idx}'].value
    aa_value = ws[f'AA{row_idx}'].value
    ag_value = ws[f'AG{row_idx}'].value
    data.append({
        'row': row_idx,
        'f': f_value,
        'aa': aa_value,
        'ag': ag_value
    })

print(f'\n读取了 {len(data)} 行数据')

# 按F列分组
groups = {}
for item in data:
    key = item['f']
    if key not in groups:
        groups[key] = []
    groups[key].append(item)

print(f'\n分组数量: {len(groups)}')
for key, group in sorted(groups.items()):
    print(f'  {key}: {len(group)} 行')

# 在每个分组内按序号填充
for key, group in groups.items():
    for idx, item in enumerate(group, 1):
        row_idx = item['row']
        ws[f'AA{row_idx}'] = idx
        print(f'设置第 {row_idx} 行 AA 列为 {idx} (分组: {key})')

# 保存文件
output_path = file_path.replace('.xlsx', '_已处理.xlsx')
wb.save(output_path)
print(f'\n文件已保存到: {output_path}')
