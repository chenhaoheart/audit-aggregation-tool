import openpyxl
import json
import os
import glob
from datetime import datetime
from typing import Dict, List, Tuple, Optional, Any


import numpy as np
import pandas as pd
import math





BASE_DIR = r"C:\Users\chenh\Desktop\新建文件夹\2025省平台数据入库\断面\63个横纵断面0423"

RIVER_DIRS = [
    "全省",
]

OUTPUT_DIR = BASE_DIR

COL_SECTION = 12
COL_HMZ = 24
COL_CZZ = 25
COL_BACKWATER = 3
COL_POINT_SEQ = 26
COL_LON = 27
COL_LAT = 28
COL_DISTANCE = 32
COL_ELEVATION = 33
COL_FEATURE = 34
COL_DISTRICT = 7

FEATURE_KEYWORDS = ["基点", "堤顶", "左岸", "右岸", "深泓", "主槽", "滩地", "桥", "涵", "堰", "闸", "坝"]

ANOMALY_TYPES = {
    "cross": "线段交叉",
    "reversal": "方向反转",
}

VALIDATION_CONFIG = {
    "lon_range": (73.0, 135.0),
    "lat_range": (18.0, 54.0),
    "dist_jump_factor": 10,
    "coord_jump_threshold": 0.01,
    "coord_jump_ratio": 0.2,
    "linear_fit_threshold": 0.05,
}

VALIDATION_RULES = {
    "seq_missing": {"desc": "序号缺失", "level": "error"},
    "seq_duplicate": {"desc": "序号重复", "level": "error"},
    "seq_not_monotonic": {"desc": "序号非递增", "level": "warning"},
    "seq_gap": {"desc": "序号跳号", "level": "warning"},
    "dist_negative": {"desc": "起点距为负", "level": "error"},
    "dist_not_monotonic": {"desc": "起点距回退", "level": "error"},
    "dist_duplicate": {"desc": "起点距重复", "level": "warning"},
    "dist_jump": {"desc": "起点距跳变", "level": "warning"},
    "lon_out_of_range": {"desc": "经度超范围", "level": "error"},
    "lon_zero": {"desc": "经度为零或缺失", "level": "error"},
    "lon_jump": {"desc": "经度跳变", "level": "warning"},
    "lat_out_of_range": {"desc": "纬度超范围", "level": "error"},
    "lat_zero": {"desc": "纬度为零或缺失", "level": "error"},
    "lat_jump": {"desc": "纬度跳变", "level": "warning"},
    "lonlat_swapped": {"desc": "经纬度互换", "level": "error"},
    "coord_missing": {"desc": "经纬度缺失", "level": "warning"},
    "all_same_coord": {"desc": "所有点坐标相同", "level": "error"},
    "linear_fit_deviation": {"desc": "线性拟合偏差", "level": "warning"},
    "direction_reversal": {"desc": "连线方向折返", "level": "error"},
    "adjacent_reversal": {"desc": "相邻线段折返", "level": "error"},
}


# ==========================================
# 辅助函数
# ==========================================

def haversine(lon1, lat1, lon2, lat2):
    """
    计算两个经纬度点之间的 Haversine 距离 (单位: 米)
    """
    lon1, lat1, lon2, lat2 = map(math.radians, [lon1, lat1, lon2, lat2])
    dlon = lon2 - lon1
    dlat = lat2 - lat1
    a = math.sin(dlat / 2)**2 + math.cos(lat1) * math.cos(lat2) * math.sin(dlon / 2)**2
    c = 2 * math.asin(math.sqrt(a))
    r = 6371000  # 地球平均半径，单位米
    return c * r

def lonlat_to_local_xy(lon, lat, ref_lon, ref_lat):
    """
    将经纬度转换为以 (ref_lon, ref_lat) 为原点的局部平面直角坐标 (单位: 米)
    用于线性拟合，避免经纬度直接拟合导致的尺度不一致问题
    """
    x = (lon - ref_lon) * math.cos(math.radians(ref_lat)) * 111320.0
    y = (lat - ref_lat) * 110540.0
    return x, y

# ==========================================
# 核心检测引擎
# ==========================================

class SpatialContinuityChecker:
    def __init__(self, df):
        """
        初始化检测器
        df: Pandas DataFrame, 必须包含列: ['distance', 'lon', 'lat']
        """
        self.df = df.copy().reset_index(drop=True)
        self.n = len(df)
        # 预计算相邻段数据
        self._pre_calculate()

    def _pre_calculate(self):
        """预计算相邻点的距离、坐标距离、平面坐标等基础数据"""
        if self.n < 2:
            return
        
        # 起点距差值 (间距)
        self.chain_dists = np.abs(np.diff(self.df['distance'].values))
        
        # 坐标距离
        coord_dists = []
        for i in range(self.n - 1):
            d = haversine(
                self.df.loc[i, 'lon'], self.df.loc[i, 'lat'],
                self.df.loc[i+1, 'lon'], self.df.loc[i+1, 'lat']
            )
            coord_dists.append(d)
        self.coord_dists = np.array(coord_dists)
        
        ref_lon = self.df.loc[0, 'lon']
        ref_lat = self.df.loc[0, 'lat']
        self.local_coords = []
        for i in range(self.n):
            x, y = lonlat_to_local_xy(
                self.df.loc[i, 'lon'], self.df.loc[i, 'lat'],
                ref_lon, ref_lat
            )
            self.local_coords.append((x, y))
        
        if self.n >= 2:
            self.main_direction = self._compute_main_direction()

    def _compute_main_direction(self):
        """计算断面的整体主方向向量（起点到终点）"""
        start = self.local_coords[0]
        end = self.local_coords[-1]
        dx = end[0] - start[0]
        dy = end[1] - start[1]
        length = math.sqrt(dx**2 + dy**2)
        if length < 1:
            return None
        return (dx / length, dy / length)

    def check_dist_coord_mismatch(self, threshold=3.0):
        """
        规则1: 间距与坐标距离不匹配
        条件: |间距 - 坐标距离| / min(间距, 坐标距离) > threshold
        """
        flagged_indices = set()
        if self.n < 2: return flagged_indices
        
        for i in range(self.n - 1):
            A = self.chain_dists[i]
            B = self.coord_dists[i]
            
            min_val = min(A, B)
            max_val = max(A, B)
            
            if min_val == 0:
                # 如果较小值为0，只要另一个值不为0，就视为极度不匹配
                if max_val > 0:
                    flagged_indices.update([i, i+1])
            else:
                ratio = (max_val - min_val) / min_val
                if ratio > threshold:
                    flagged_indices.update([i, i+1])
                    
        return flagged_indices

    def check_linear_fit_deviation(self, threshold=0.05):
        """
        规则2: 线性拟合偏差
        条件: 点偏离断面拟合直线的距离 > 断面长度的 threshold (5%)
        """
        flagged_indices = set()
        if self.n < 3: return flagged_indices  # 少于3点无法拟合
        
        # 1. 转换为局部平面坐标
        ref_lon = self.df.loc[0, 'lon']
        ref_lat = self.df.loc[0, 'lat']
        x_vals = np.array([lonlat_to_local_xy(row['lon'], row['lat'], ref_lon, ref_lat)[0] for _, row in self.df.iterrows()])
        y_vals = np.array([lonlat_to_local_xy(row['lon'], row['lat'], ref_lon, ref_lat)[1] for _, row in self.df.iterrows()])
        
        # 2. 计算断面总长度 (使用起点距极差更符合业务逻辑)
        L = np.max(self.df['distance']) - np.min(self.df['distance'])
        if L <= 0:
            L = np.max(self.coord_dists) # 极端情况退化为坐标距离
            
        # 3. 使用 PCA (主成分分析) 拟合直线，比最小二乘法更稳健
        X = np.vstack([x_vals, y_vals]).T
        X_centered = X - np.mean(X, axis=0)
        cov_matrix = np.cov(X_centered.T)
        eig_vals, eig_vecs = np.linalg.eig(cov_matrix)
        
        # 最大特征值对应的特征向量即为直线方向
        max_idx = np.argmax(eig_vals)
        direction = eig_vecs[:, max_idx]
        
        # 计算法向量并归一化 (直线方程 Ax + By + C = 0)
        A = direction[1]
        B = -direction[0]
        C = -(A * np.mean(x_vals) + B * np.mean(y_vals))
        norm = np.sqrt(A**2 + B**2)
        A, B, C = A/norm, B/norm, C/norm
        
        # 4. 计算每个点到直线的垂距
        distances = np.abs(A * x_vals + B * y_vals + C)
        
        # 5. 判定
        for i in range(self.n):
            if distances[i] > threshold * L:
                flagged_indices.add(i)
                
        return flagged_indices

    def check_direction_consistency(self, reversal_threshold=90):
        """
        规则3: 方向一致性检测
        检查每条线段是否与断面主方向一致，夹角大于reversal_threshold度视为折返
        
        reversal_threshold: 折返判定阈值（度），默认90度
        """
        anomalies = []
        if self.n < 3 or self.main_direction is None:
            return anomalies
        
        reversal_angle_rad = math.radians(reversal_threshold)
        
        for i in range(self.n - 1):
            p1 = self.local_coords[i]
            p2 = self.local_coords[i + 1]
            
            dx = p2[0] - p1[0]
            dy = p2[1] - p1[1]
            seg_length = math.sqrt(dx**2 + dy**2)
            
            if seg_length < 0.5:
                continue
            
            seg_unit = (dx / seg_length, dy / seg_length)
            
            dot = seg_unit[0] * self.main_direction[0] + seg_unit[1] * self.main_direction[1]
            
            dot_clamped = max(-1.0, min(1.0, dot))
            angle_rad = math.acos(dot_clamped)
            angle_deg = math.degrees(angle_rad)
            
            if angle_deg > reversal_threshold:
                anomalies.append({
                    'type': 'reversal',
                    'point_index': i + 1,
                    'point_seq': self.df.loc[i + 1, 'seq'] if 'seq' in self.df.columns else i + 1,
                    'angle': angle_deg,
                    'lon': self.df.loc[i + 1, 'lon'],
                    'lat': self.df.loc[i + 1, 'lat'],
                    'desc': f'线段与主方向夹角{angle_deg:.1f}°，超过{reversal_threshold}°阈值'
                })
        
        return anomalies

    def check_adjacent_segment_angle(self, sharp_turn_threshold=90):
        """
        规则4: 相邻线段角度突变检测
        检查相邻线段之间的夹角，大于sharp_turn_threshold度视为折返
        
        sharp_turn_threshold: 折返判定阈值（度），默认90度
        """
        anomalies = []
        if self.n < 3:
            return anomalies
        
        prev_seg_unit = None
        reversal_angle_rad = math.radians(sharp_turn_threshold)
        
        for i in range(self.n - 1):
            p1 = self.local_coords[i]
            p2 = self.local_coords[i + 1]
            
            dx = p2[0] - p1[0]
            dy = p2[1] - p1[1]
            seg_length = math.sqrt(dx**2 + dy**2)
            
            if seg_length < 0.5:
                continue
            
            seg_unit = (dx / seg_length, dy / seg_length)
            
            if prev_seg_unit is not None:
                dot = seg_unit[0] * prev_seg_unit[0] + seg_unit[1] * prev_seg_unit[1]
                
                dot_clamped = max(-1.0, min(1.0, dot))
                angle_rad = math.acos(dot_clamped)
                angle_deg = math.degrees(angle_rad)
                
                if angle_deg > sharp_turn_threshold:
                    anomalies.append({
                        'type': 'sharp_turn',
                        'point_index': i + 1,
                        'point_seq': self.df.loc[i + 1, 'seq'] if 'seq' in self.df.columns else i + 1,
                        'angle': angle_deg,
                        'lon': self.df.loc[i + 1, 'lon'],
                        'lat': self.df.loc[i + 1, 'lat'],
                        'desc': f'相邻线段夹角{angle_deg:.1f}°，超过{sharp_turn_threshold}°阈值'
                    })
            
            prev_seg_unit = seg_unit
        
        return anomalies



def cross_product(o: Tuple[float, float], a: Tuple[float, float], b: Tuple[float, float]) -> float:
    return (a[0] - o[0]) * (b[1] - o[1]) - (a[1] - o[1]) * (b[0] - o[0])


def segments_intersect(p1: Tuple[float, float], p2: Tuple[float, float], 
                       p3: Tuple[float, float], p4: Tuple[float, float]) -> bool:
    d1 = cross_product(p3, p4, p1)
    d2 = cross_product(p3, p4, p2)
    d3 = cross_product(p1, p2, p3)
    d4 = cross_product(p1, p2, p4)
    return ((d1 > 0 and d2 < 0) or (d1 < 0 and d2 > 0)) and \
           ((d3 > 0 and d4 < 0) or (d3 < 0 and d4 > 0))


def detect_anomaly(sec: Dict, reversal_threshold: float = 90.0) -> Tuple[bool, List[Dict]]:
    """
    检测断面测量点连线的异常情况
    
    检测内容：
    1. 方向反转：线段与主方向夹角超过阈值（默认90度）
    2. 相邻线段折返：相邻线段夹角超过阈值（默认90度）
    3. 线段交叉：非相邻线段相交
    
    reversal_threshold: 折返判定阈值（度），大于此角度视为折返
    """
    anomalies = []
    points = [p for p in sec["points"] if p.get("lon") is not None and p.get("lat") is not None]
    if len(points) < 3:
        return False, anomalies
    
    points.sort(key=lambda p: p.get("seq", 0) if p.get("seq") is not None else 0)
    
    ref_lon = points[0]["lon"]
    ref_lat = points[0]["lat"]
    
    local_coords = []
    for p in points:
        x, y = lonlat_to_local_xy(p["lon"], p["lat"], ref_lon, ref_lat)
        local_coords.append((x, y))
    
    start = local_coords[0]
    end = local_coords[-1]
    dx_main = end[0] - start[0]
    dy_main = end[1] - start[1]
    main_length = math.sqrt(dx_main**2 + dy_main**2)
    
    if main_length < 1:
        coords = [(p["lon"], p["lat"]) for p in points]
        for i in range(len(coords) - 1):
            for j in range(i + 2, len(coords) - 1):
                if segments_intersect(coords[i], coords[i + 1], coords[j], coords[j + 1]):
                    anomalies.append({
                        "type": "cross",
                        "desc": ANOMALY_TYPES["cross"],
                        "segment1_start_seq": points[i]["seq"],
                        "segment1_end_seq": points[i + 1]["seq"],
                        "segment2_start_seq": points[j]["seq"],
                        "segment2_end_seq": points[j + 1]["seq"],
                        "segment1_coords": [coords[i], coords[i + 1]],
                        "segment2_coords": [coords[j], coords[j + 1]],
                    })
        return len(anomalies) > 0, anomalies
    
    main_unit = (dx_main / main_length, dy_main / main_length)
    
    for i in range(len(local_coords) - 1):
        p1 = local_coords[i]
        p2 = local_coords[i + 1]
        
        dx = p2[0] - p1[0]
        dy = p2[1] - p1[1]
        seg_length = math.sqrt(dx**2 + dy**2)
        
        if seg_length < 0.5:
            continue
        
        seg_unit = (dx / seg_length, dy / seg_length)
        
        dot = seg_unit[0] * main_unit[0] + seg_unit[1] * main_unit[1]
        dot_clamped = max(-1.0, min(1.0, dot))
        angle_deg = math.degrees(math.acos(dot_clamped))
        
        if angle_deg > reversal_threshold:
            anomalies.append({
                "type": "reversal",
                "desc": ANOMALY_TYPES["reversal"],
                "point_seq": points[i + 1]["seq"],
                "lon": points[i + 1]["lon"],
                "lat": points[i + 1]["lat"],
                "prev_lon": points[i]["lon"],
                "prev_lat": points[i]["lat"],
                "angle": angle_deg,
                "detail": f"线段与主方向夹角{angle_deg:.1f}°，超过{reversal_threshold}°阈值"
            })
    
    prev_seg_unit = None
    for i in range(len(local_coords) - 1):
        p1 = local_coords[i]
        p2 = local_coords[i + 1]
        
        dx = p2[0] - p1[0]
        dy = p2[1] - p1[1]
        seg_length = math.sqrt(dx**2 + dy**2)
        
        if seg_length < 0.5:
            continue
        
        seg_unit = (dx / seg_length, dy / seg_length)
        
        if prev_seg_unit is not None:
            dot = seg_unit[0] * prev_seg_unit[0] + seg_unit[1] * prev_seg_unit[1]
            dot_clamped = max(-1.0, min(1.0, dot))
            angle_deg = math.degrees(math.acos(dot_clamped))
            
            if angle_deg > reversal_threshold:
                anomalies.append({
                    "type": "sharp_turn",
                    "desc": "相邻线段折返",
                    "point_seq": points[i + 1]["seq"],
                    "lon": points[i + 1]["lon"],
                    "lat": points[i + 1]["lat"],
                    "prev_lon": points[i]["lon"],
                    "prev_lat": points[i]["lat"],
                    "angle": angle_deg,
                    "detail": f"相邻线段夹角{angle_deg:.1f}°，超过{reversal_threshold}°阈值"
                })
        
        prev_seg_unit = seg_unit
    
    coords = [(p["lon"], p["lat"]) for p in points]
    for i in range(len(coords) - 1):
        for j in range(i + 2, len(coords) - 1):
            if segments_intersect(coords[i], coords[i + 1], coords[j], coords[j + 1]):
                anomalies.append({
                    "type": "cross",
                    "desc": ANOMALY_TYPES["cross"],
                    "segment1_start_seq": points[i]["seq"],
                    "segment1_end_seq": points[i + 1]["seq"],
                    "segment2_start_seq": points[j]["seq"],
                    "segment2_end_seq": points[j + 1]["seq"],
                    "segment1_coords": [coords[i], coords[i + 1]],
                    "segment2_coords": [coords[j], coords[j + 1]],
                })
    
    return len(anomalies) > 0, anomalies


def validate_section_data(sec: Dict) -> List[Dict]:
    issues = []
    points = sec.get("points", [])
    table_type = sec.get("table_type", "")
    is_chengtu = table_type == "chengtu"
    if len(points) == 0:
        return issues
    
    seq_values = []
    seq_missing_indices = []
    for i, p in enumerate(points):
        seq = p.get("seq")
        if seq is None:
            seq_missing_indices.append(i)
        else:
            try:
                seq_int = int(seq)
                seq_values.append((seq_int, i))
            except (ValueError, TypeError):
                seq_missing_indices.append(i)
    
    if seq_missing_indices:
        issues.append({
            "rule": "seq_missing",
            "level": VALIDATION_RULES["seq_missing"]["level"],
            "desc": VALIDATION_RULES["seq_missing"]["desc"],
            "details": f"第 {', '.join(str(i+1) for i in seq_missing_indices)} 个点序号缺失",
            "point_index": seq_missing_indices,
        })
    
    if seq_values:
        seq_counts = {}
        for seq, idx in seq_values:
            if seq in seq_counts:
                seq_counts[seq].append(idx)
            else:
                seq_counts[seq] = [idx]
        
        for seq, indices in seq_counts.items():
            if len(indices) > 1:
                issues.append({
                    "rule": "seq_duplicate",
                    "level": VALIDATION_RULES["seq_duplicate"]["level"],
                    "desc": VALIDATION_RULES["seq_duplicate"]["desc"],
                    "details": f"序号 {seq} 出现 {len(indices)} 次",
                    "point_index": indices,
                })
        
        sorted_seq = sorted(seq_values, key=lambda x: x[0])
        not_monotonic_indices = []
        for i in range(1, len(sorted_seq)):
            if sorted_seq[i][0] <= sorted_seq[i-1][0]:
                not_monotonic_indices.append(sorted_seq[i][1])
        
        if not_monotonic_indices:
            issues.append({
                "rule": "seq_not_monotonic",
                "level": VALIDATION_RULES["seq_not_monotonic"]["level"],
                "desc": VALIDATION_RULES["seq_not_monotonic"]["desc"],
                "details": f"序号非递增，涉及第 {', '.join(str(i+1) for i in not_monotonic_indices)} 个点",
                "point_index": not_monotonic_indices,
            })
        
        gap_indices = []
        for i in range(1, len(sorted_seq)):
            gap = sorted_seq[i][0] - sorted_seq[i-1][0]
            if gap > 1:
                gap_indices.append(sorted_seq[i][1])
        
        if gap_indices:
            issues.append({
                "rule": "seq_gap",
                "level": VALIDATION_RULES["seq_gap"]["level"],
                "desc": VALIDATION_RULES["seq_gap"]["desc"],
                "details": f"序号跳号，涉及第 {', '.join(str(i+1) for i in gap_indices)} 个点",
                "point_index": gap_indices,
            })
    
    dist_values = [(p.get("distance"), i) for i, p in enumerate(points)]
    negative_indices = [idx for dist, idx in dist_values if dist is not None and dist < 0]
    if negative_indices:
        issues.append({
            "rule": "dist_negative",
            "level": VALIDATION_RULES["dist_negative"]["level"],
            "desc": VALIDATION_RULES["dist_negative"]["desc"],
            "details": f"起点距为负，涉及第 {', '.join(str(i+1) for i in negative_indices)} 个点",
            "point_index": negative_indices,
        })
    
    valid_dist = [(dist, idx) for dist, idx in dist_values if dist is not None]
    if valid_dist:
        sorted_dist = sorted(valid_dist, key=lambda x: x[0])
        not_monotonic_indices = []
        for i in range(1, len(sorted_dist)):
            if sorted_dist[i][0] < sorted_dist[i-1][0]:
                not_monotonic_indices.append(sorted_dist[i][1])
        
        if not_monotonic_indices:
            issues.append({
                "rule": "dist_not_monotonic",
                "level": VALIDATION_RULES["dist_not_monotonic"]["level"],
                "desc": VALIDATION_RULES["dist_not_monotonic"]["desc"],
                "details": f"起点距回退，涉及第 {', '.join(str(i+1) for i in not_monotonic_indices)} 个点",
                "point_index": not_monotonic_indices,
            })
        
        dist_counts = {}
        for dist, idx in valid_dist:
            key = round(dist, 3)
            if key in dist_counts:
                dist_counts[key].append(idx)
            else:
                dist_counts[key] = [idx]
        
        for dist_key, indices in dist_counts.items():
            if len(indices) > 1:
                issues.append({
                    "rule": "dist_duplicate",
                    "level": VALIDATION_RULES["dist_duplicate"]["level"],
                    "desc": VALIDATION_RULES["dist_duplicate"]["desc"],
                    "details": f"起点距 {dist_key:.3f}m 重复出现 {len(indices)} 次",
                    "point_index": indices,
                })
        
        if len(valid_dist) > 1:
            dist_diffs = [abs(valid_dist[i+1][0] - valid_dist[i][0]) for i in range(len(valid_dist)-1)]
            avg_diff = sum(dist_diffs) / len(dist_diffs) if dist_diffs else 0
            jump_indices = []
            for i, diff in enumerate(dist_diffs):
                if avg_diff > 0 and diff > avg_diff * VALIDATION_CONFIG["dist_jump_factor"]:
                    jump_indices.append(valid_dist[i+1][1])
            
            if jump_indices:
                issues.append({
                    "rule": "dist_jump",
                    "level": VALIDATION_RULES["dist_jump"]["level"],
                    "desc": VALIDATION_RULES["dist_jump"]["desc"],
                    "details": f"起点距跳变(平均间距{avg_diff:.1f}m)，涉及第 {', '.join(str(i+1) for i in jump_indices)} 个点",
                    "point_index": jump_indices,
                })
    
    lon_values = [(p.get("lon"), i) for i, p in enumerate(points)]
    lat_values = [(p.get("lat"), i) for i, p in enumerate(points)]
    
    lon_range = VALIDATION_CONFIG["lon_range"]
    lat_range = VALIDATION_CONFIG["lat_range"]
    
    out_of_range_indices = []
    for lon, idx in lon_values:
        if lon is not None and (lon < lon_range[0] or lon > lon_range[1]):
            out_of_range_indices.append(idx)
    if out_of_range_indices:
        issues.append({
            "rule": "lon_out_of_range",
            "level": VALIDATION_RULES["lon_out_of_range"]["level"],
            "desc": VALIDATION_RULES["lon_out_of_range"]["desc"],
            "details": f"经度超出{lon_range[0]}~{lon_range[1]}范围，涉及第 {', '.join(str(i+1) for i in out_of_range_indices)} 个点",
            "point_index": out_of_range_indices,
        })
    
    out_of_range_indices = []
    for lat, idx in lat_values:
        if lat is not None and (lat < lat_range[0] or lat > lat_range[1]):
            out_of_range_indices.append(idx)
    if out_of_range_indices:
        issues.append({
            "rule": "lat_out_of_range",
            "level": VALIDATION_RULES["lat_out_of_range"]["level"],
            "desc": VALIDATION_RULES["lat_out_of_range"]["desc"],
            "details": f"纬度超出{lat_range[0]}~{lat_range[1]}范围，涉及第 {', '.join(str(i+1) for i in out_of_range_indices)} 个点",
            "point_index": out_of_range_indices,
        })
    
    zero_lon_indices = [idx for lon, idx in lon_values if lon is None or lon == 0]
    if zero_lon_indices and not is_chengtu:
        issues.append({
            "rule": "lon_zero",
            "level": VALIDATION_RULES["lon_zero"]["level"],
            "desc": VALIDATION_RULES["lon_zero"]["desc"],
            "details": f"经度为零或缺失，涉及第 {', '.join(str(i+1) for i in zero_lon_indices)} 个点",
            "point_index": zero_lon_indices,
        })
    
    zero_lat_indices = [idx for lat, idx in lat_values if lat is None or lat == 0]
    if zero_lat_indices and not is_chengtu:
        issues.append({
            "rule": "lat_zero",
            "level": VALIDATION_RULES["lat_zero"]["level"],
            "desc": VALIDATION_RULES["lat_zero"]["desc"],
            "details": f"纬度为零或缺失，涉及第 {', '.join(str(i+1) for i in zero_lat_indices)} 个点",
            "point_index": zero_lat_indices,
        })
    
    valid_lon = [(lon, idx) for lon, idx in lon_values if lon is not None and lon != 0]
    valid_lat = [(lat, idx) for lat, idx in lat_values if lat is not None and lat != 0]
    
    if len(valid_lon) > 1:
        lon_data = [lon for lon, idx in valid_lon]
        lon_span = max(lon_data) - min(lon_data)
        threshold = max(VALIDATION_CONFIG["coord_jump_threshold"], lon_span * VALIDATION_CONFIG["coord_jump_ratio"])
        jump_indices = []
        for i in range(1, len(valid_lon)):
            diff = abs(valid_lon[i][0] - valid_lon[i-1][0])
            if diff > threshold:
                jump_indices.append(valid_lon[i][1])
        
        if jump_indices:
            issues.append({
                "rule": "lon_jump",
                "level": VALIDATION_RULES["lon_jump"]["level"],
                "desc": VALIDATION_RULES["lon_jump"]["desc"],
                "details": f"经度跳变(跨度{lon_span:.4f}°)，涉及第 {', '.join(str(i+1) for i in jump_indices)} 个点",
                "point_index": jump_indices,
            })
    
    if len(valid_lat) > 1:
        lat_data = [lat for lat, idx in valid_lat]
        lat_span = max(lat_data) - min(lat_data)
        threshold = max(VALIDATION_CONFIG["coord_jump_threshold"], lat_span * VALIDATION_CONFIG["coord_jump_ratio"])
        jump_indices = []
        for i in range(1, len(valid_lat)):
            diff = abs(valid_lat[i][0] - valid_lat[i-1][0])
            if diff > threshold:
                jump_indices.append(valid_lat[i][1])
        
        if jump_indices:
            issues.append({
                "rule": "lat_jump",
                "level": VALIDATION_RULES["lat_jump"]["level"],
                "desc": VALIDATION_RULES["lat_jump"]["desc"],
                "details": f"纬度跳变(跨度{lat_span:.4f}°)，涉及第 {', '.join(str(i+1) for i in jump_indices)} 个点",
                "point_index": jump_indices,
            })
    
    swapped_indices = []
    for i, p in enumerate(points):
        lon = p.get("lon")
        lat = p.get("lat")
        if lon is not None and lat is not None and lon != 0 and lat != 0:
            if lat_range[0] <= lon <= lat_range[1] and lon_range[0] <= lat <= lon_range[1]:
                swapped_indices.append(i)
    
    if swapped_indices:
        issues.append({
            "rule": "lonlat_swapped",
            "level": VALIDATION_RULES["lonlat_swapped"]["level"],
            "desc": VALIDATION_RULES["lonlat_swapped"]["desc"],
            "details": f"经纬度可能互换，涉及第 {', '.join(str(i+1) for i in swapped_indices)} 个点",
            "point_index": swapped_indices,
        })
    
    coord_missing_indices = []
    for i, p in enumerate(points):
        if p.get("distance") is not None and p.get("elevation") is not None:
            lon = p.get("lon")
            lat = p.get("lat")
            if lon is None or lat is None or lon == 0 or lat == 0:
                coord_missing_indices.append(i)
    
    if coord_missing_indices and not is_chengtu:
        issues.append({
            "rule": "coord_missing",
            "level": VALIDATION_RULES["coord_missing"]["level"],
            "desc": VALIDATION_RULES["coord_missing"]["desc"],
            "details": f"有起点距高程但缺经纬度，涉及第 {', '.join(str(i+1) for i in coord_missing_indices)} 个点",
            "point_index": coord_missing_indices,
        })
    
    valid_coords = [(p.get("lon"), p.get("lat")) for p in points if p.get("lon") and p.get("lat") and p.get("lon") != 0 and p.get("lat") != 0]
    if len(valid_coords) >= 2:
        first_coord = valid_coords[0]
        all_same = all(c == first_coord for c in valid_coords)
        if all_same:
            issues.append({
                "rule": "all_same_coord",
                "level": VALIDATION_RULES["all_same_coord"]["level"],
                "desc": VALIDATION_RULES["all_same_coord"]["desc"],
                "details": f"所有{len(valid_coords)}个点的经纬度完全相同: ({first_coord[0]:.6f}, {first_coord[1]:.6f})",
                "point_index": list(range(len(points))),
            })
    
    if len(valid_coords) >= 3:
        valid_points_data = []
        valid_points_indices = []
        for i, p in enumerate(points):
            lon = p.get("lon")
            lat = p.get("lat")
            dist = p.get("distance")
            seq = p.get("seq")
            if lon and lat and lon != 0 and lat != 0 and dist is not None:
                valid_points_data.append({"distance": dist, "lon": lon, "lat": lat, "seq": seq if seq is not None else i + 1})
                valid_points_indices.append(i)
        
        if len(valid_points_data) >= 3:
            df = pd.DataFrame(valid_points_data)
            checker = SpatialContinuityChecker(df)
            
            flagged_indices = checker.check_linear_fit_deviation(
                threshold=VALIDATION_CONFIG["linear_fit_threshold"]
            )
            
            if flagged_indices:
                deviation_indices = [valid_points_indices[i] for i in flagged_indices if i < len(valid_points_indices)]
                if deviation_indices:
                    L = np.max(df['distance']) - np.min(df['distance'])
                    issues.append({
                        "rule": "linear_fit_deviation",
                        "level": VALIDATION_RULES["linear_fit_deviation"]["level"],
                        "desc": VALIDATION_RULES["linear_fit_deviation"]["desc"],
                        "details": f"偏离PCA拟合直线超过断面长度{L:.1f}m的{VALIDATION_CONFIG['linear_fit_threshold']*100}%，涉及第 {', '.join(str(i+1) for i in deviation_indices)} 个点",
                        "point_index": deviation_indices,
                    })
            
            direction_anomalies = checker.check_direction_consistency(reversal_threshold=90)
            for anomaly in direction_anomalies:
                idx = anomaly.get('point_index', 0)
                if idx < len(valid_points_indices):
                    point_idx = valid_points_indices[idx]
                    issues.append({
                        "rule": "direction_reversal",
                        "level": VALIDATION_RULES["direction_reversal"]["level"],
                        "desc": VALIDATION_RULES["direction_reversal"]["desc"],
                        "details": f"第{point_idx+1}个点连线方向折返，{anomaly.get('desc', '')}",
                        "point_index": [point_idx],
                    })
            
            adjacent_anomalies = checker.check_adjacent_segment_angle(sharp_turn_threshold=90)
            for anomaly in adjacent_anomalies:
                idx = anomaly.get('point_index', 0)
                if idx < len(valid_points_indices):
                    point_idx = valid_points_indices[idx]
                    issues.append({
                        "rule": "adjacent_reversal",
                        "level": VALIDATION_RULES["adjacent_reversal"]["level"],
                        "desc": VALIDATION_RULES["adjacent_reversal"]["desc"],
                        "details": f"第{point_idx+1}个点相邻线段折返，{anomaly.get('desc', '')}",
                        "point_index": [point_idx],
                    })
    
    return issues



def find_excel_files() -> List[Tuple[str, str]]:
    files = []
    for river in RIVER_DIRS:
        river_path = os.path.join(BASE_DIR, river)
        if not os.path.exists(river_path):
            print(f"  目录不存在: {river_path}")
            continue
        xlsx_files = glob.glob(os.path.join(river_path, "*.xlsx"))
        for f in xlsx_files:
            if "横断面" in f:
                files.append((river, f))
    return files


def safe_float(val: Any) -> Optional[float]:
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def read_sections_from_file(file_path: str, river_name: str) -> Dict:
    sections = {}
    wb = None
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        ws = wb["横断面"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            sec_name = row[COL_SECTION]
            if not sec_name:
                continue
            district = row[COL_DISTRICT] or "未知区域"
            full_key = f"{river_name}|{sec_name}"
            if full_key not in sections:
                sections[full_key] = {
                    "name": sec_name,
                    "river": river_name,
                    "district": district,
                    "hmz": safe_float(row[COL_HMZ]),
                    "czz": safe_float(row[COL_CZZ]),
                    "backwater": safe_float(row[COL_BACKWATER]),
                    "points": [],
                }
            
            dist = safe_float(row[COL_DISTANCE])
            elev = safe_float(row[COL_ELEVATION])
            if dist is None or elev is None:
                continue
            
            lon = safe_float(row[COL_LON])
            lat = safe_float(row[COL_LAT])
            seq = row[COL_POINT_SEQ]
            
            feature_desc = row[COL_FEATURE]
            if feature_desc is None:
                feature_desc = ""
            elif not isinstance(feature_desc, str):
                feature_desc = str(feature_desc)
            is_feature = any(kw in feature_desc for kw in FEATURE_KEYWORDS) if feature_desc else False
            
            sections[full_key]["points"].append({
                "seq": seq if seq is not None else len(sections[full_key]["points"]) + 1,
                "distance": round(dist, 3),
                "elevation": round(elev, 3),
                "lon": lon,
                "lat": lat,
                "feature": feature_desc,
                "isFeature": is_feature,
            })
    except Exception as e:
        print(f"  读取失败: {file_path}, 错误: {e}")
    finally:
        if wb:
            wb.close()
    return sections


def build_tree_data(sections: Dict) -> Dict:
    tree = {}
    for key, sec in sections.items():
        river = sec["river"]
        district = sec["district"]
        if river not in tree:
            tree[river] = {}
        if district not in tree[river]:
            tree[river][district] = []
        tree[river][district].append({
            "key": key,
            "name": sec["name"],
            "points": len(sec["points"]),
            "features": sum(1 for p in sec["points"] if p["isFeature"]),
            "anomaly": sec.get("anomaly", False),
            "validation_error": sec.get("validation_error", False),
            "validation_warning": sec.get("validation_warning", False),
            "validation_issues_count": len(sec.get("validation_issues", [])),
        })
    return tree


def write_anomaly_report(all_sections: Dict, output_path: str):
    anomaly_data = []
    for key, sec in all_sections.items():
        if sec.get("anomaly"):
            anomaly_info = {
                "断面名称": sec["name"],
                "流域": sec["river"],
                "区域": sec["district"],
                "测量点数": len(sec["points"]),
                "异常类型": [],
                "异常详情": sec.get("anomaly_details", []),
            }
            for detail in sec.get("anomaly_details", []):
                anomaly_info["异常类型"].append(detail.get("desc", detail.get("type")))
            anomaly_data.append(anomaly_info)
    
    validation_data = []
    for key, sec in all_sections.items():
        issues = sec.get("validation_issues", [])
        if issues:
            validation_info = {
                "断面名称": sec["name"],
                "流域": sec["river"],
                "区域": sec["district"],
                "测量点数": len(sec["points"]),
                "错误数": sum(1 for i in issues if i["level"] == "error"),
                "警告数": sum(1 for i in issues if i["level"] == "warning"),
                "问题详情": issues,
            }
            validation_data.append(validation_info)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "异常断面汇总"
    
    headers = ["断面名称", "流域", "区域", "测量点数", "异常类型", "异常详情(JSON)"]
    ws.append(headers)
    
    for item in anomaly_data:
        ws.append([
            item["断面名称"],
            item["流域"],
            item["区域"],
            item["测量点数"],
            ", ".join(item["异常类型"]),
            json.dumps(item["异常详情"], ensure_ascii=False),
        ])
    
    ws2 = wb.create_sheet("异常点详情")
    detail_headers = ["断面名称", "流域", "异常类型", "点序号", "经度", "纬度", "详情"]
    ws2.append(detail_headers)
    
    for item in anomaly_data:
        for detail in item["异常详情"]:
            row_data = [
                item["断面名称"],
                item["流域"],
                detail.get("desc", detail.get("type")),
            ]
            if detail.get("type") == "reversal":
                row_data.extend([
                    detail.get("point_seq"),
                    detail.get("lon"),
                    detail.get("lat"),
                    f"前一坐标({detail.get('prev_lon')}, {detail.get('prev_lat')})",
                ])
            elif detail.get("type") == "cross":
                row_data.extend([
                    f"{detail.get('segment1_start_seq')}-{detail.get('segment1_end_seq')} 与 {detail.get('segment2_start_seq')}-{detail.get('segment2_end_seq')}",
                    f"线段1: {detail.get('segment1_coords')}",
                    f"线段2: {detail.get('segment2_coords')}",
                    "线段交叉",
                ])
            else:
                row_data.extend(["-", "-", "-", json.dumps(detail, ensure_ascii=False)])
            ws2.append(row_data)
    
    ws3 = wb.create_sheet("数据校验详情")
    validation_headers = ["断面名称", "流域", "区域", "测量点数", "错误数", "警告数", "问题类型", "问题详情"]
    ws3.append(validation_headers)
    
    for item in validation_data:
        for issue in item["问题详情"]:
            ws3.append([
                item["断面名称"],
                item["流域"],
                item["区域"],
                item["测量点数"],
                item["错误数"],
                item["警告数"],
                f"{issue['desc']}({issue['level']})",
                issue["details"],
            ])
    
    for col in range(1, 7):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20
    for col in range(1, 8):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 25
    for col in range(1, 9):
        ws3.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 25
    
    wb.save(output_path)
    wb.close()
    print(f"  异常断面报告已保存: {output_path} ({len(anomaly_data)}个异常断面, {len(validation_data)}个校验问题断面)")


HTML_TEMPLATE = """<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>__TITLE__</title>
<script src="https://cdn.jsdelivr.net/npm/echarts@5.5.0/dist/echarts.min.js"></script>
<style>
  * { margin: 0; padding: 0; box-sizing: border-box; }
  body { font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', 'PingFang SC', 'Microsoft YaHei', sans-serif; background: #0f172a; color: #e2e8f0; }

  .sidebar {
    position: fixed; left: 0; top: 0; width: 320px; height: 100vh;
    background: #1e293b; border-right: 1px solid #334155;
    display: flex; flex-direction: column; overflow: hidden; z-index: 100;
  }
  .sidebar-header {
    padding: 16px 20px; border-bottom: 1px solid #334155;
    background: linear-gradient(135deg, #1e40af, #3b82f6);
  }
  .sidebar-header h1 { font-size: 16px; font-weight: 600; color: #fff; margin-bottom: 4px; }
  .sidebar-header p { font-size: 12px; color: rgba(255,255,255,0.7); }
  .sidebar-search { padding: 12px 16px; border-bottom: 1px solid #334155; }
  .sidebar-search input {
    width: 100%; padding: 8px 12px; border-radius: 8px; border: 1px solid #475569;
    background: #0f172a; color: #e2e8f0; font-size: 13px; outline: none;
  }
  .sidebar-search input:focus { border-color: #3b82f6; box-shadow: 0 0 0 2px rgba(59,130,246,0.2); }
  .sidebar-stats {
    padding: 10px 16px; border-bottom: 1px solid #334155; display: flex; gap: 12px; font-size: 12px; color: #94a3b8;
  }
  .sidebar-stats span { color: #3b82f6; font-weight: 600; }
  .sidebar-pagination {
    padding: 8px 16px; border-bottom: 1px solid #334155; display: flex; align-items: center; justify-content: space-between; gap: 8px; font-size: 12px; flex-wrap: wrap;
  }
  .sidebar-pagination .page-nav { display: flex; align-items: center; gap: 8px; }
  .sidebar-pagination .page-size { display: flex; align-items: center; gap: 4px; }
  .sidebar-pagination .page-size select {
    padding: 4px 8px; border-radius: 4px; border: 1px solid #475569; background: #334155; color: #e2e8f0; cursor: pointer; font-size: 11px;
  }
  .sidebar-pagination button {
    padding: 4px 10px; border-radius: 4px; border: 1px solid #475569; background: #334155; color: #e2e8f0; cursor: pointer;
  }
  .sidebar-pagination button:hover { background: #475569; }
  .sidebar-pagination button:disabled { opacity: 0.5; cursor: not-allowed; }
  .sidebar-pagination .page-info { color: #94a3b8; }
  .sidebar-pagination .expand-btn {
    padding: 4px 8px; border-radius: 4px; border: 1px solid #475569; background: #334155; color: #e2e8f0; cursor: pointer; font-size: 11px;
  }
  .sidebar-pagination .expand-btn:hover { background: #475569; }
  .sidebar-pagination .expand-btn.btn-primary {
    background: linear-gradient(135deg, #2563eb, #3b82f6); border-color: transparent; color: #fff;
  }
  .sidebar-pagination .expand-btn.btn-primary:hover { background: linear-gradient(135deg, #1d4ed8, #2563eb); }
  .section-list { flex: 1; overflow-y: auto; padding: 8px; }
  .section-list::-webkit-scrollbar { width: 4px; }
  .section-list::-webkit-scrollbar-thumb { background: #475569; border-radius: 2px; }

  .tree-node { margin-bottom: 2px; }
  .tree-header {
    padding: 8px 12px; border-radius: 6px; cursor: pointer;
    display: flex; align-items: center; gap: 8px; font-size: 13px;
    color: #cbd5e1; transition: all 0.15s;
  }
  .tree-header:hover { background: #334155; }
  .tree-header.river { font-weight: 600; color: #60a5fa; }
  .tree-header.district { font-size: 12px; color: #94a3b8; padding-left: 24px; }
  .tree-header .arrow {
    width: 16px; height: 16px; display: flex; align-items: center; justify-content: center;
    transition: transform 0.2s;
  }
  .tree-header .arrow.expanded { transform: rotate(90deg); }
  .tree-header .count {
    font-size: 11px; padding: 1px 6px; border-radius: 10px;
    background: rgba(59,130,246,0.15); color: #60a5fa;
  }
  .tree-children { padding-left: 12px; }
  .tree-children.hidden { display: none; }

  .section-item {
    padding: 8px 14px 8px 14px; border-radius: 6px; cursor: pointer; margin-bottom: 1px;
    font-size: 12px; color: #cbd5e1; transition: all 0.15s; display: flex;
    align-items: center; justify-content: space-between;
  }
  .section-item:hover { background: #334155; }
  .section-item.active { background: rgba(59,130,246,0.15); color: #60a5fa; }
  .section-item .checkbox {
    width: 16px; height: 16px; border: 2px solid #475569; border-radius: 3px;
    margin-right: 8px; display: flex; align-items: center; justify-content: center;
    transition: all 0.15s;
  }
  .section-item.active .checkbox { border-color: #3b82f6; background: #3b82f6; }
  .section-item.active .checkbox::after { content: '✓'; color: #fff; font-size: 10px; }
  .section-item .badge {
    font-size: 10px; padding: 1px 5px; border-radius: 8px;
    background: rgba(59,130,246,0.15); color: #60a5fa;
  }
  .section-item .badge.has-feature { background: rgba(239,68,68,0.15); color: #f87171; }
  .section-item .badge.has-anomaly { background: rgba(250,204,21,0.15); color: #facc15; }
  .section-item .badge.has-validation-error { background: rgba(220,38,38,0.15); color: #ef4444; }
  .section-item .badge.has-validation-warning { background: rgba(234,179,8,0.15); color: #eab308; }

  .sidebar-filter {
    padding: 8px 16px; border-bottom: 1px solid #334155; display: flex; gap: 6px; flex-wrap: wrap;
  }
  .filter-btn {
    padding: 4px 10px; border-radius: 4px; font-size: 11px; cursor: pointer;
    border: 1px solid #475569; background: #334155; color: #e2e8f0; transition: all 0.15s;
  }
  .filter-btn:hover { background: #475569; }
  .filter-btn.active { background: rgba(250,204,21,0.15); border-color: #facc15; color: #facc15; }
  .filter-btn .filter-count { margin-left: 4px; font-weight: 600; }

  .main { margin-left: 320px; min-height: 100vh; padding: 20px; }
  .toolbar {
    position: sticky; top: 0; padding: 12px 24px; border-radius: 12px;
    background: #1e293b; border: 1px solid #334155; margin-bottom: 20px;
    display: flex; align-items: center; gap: 16px; z-index: 50;
  }
  .toolbar-title { font-size: 15px; font-weight: 600; color: #f1f5f9; }
  .toolbar-info { font-size: 12px; color: #94a3b8; }
  .toolbar-info span { color: #60a5fa; font-weight: 500; }
  .toolbar-actions { margin-left: auto; display: flex; gap: 8px; }
  .btn {
    padding: 6px 14px; border-radius: 6px; font-size: 12px; cursor: pointer;
    border: 1px solid #475569; background: #334155; color: #e2e8f0; transition: all 0.15s;
  }
  .btn:hover { background: #475569; }
  .btn.active { background: rgba(59,130,246,0.2); border-color: #3b82f6; color: #60a5fa; }
  .btn-primary { background: linear-gradient(135deg, #2563eb, #3b82f6); border-color: transparent; color: #fff; }
  .btn-primary:hover { background: linear-gradient(135deg, #1d4ed8, #2563eb); }

  .chart-area { display: grid; gap: 20px; }
  .chart-area.cols-1 { grid-template-columns: 1fr; }
  .chart-area.cols-2 { grid-template-columns: repeat(2, 1fr); }
  .chart-area.cols-3 { grid-template-columns: repeat(3, 1fr); }
  .chart-area.cols-4 { grid-template-columns: repeat(4, 1fr); }

  .chart-card {
    background: #1e293b; border: 1px solid #334155; border-radius: 12px;
    overflow: hidden; display: flex; flex-direction: column;
  }
  .chart-card.anomaly-card { border-color: #facc15; box-shadow: 0 0 0 1px rgba(250,204,21,0.2); }
  .chart-card.validation-error-card { border-color: #ef4444; box-shadow: 0 0 0 1px rgba(220,38,38,0.2); }
  .chart-card-header {
    padding: 12px 16px; border-bottom: 1px solid #334155;
    display: flex; align-items: center; justify-content: space-between; flex-wrap: wrap; gap: 8px;
  }
  .chart-card-title { font-size: 14px; font-weight: 600; color: #f1f5f9; }
  .chart-card-subtitle { font-size: 11px; color: #64748b; }
  .chart-card-meta { display: flex; gap: 8px; font-size: 11px; color: #94a3b8; flex-wrap: wrap; }
  .chart-card-meta .tag { padding: 2px 8px; border-radius: 4px; }
  .tag-hmz { background: rgba(249,115,22,0.15); color: #fb923c; }
  .tag-czz { background: rgba(220,38,38,0.15); color: #f87171; }
  .tag-backwater { background: rgba(59,130,246,0.15); color: #60a5fa; }

  .chart-card-body { display: flex; flex-direction: column; gap: 4px; padding: 4px; height: 700px; min-height: 500px; }
  .chart-section { flex: 1; min-height: 0; }
  .chart-section-title {
    padding: 6px 10px; font-size: 11px; color: #94a3b8; background: #0f172a;
    border-radius: 4px 4px 0 0; text-align: center;
  }
  .chart-section-chart { height: calc(100% - 24px); }

  .water-level-legend {
    display: flex; gap: 12px; padding: 8px 16px; border-top: 1px solid #334155;
    font-size: 11px; color: #94a3b8; flex-wrap: wrap;
  }
  .water-level-legend .item { display: flex; align-items: center; gap: 4px; }
  .water-level-legend .line { width: 20px; height: 2px; border-top: 2px dashed; }

  .empty-state {
    display: flex; align-items: center; justify-content: center;
    height: 300px; color: #64748b; font-size: 14px;
    background: #1e293b; border-radius: 12px; border: 1px solid #334155;
  }

  .water-toggle { display: flex; gap: 6px; }
  .water-toggle .toggle-btn {
    padding: 3px 8px; border-radius: 4px; font-size: 11px; cursor: pointer;
    border: 1px solid transparent; transition: all 0.15s;
  }
  .toggle-btn.hmz { color: #fb923c; }
  .toggle-btn.hmz.on { background: rgba(249,115,22,0.15); border-color: rgba(249,115,22,0.3); }
  .toggle-btn.czz { color: #f87171; }
  .toggle-btn.czz.on { background: rgba(220,38,38,0.15); border-color: rgba(220,38,38,0.3); }
  .toggle-btn.backwater { color: #60a5fa; }
  .toggle-btn.backwater.on { background: rgba(59,130,246,0.15); border-color: rgba(59,130,246,0.3); }

  .coord-toggle { display: flex; gap: 6px; margin-left: 12px; }
  .coord-toggle .toggle-btn {
    padding: 3px 8px; border-radius: 4px; font-size: 11px; cursor: pointer;
    border: 1px solid transparent; transition: all 0.15s;
  }
  .toggle-btn.seq { color: #a78bfa; }
  .toggle-btn.seq.on { background: rgba(167,139,250,0.15); border-color: rgba(167,139,250,0.3); }
  .toggle-btn.coord { color: #22d3ee; }
  .toggle-btn.coord.on { background: rgba(34,211,238,0.15); border-color: rgba(34,211,238,0.3); }
</style>
</head>
<body>

<div class="sidebar">
  <div class="sidebar-header">
    <h1>__TITLE__</h1>
    <p>断面形态 + 测量点位置组合图</p>
  </div>
  <div class="sidebar-search">
    <input type="text" id="searchInput" placeholder="搜索断面名称..." oninput="filterSections()">
  </div>
  <div class="sidebar-filter">
    <div class="filter-btn" onclick="toggleAnomalyFilter()" id="anomalyFilterBtn">仅显示异常断面<span class="filter-count" id="anomalyCount"></span></div>
    <div class="filter-btn" onclick="toggleValidationErrorFilter()" id="validationErrorFilterBtn">仅显示校验错误<span class="filter-count" id="validationErrorFilterCount"></span></div>
  </div>
  <div class="sidebar-stats">
    <div>流域: <span id="riverCount">0</span></div>
    <div>断面: <span id="totalCount">0</span></div>
    <div>测量点: <span id="totalPoints">0</span></div>
    <div>异常: <span id="anomalyTotalCount">0</span></div>
    <div>校验错误: <span id="validationErrorCount">0</span></div>
  </div>
  <div class="sidebar-pagination">
    <div class="page-nav">
      <button onclick="prevPage()" id="prevBtn">上一页</button>
      <span class="page-info" id="pageInfo">第 1 页</span>
      <button onclick="nextPage()" id="nextBtn">下一页</button>
    </div>
    <div class="page-size">
      <span>每页:</span>
      <select id="pageSizeSelect" onchange="changePageSize()">
        <option value="10">10</option>
        <option value="20">20</option>
        <option value="50" selected>50</option>
        <option value="100">100</option>
      </select>
    </div>
    <div>
      <button class="expand-btn" onclick="expandAll()">全部展开</button>
      <button class="expand-btn" onclick="collapseAll()">全部折叠</button>
    </div>
    <div>
      <button class="expand-btn" onclick="selectAllVisible()" id="selectAllBtn">全选当前页</button>
      <button class="expand-btn btn-primary" onclick="loadSelectedCharts()" id="loadChartsBtn">加载成图</button>
    </div>
  </div>
  <div class="section-list" id="sectionList"></div>
</div>

<div class="main">
  <div class="toolbar">
    <div class="toolbar-title" id="currentSection">请选择断面</div>
    <div class="toolbar-info" id="toolbarInfo"></div>
    <div class="water-toggle" id="waterToggle" style="display:none;">
      <div class="toggle-btn hmz on" onclick="toggleWater('hmz')" id="toggleHmz">历史最高水位</div>
      <div class="toggle-btn czz on" onclick="toggleWater('czz')" id="toggleCzz">成灾水位</div>
      <div class="toggle-btn backwater on" onclick="toggleWater('backwater')" id="toggleBackwater">壅水高程</div>
    </div>
    <div class="coord-toggle" id="coordToggle" style="display:none;">
      <div class="toggle-btn seq on" onclick="toggleCoordDisplay('seq')" id="toggleSeq">显示序号</div>
      <div class="toggle-btn coord on" onclick="toggleCoordDisplay('coord')" id="toggleCoord">显示经纬度</div>
    </div>
    <div class="toolbar-actions">
      <button class="btn" onclick="setCols(1)" id="col1">1列</button>
      <button class="btn active" onclick="setCols(2)" id="col2">2列</button>
      <button class="btn" onclick="setCols(3)" id="col3">3列</button>
      <button class="btn" onclick="setCols(4)" id="col4">4列</button>
    </div>
  </div>
  <div class="chart-area cols-2" id="chartArea">
    <div class="empty-state">请在左侧选择断面查看成图</div>
  </div>
</div>

<script>
const SECTIONS = __SECTIONS_JSON__;
const TREE = __TREE_JSON__;
const DEFAULT_FILTER_ANOMALY = __DEFAULT_FILTER__;
let selectedSections = [];
let chartInstances = {};
let cols = 2;
let showHmz = true, showCzz = true, showBackwater = true;
let showSeq = true, showCoord = false;
let filterAnomaly = DEFAULT_FILTER_ANOMALY;
let filterValidationError = false;
let expandedRivers = {};
let expandedDistricts = {};
let currentPage = 1;
let pageSize = 50;
let flatList = [];

function init() {
  let totalPts = 0;
  let totalSecs = 0;
  let anomalyCount = 0;
  let validationErrorCount = 0;
  Object.keys(SECTIONS).forEach(function(k) {
    totalSecs++;
    totalPts += SECTIONS[k].points.length;
    if (SECTIONS[k].anomaly) anomalyCount++;
    if (SECTIONS[k].validation_error) validationErrorCount++;
  });
  document.getElementById('riverCount').textContent = Object.keys(TREE).length;
  document.getElementById('totalCount').textContent = totalSecs;
  document.getElementById('totalPoints').textContent = totalPts;
  document.getElementById('anomalyTotalCount').textContent = anomalyCount;
  document.getElementById('anomalyCount').textContent = '(' + anomalyCount + ')';
  document.getElementById('validationErrorCount').textContent = validationErrorCount;
  document.getElementById('validationErrorFilterCount').textContent = '(' + validationErrorCount + ')';
  
  if (filterAnomaly) {
    document.getElementById('anomalyFilterBtn').classList.add('active');
  }
  if (filterValidationError) {
    document.getElementById('validationErrorFilterBtn').classList.add('active');
  }

  Object.keys(TREE).forEach(function(r) { expandedRivers[r] = false; });
  Object.keys(TREE).forEach(function(r) {
    Object.keys(TREE[r]).forEach(function(d) { expandedDistricts[r + '|' + d] = false; });
  });

  buildFlatList();
  selectAllVisible();
}

function buildFlatList(filter) {
  filter = filter || '';
  flatList = [];
  Object.keys(TREE).forEach(function(river) {
    Object.keys(TREE[river]).forEach(function(district) {
      TREE[river][district].forEach(function(item) {
        if (filterAnomaly && !item.anomaly) return;
        if (filterValidationError && !item.validation_error) return;
        if (!filter || item.name.includes(filter) || river.includes(filter) || district.includes(filter)) {
          flatList.push(item);
        }
      });
    });
  });
}

function toggleAnomalyFilter() {
  filterAnomaly = !filterAnomaly;
  document.getElementById('anomalyFilterBtn').classList.toggle('active', filterAnomaly);
  currentPage = 1;
  buildFlatList(document.getElementById('searchInput').value);
  selectAllVisible();
}

function toggleValidationErrorFilter() {
  filterValidationError = !filterValidationError;
  document.getElementById('validationErrorFilterBtn').classList.toggle('active', filterValidationError);
  currentPage = 1;
  buildFlatList(document.getElementById('searchInput').value);
  selectAllVisible();
}

function renderTree() {
  var list = document.getElementById('sectionList');
  var start = (currentPage - 1) * pageSize;
  var end = start + pageSize;
  var pageItems = flatList.slice(start, end);

  var totalPages = Math.ceil(flatList.length / pageSize);
  document.getElementById('pageInfo').textContent = '第 ' + currentPage + ' / ' + (totalPages || 1) + ' 页';
  document.getElementById('prevBtn').disabled = currentPage <= 1;
  document.getElementById('nextBtn').disabled = currentPage >= totalPages;

  var html = '';
  var seenRivers = new Set();
  var seenDistricts = new Set();

  pageItems.forEach(function(item) {
    var river = item.key.split('|')[0];
    var district = SECTIONS[item.key].district;
    var riverKey = river;
    var districtKey = river + '|' + district;

    if (!seenRivers.has(riverKey)) {
      var riverExpanded = expandedRivers[river];
      var riverCount = Object.keys(TREE[river]).reduce(function(s, d) { return s + TREE[river][d].length; }, 0);
      html += '<div class="tree-node">' +
        '<div class="tree-header river" onclick="toggleRiver(\\'' + river + '\\')">' +
        '<span class="arrow ' + (riverExpanded ? 'expanded' : '') + '">▶</span>' +
        '<span>' + river + '</span>' +
        '<span class="count">' + riverCount + '</span>' +
        '</div>' +
        '<div class="tree-children ' + (riverExpanded ? '' : 'hidden') + '" id="river_' + river + '"></div>' +
        '</div>';
      seenRivers.add(riverKey);
    }

    if (!seenDistricts.has(districtKey) && expandedRivers[river]) {
      var districtExpanded = expandedDistricts[districtKey];
      var districtCount = TREE[river][district].length;
      html += '<div class="tree-node" id="dist_' + districtKey + '">' +
        '<div class="tree-header district" onclick="toggleDistrict(\\'' + river + '\\', \\'' + district + '\\')">' +
        '<span class="arrow ' + (districtExpanded ? 'expanded' : '') + '">▶</span>' +
        '<span>' + district + '</span>' +
        '<span class="count">' + districtCount + '</span>' +
        '</div>' +
        '<div class="tree-children ' + (districtExpanded ? '' : 'hidden') + '" id="district_' + districtKey + '"></div>' +
        '</div>';
      seenDistricts.add(districtKey);
    }

    if (expandedRivers[river] && expandedDistricts[districtKey]) {
      var active = selectedSections.includes(item.key) ? 'active' : '';
      var badgeClass = item.features > 0 ? 'badge has-feature' : 'badge';
      var anomalyBadge = item.anomaly ? '<span class="badge has-anomaly">异常</span>' : '';
      var validationErrorBadge = item.validation_error ? '<span class="badge has-validation-error">校验错误</span>' : '';
      var validationWarningBadge = item.validation_warning && !item.validation_error ? '<span class="badge has-validation-warning">校验警告</span>' : '';
      html += '<div class="section-item ' + active + '" onclick="toggleSection(\\'' + item.key + '\\')">' +
        '<div style="display:flex;align-items:center;"><div class="checkbox"></div><span>' + item.name + '</span></div>' +
        '<div style="display:flex;gap:4px;">' + validationErrorBadge + validationWarningBadge + anomalyBadge + '<span class="' + badgeClass + '">' + item.points + '点' + (item.features > 0 ? '/' + item.features + '特征' : '') + '</span></div>' +
        '</div>';
    }
  });

  list.innerHTML = html;
}

function toggleRiver(river) {
  expandedRivers[river] = !expandedRivers[river];
  renderTree();
}

function toggleDistrict(river, district) {
  var key = river + '|' + district;
  expandedDistricts[key] = !expandedDistricts[key];
  renderTree();
}

function expandAll() {
  Object.keys(TREE).forEach(function(r) {
    expandedRivers[r] = true;
    Object.keys(TREE[r]).forEach(function(d) { expandedDistricts[r + '|' + d] = true; });
  });
  renderTree();
}

function collapseAll() {
  Object.keys(TREE).forEach(function(r) {
    expandedRivers[r] = false;
    Object.keys(TREE[r]).forEach(function(d) { expandedDistricts[r + '|' + d] = false; });
  });
  renderTree();
}

function changePageSize() {
  pageSize = parseInt(document.getElementById('pageSizeSelect').value);
  currentPage = 1;
  renderTree();
}

function prevPage() {
  if (currentPage > 1) { currentPage--; selectAllVisible(); }
}

function nextPage() {
  var totalPages = Math.ceil(flatList.length / pageSize);
  if (currentPage < totalPages) { currentPage++; selectAllVisible(); }
}

function filterSections() {
  var filter = document.getElementById('searchInput').value;
  buildFlatList(filter);
  currentPage = 1;
  renderTree();
}

function toggleSection(key) {
  var idx = selectedSections.indexOf(key);
  if (idx >= 0) { selectedSections.splice(idx, 1); }
  else { selectedSections.push(key); }
  updateSelectAllBtn();
  updateLoadBtn();
  renderTree();
}

function loadSelectedCharts() { renderCharts(); }

function updateLoadBtn() {
  var btn = document.getElementById('loadChartsBtn');
  if (btn) {
    btn.textContent = selectedSections.length > 0 ? '加载成图(' + selectedSections.length + ')' : '加载成图';
    btn.disabled = selectedSections.length === 0;
  }
}

function selectAllVisible() {
  var start = (currentPage - 1) * pageSize;
  var end = start + pageSize;
  var pageItems = flatList.slice(start, end);
  selectedSections = [];
  pageItems.forEach(function(item) { selectedSections.push(item.key); });
  updateSelectAllBtn();
  updateLoadBtn();
  renderTree();
  renderCharts();
}

function updateSelectAllBtn() {
  var btn = document.getElementById('selectAllBtn');
  if (btn) { btn.textContent = '全选当前页'; }
}

function setCols(n) {
  cols = n;
  document.querySelectorAll('.toolbar-actions .btn:not(.btn-primary)').forEach(function(b) { b.classList.remove('active'); });
  document.getElementById('col' + n).classList.add('active');
  document.getElementById('chartArea').className = 'chart-area cols-' + n;
  setTimeout(function() {
    Object.keys(chartInstances).forEach(function(k) {
      if (chartInstances[k].section) chartInstances[k].section.resize();
      if (chartInstances[k].coord) chartInstances[k].coord.resize();
    });
  }, 100);
}

function toggleWater(type) {
  if (type === 'hmz') showHmz = !showHmz;
  if (type === 'czz') showCzz = !showCzz;
  if (type === 'backwater') showBackwater = !showBackwater;
  document.getElementById('toggleHmz').classList.toggle('on', showHmz);
  document.getElementById('toggleCzz').classList.toggle('on', showCzz);
  document.getElementById('toggleBackwater').classList.toggle('on', showBackwater);
  renderCharts();
}

function toggleCoordDisplay(type) {
  if (type === 'seq') { showSeq = !showSeq; document.getElementById('toggleSeq').classList.toggle('on', showSeq); }
  if (type === 'coord') { showCoord = !showCoord; document.getElementById('toggleCoord').classList.toggle('on', showCoord); }
  renderCharts();
}

function renderCharts() {
  Object.keys(chartInstances).forEach(function(k) {
    if (chartInstances[k].section) chartInstances[k].section.dispose();
    if (chartInstances[k].coord) chartInstances[k].coord.dispose();
  });
  chartInstances = {};

  var area = document.getElementById('chartArea');
  var toggle = document.getElementById('waterToggle');
  var coordToggle = document.getElementById('coordToggle');

  if (selectedSections.length === 0) {
    area.innerHTML = '<div class="empty-state">请在左侧选择断面查看成图</div>';
    toggle.style.display = 'none';
    coordToggle.style.display = 'none';
    document.getElementById('currentSection').textContent = '请选择断面';
    document.getElementById('toolbarInfo').innerHTML = '';
    return;
  }

  toggle.style.display = 'flex';
  coordToggle.style.display = 'flex';
  document.getElementById('currentSection').textContent = '断面经纬度成图';
  document.getElementById('toolbarInfo').innerHTML = '已选 <span>' + selectedSections.length + '</span> 个断面';

  area.innerHTML = selectedSections.map(function(key) {
    var sec = SECTIONS[key];
    var river = sec.river;
    var tags = [];
    if (sec.hmz) tags.push('<span class="tag tag-hmz">历史最高水位: ' + sec.hmz.toFixed(2) + 'm</span>');
    if (sec.czz) tags.push('<span class="tag tag-czz">成灾水位: ' + sec.czz.toFixed(2) + 'm</span>');
    if (sec.backwater) tags.push('<span class="tag tag-backwater">壅水高程: ' + sec.backwater.toFixed(2) + 'm</span>');

    var legendItems = [];
    if (sec.hmz) legendItems.push('<div class="item"><div class="line" style="border-color:#f97316;"></div>历史最高水位 ' + sec.hmz.toFixed(2) + 'm</div>');
    if (sec.czz) legendItems.push('<div class="item"><div class="line" style="border-color:#dc2626;"></div>成灾水位 ' + sec.czz.toFixed(2) + 'm</div>');
    if (sec.backwater) legendItems.push('<div class="item"><div class="line" style="border-color:#3b82f6;"></div>壅水高程 ' + sec.backwater.toFixed(2) + 'm</div>');

    var safeKey = encodeURIComponent(key).replace(/[%'"]/g, '_');
    var anomalyTag = sec.anomaly ? '<span class="tag" style="background:rgba(250,204,21,0.15);color:#facc15;">经纬度异常</span>' : '';
    var validationErrorTag = sec.validation_error ? '<span class="tag" style="background:rgba(220,38,38,0.15);color:#ef4444;">校验错误(' + sec.validation_issues.length + ')</span>' : '';
    var validationWarningTag = sec.validation_warning && !sec.validation_error ? '<span class="tag" style="background:rgba(234,179,8,0.15);color:#eab308;">校验警告(' + sec.validation_issues.length + ')</span>' : '';
    return '<div class="chart-card' + (sec.anomaly ? ' anomaly-card' : '') + (sec.validation_error ? ' validation-error-card' : '') + '" id="card_' + safeKey + '">' +
      '<div class="chart-card-header">' +
      '<div>' +
      '<div class="chart-card-title">' + sec.name + '</div>' +
      '<div class="chart-card-subtitle">' + river + ' · ' + sec.district + ' · ' + sec.points.length + '个测量点</div>' +
      '</div>' +
      '<div class="chart-card-meta">' + validationErrorTag + validationWarningTag + anomalyTag + tags.join('') + '</div>' +
      '</div>' +
      '<div class="chart-card-body">' +
      '<div class="chart-section">' +
      '<div class="chart-section-title">断面形态 (起点距-高程)</div>' +
      '<div class="chart-section-chart" id="section_' + safeKey + '"></div>' +
      '</div>' +
      '<div class="chart-section">' +
      '<div class="chart-section-title">测量点位置 (经度-纬度)' + (sec.anomaly ? ' ⚠ 经纬度连线异常' : '') + (sec.validation_error ? ' ⚠ 数据校验错误' : '') + '</div>' +
      '<div class="chart-section-chart" id="coord_' + safeKey + '"></div>' +
      '</div>' +
      '</div>' +
      (legendItems.length > 0 ? '<div class="water-level-legend">' + legendItems.join('') + '</div>' : '') +
      '</div>';
  }).join('');

  area.className = 'chart-area cols-' + cols;

  selectedSections.forEach(function(key) {
    var sec = SECTIONS[key];
    if (!sec.points || sec.points.length === 0) return;
    var safeKey = encodeURIComponent(key).replace(/[%'"]/g, '_');

    var sectionDom = document.getElementById('section_' + safeKey);
    var coordDom = document.getElementById('coord_' + safeKey);

    if (sectionDom) {
      var sectionChart = echarts.init(sectionDom, null, { renderer: 'canvas' });
      sectionChart.setOption(buildSectionChartOption(sec));
      chartInstances[key] = chartInstances[key] || {};
      chartInstances[key].section = sectionChart;
    }

    if (coordDom) {
      var coordChart = echarts.init(coordDom, null, { renderer: 'canvas' });
      coordChart.setOption(buildCoordChartOption(sec));
      chartInstances[key] = chartInstances[key] || {};
      chartInstances[key].coord = coordChart;
    }
  });
}

function buildSectionChartOption(sec) {
  var points = sec.points;
  if (!points || points.length === 0) return {};

  var xData = points.map(function(p) { return p.distance; });
  var yData = points.map(function(p) { return p.elevation; });

  var yMin = Math.min.apply(null, yData);
  var yMax = Math.max.apply(null, yData);
  var yPad = (yMax - yMin) * 0.15 || 1;

  var markPoints = [];
  points.forEach(function(p, i) {
    if (p.isFeature) {
      markPoints.push({
        coord: [p.distance, p.elevation],
        name: p.feature,
        itemStyle: { color: '#ef4444' },
        symbolSize: 12,
      });
    }
  });

  var labelPoints = [];
  points.forEach(function(p, i) {
    var labelText = '';
    if (showSeq) { labelText = '#' + (p.seq || (i + 1)); }
    if (showCoord && p.lon && p.lat) {
      if (labelText) labelText += ' ';
      labelText += '(' + p.lon.toFixed(6) + ', ' + p.lat.toFixed(6) + ')';
    }
    if (labelText) {
      labelPoints.push({
        coord: [p.distance, p.elevation],
        name: labelText,
        value: labelText,
        itemStyle: { color: showCoord ? '#22d3ee' : '#a78bfa' },
        symbolSize: 8,
      });
    }
  });

  var series = [{
    type: 'line',
    name: '断面线',
    data: yData.map(function(y, i) { return [xData[i], y]; }),
    smooth: false,
    symbol: 'circle',
    symbolSize: 6,
    lineStyle: { color: '#10b981', width: 2 },
    itemStyle: { color: '#10b981' },
    areaStyle: {
      color: new echarts.graphic.LinearGradient(0, 0, 0, 1, [
        { offset: 0, color: 'rgba(16,185,129,0.3)' },
        { offset: 1, color: 'rgba(16,185,129,0.02)' }
      ])
    },
    markPoint: {
      data: markPoints.concat(labelPoints),
      symbol: 'circle',
      label: {
        show: true,
        position: 'top',
        formatter: function(p) { return p.name || ''; },
        fontSize: 10,
        color: '#a78bfa',
      }
    },
  }];

  var markLines = [];
  if (sec.hmz && showHmz) {
    markLines.push({
      yAxis: sec.hmz, name: '历史最高水位',
      lineStyle: { color: '#f97316', type: 'dashed', width: 2 },
      label: { show: true, formatter: '历史最高水位 ' + sec.hmz.toFixed(2) + 'm', position: 'insideEndTop', color: '#f97316', fontSize: 11 }
    });
  }
  if (sec.czz && showCzz) {
    markLines.push({
      yAxis: sec.czz, name: '成灾水位',
      lineStyle: { color: '#dc2626', type: 'dashed', width: 2 },
      label: { show: true, formatter: '成灾水位 ' + sec.czz.toFixed(2) + 'm', position: 'insideEndTop', color: '#dc2626', fontSize: 11 }
    });
  }
  if (sec.backwater && showBackwater) {
    markLines.push({
      yAxis: sec.backwater, name: '壅水高程',
      lineStyle: { color: '#3b82f6', type: 'dashed', width: 2 },
      label: { show: true, formatter: '壅水高程 ' + sec.backwater.toFixed(2) + 'm', position: 'insideEndTop', color: '#3b82f6', fontSize: 11 }
    });
  }

  if (markLines.length > 0) {
    series[0].markLine = { silent: true, symbol: 'none', data: markLines };
  }

  return {
    animation: true,
    animationDuration: 600,
    grid: { left: 50, right: 20, top: 20, bottom: 30 },
    tooltip: {
      trigger: 'axis',
      backgroundColor: 'rgba(15,23,42,0.9)',
      borderColor: '#334155',
      textStyle: { color: '#e2e8f0', fontSize: 12 },
      formatter: function(params) {
        var p = params[0];
        var idx = p.dataIndex;
        var pt = points[idx];
        var html = '<b>' + sec.name + '</b><br/>';
        html += '序号: #' + (pt.seq || (idx + 1)) + '<br/>';
        html += '起点距: ' + pt.distance.toFixed(3) + ' m<br/>';
        html += '高程: ' + pt.elevation.toFixed(3) + ' m';
        if (pt.lon && pt.lat) {
          html += '<br/>经度: ' + pt.lon.toFixed(6) + '°';
          html += '<br/>纬度: ' + pt.lat.toFixed(6) + '°';
        }
        if (pt.feature) html += '<br/>特征: ' + pt.feature;
        return html;
      }
    },
    xAxis: {
      type: 'value', name: '起点距(m)',
      nameTextStyle: { color: '#94a3b8', fontSize: 10 },
      axisLine: { lineStyle: { color: '#475569' } },
      axisLabel: { color: '#94a3b8', fontSize: 10 },
      splitLine: { lineStyle: { color: '#1e293b' } },
      min: function(value) { return Math.floor(value.min - (value.max - value.min) * 0.02); },
    },
    yAxis: {
      type: 'value', name: '高程(m)',
      nameTextStyle: { color: '#94a3b8', fontSize: 10 },
      axisLine: { lineStyle: { color: '#475569' } },
      axisLabel: { color: '#94a3b8', fontSize: 10, formatter: function(v) { return v.toFixed(1); } },
      splitLine: { lineStyle: { color: '#1e293b' } },
      min: yMin - yPad, max: yMax + yPad,
    },
    series: series,
    dataZoom: [
      { type: 'inside', xAxisIndex: 0, filterMode: 'none' },
      { type: 'inside', yAxisIndex: 0, filterMode: 'none' },
    ],
  };
}

function buildCoordChartOption(sec) {
  var points = sec.points;
  if (!points || points.length === 0) return {};

  var validPoints = points.filter(function(p) { return p.lon && p.lat; });
  if (validPoints.length === 0) {
    return {
      title: { text: '无经纬度数据', left: 'center', top: 'center', textStyle: { color: '#64748b', fontSize: 12 } }
    };
  }

  validPoints.sort(function(a, b) {
    var seqA = a.seq || points.indexOf(a) + 1;
    var seqB = b.seq || points.indexOf(b) + 1;
    return seqA - seqB;
  });

  var lonData = validPoints.map(function(p) { return p.lon; });
  var latData = validPoints.map(function(p) { return p.lat; });

  var lonMin = Math.min.apply(null, lonData);
  var lonMax = Math.max.apply(null, lonData);
  var latMin = Math.min.apply(null, latData);
  var latMax = Math.max.apply(null, latData);

  var lonPad = (lonMax - lonMin) * 0.1 || 0.001;
  var latPad = (latMax - latMin) * 0.1 || 0.001;

  var scatterData = validPoints.map(function(p, i) {
    return {
      value: [p.lon, p.lat],
      seq: p.seq || (i + 1),
      distance: p.distance,
      elevation: p.elevation,
      feature: p.feature,
      isFeature: p.isFeature,
      itemStyle: { color: p.isFeature ? '#ef4444' : '#22d3ee' }
    };
  });

  var lineData = scatterData.map(function(d) { return d.value; });
  var isAnomaly = sec.anomaly || false;
  var lineColor = isAnomaly ? '#facc15' : '#22d3ee';

  return {
    animation: true,
    animationDuration: 600,
    grid: { left: 55, right: 20, top: 20, bottom: 30 },
    tooltip: {
      trigger: 'item',
      backgroundColor: 'rgba(15,23,42,0.9)',
      borderColor: '#334155',
      textStyle: { color: '#e2e8f0', fontSize: 12 },
      formatter: function(params) {
        var d = params.data;
        var html = '<b>' + sec.name + '</b><br/>';
        if (d.seq) {
          html += '序号: #' + d.seq + '<br/>';
          html += '经度: ' + d.value[0].toFixed(6) + '°<br/>';
          html += '纬度: ' + d.value[1].toFixed(6) + '°<br/>';
          html += '起点距: ' + d.distance.toFixed(3) + ' m<br/>';
          html += '高程: ' + d.elevation.toFixed(3) + ' m';
          if (d.feature) html += '<br/>特征: ' + d.feature;
        } else {
          html += '测量轨迹线';
        }
        return html;
      }
    },
    xAxis: {
      type: 'value', name: '经度(°)',
      nameTextStyle: { color: '#94a3b8', fontSize: 10 },
      axisLine: { lineStyle: { color: '#475569' } },
      axisLabel: { color: '#94a3b8', fontSize: 10, formatter: function(v) { return v.toFixed(4); } },
      splitLine: { lineStyle: { color: '#1e293b' } },
      min: lonMin - lonPad, max: lonMax + lonPad,
    },
    yAxis: {
      type: 'value', name: '纬度(°)',
      nameTextStyle: { color: '#94a3b8', fontSize: 10 },
      axisLine: { lineStyle: { color: '#475569' } },
      axisLabel: { color: '#94a3b8', fontSize: 10, formatter: function(v) { return v.toFixed(4); } },
      splitLine: { lineStyle: { color: '#1e293b' } },
      min: latMin - latPad, max: latMax + latPad,
    },
    series: [
      {
        type: 'line',
        name: '测量轨迹',
        data: lineData,
        smooth: false,
        symbol: 'none',
        lineStyle: { color: lineColor, width: 2, type: isAnomaly ? 'dashed' : 'solid' },
        tooltip: { show: false }
      },
      {
        type: 'scatter',
        name: '测量点',
        data: scatterData,
        symbolSize: function(val, params) {
          return params.data.isFeature ? 12 : 8;
        },
        label: {
          show: showSeq,
          position: 'top',
          formatter: function(params) { return '#' + params.data.seq; },
          fontSize: 10,
          color: '#a78bfa'
        },
        itemStyle: { borderColor: '#fff', borderWidth: 1 },
        emphasis: { itemStyle: { shadowBlur: 10, shadowColor: 'rgba(34,211,238,0.5)' } }
      }
    ],
    dataZoom: [
      { type: 'inside', xAxisIndex: 0, filterMode: 'none' },
      { type: 'inside', yAxisIndex: 0, filterMode: 'none' },
    ],
  };
}

window.addEventListener('resize', function() {
  Object.keys(chartInstances).forEach(function(k) {
    if (chartInstances[k].section) chartInstances[k].section.resize();
    if (chartInstances[k].coord) chartInstances[k].coord.resize();
  });
});

init();
</script>
</body>
</html>"""


def generate_html(sections: Dict, tree: Dict, river_name: str = "松花江水系", 
                  default_filter_anomaly: bool = False) -> str:
    sections_json = json.dumps(sections, ensure_ascii=False).replace("</script>", "<\\/script>")
    tree_json = json.dumps(tree, ensure_ascii=False).replace("</script>", "<\\/script>")
    title = f"{river_name} 横断面经纬度成图"
    
    html = HTML_TEMPLATE
    html = html.replace("__SECTIONS_JSON__", sections_json)
    html = html.replace("__TREE_JSON__", tree_json)
    html = html.replace("__TITLE__", title)
    html = html.replace("__DEFAULT_FILTER__", "true" if default_filter_anomaly else "false")
    return html


def generate_anomaly_html(all_sections: Dict, output_path: str):
    anomaly_sections = {k: v for k, v in all_sections.items() if v.get("anomaly")}
    
    if not anomaly_sections:
        print("  无异常断面数据，不生成异常断面HTML")
        return
    
    anomaly_tree = build_tree_data(anomaly_sections)
    html = generate_html(anomaly_sections, anomaly_tree, "异常断面汇总", default_filter_anomaly=True)
    
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(html)
    
    sec_count = len(anomaly_sections)
    pt_count = sum(len(s["points"]) for s in anomaly_sections.values())
    file_size = os.path.getsize(output_path) / 1024
    print(f"  异常断面HTML已保存: {output_path} ({sec_count}个异常断面, {pt_count}个测量点, {file_size:.1f} KB)")


def main():
    print("开始批量读取横断面数据（含经纬度）...")
    files = find_excel_files()
    print(f"找到 {len(files)} 个横断面Excel文件")

    all_sections = {}
    for river, file_path in files:
        print(f"  读取: {river} - {os.path.basename(file_path)}")
        sections = read_sections_from_file(file_path, river)
        all_sections.update(sections)

    print(f"共解析 {len(all_sections)} 个断面")
    total_points = sum(len(s["points"]) for s in all_sections.values())
    print(f"共 {total_points} 个测量点")

    coords_count = sum(1 for s in all_sections.values() for p in s["points"] if p["lon"] and p["lat"])
    print(f"含经纬度数据的测量点: {coords_count}")

    anomaly_count = 0
    for key, sec in all_sections.items():
        is_anomaly, anomaly_details = detect_anomaly(sec)
        sec["anomaly"] = is_anomaly
        sec["anomaly_details"] = anomaly_details
        if is_anomaly:
            anomaly_count += 1
    print(f"异常断面（经纬度连线交叉/方向不一致）: {anomaly_count}")

    validation_error_count = 0
    validation_warning_count = 0
    total_issues = 0
    for key, sec in all_sections.items():
        issues = validate_section_data(sec)
        sec["validation_issues"] = issues
        sec["validation_error"] = any(i["level"] == "error" for i in issues)
        sec["validation_warning"] = any(i["level"] == "warning" for i in issues)
        if sec["validation_error"]:
            validation_error_count += 1
        if sec["validation_warning"] or sec["validation_error"]:
            validation_warning_count += 1
        total_issues += len(issues)
    print(f"数据校验: {validation_error_count} 个断面有错误, {validation_warning_count} 个断面有警告, 共 {total_issues} 个问题")

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    anomaly_xlsx_path = os.path.join(OUTPUT_DIR, f"异常断面汇总_{timestamp}.xlsx")
    anomaly_html_path = os.path.join(OUTPUT_DIR, f"异常断面汇总_{timestamp}.html")
    
    write_anomaly_report(all_sections, anomaly_xlsx_path)
    generate_anomaly_html(all_sections, anomaly_html_path)

    for river_name in RIVER_DIRS:
        river_sections = {k: v for k, v in all_sections.items() if v["river"] == river_name}
        if not river_sections:
            print(f"  跳过无数据的流域: {river_name}")
            continue

        river_tree = build_tree_data(river_sections)
        html = generate_html(river_sections, river_tree, river_name)

        output_path = os.path.join(OUTPUT_DIR, f"{river_name}_横断面经纬度成图.html")
        with open(output_path, "w", encoding="utf-8") as f:
            f.write(html)

        sec_count = len(river_sections)
        pt_count = sum(len(s["points"]) for s in river_sections.values())
        file_size = os.path.getsize(output_path) / 1024
        print(f"  生成: {river_name}_横断面经纬度成图.html ({sec_count}个断面, {pt_count}个测量点, {file_size:.1f} KB)")

    print(f"\n全部完成! 共 {len(RIVER_DIRS)} 个流域目录")


if __name__ == "__main__":
    main()