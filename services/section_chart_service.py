# -*- coding: utf-8 -*-
import os
import json
import sqlite3
import math
import glob
from typing import Dict, List, Tuple, Optional, Any
from datetime import datetime

import numpy as np
import pandas as pd
import openpyxl


DB_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "data")
DB_PATH = os.path.join(DB_DIR, "section_chart.db")

DEFAULT_FEATURE_KEYWORDS = ["基点", "堤顶", "坡顶", "坡脚", "堤脚", "左岸", "右岸", "深泓", "深泓点", "主槽", "滩地", "桥", "涵", "堰", "闸", "坝"]

FEATURE_KEYWORDS = DEFAULT_FEATURE_KEYWORDS.copy()

VALIDATION_CONFIG = {
    "lon_range": (73.0, 135.0),
    "lat_range": (18.0, 54.0),
    "dist_jump_factor": 10,
    "coord_jump_threshold": 0.01,
    "coord_jump_ratio": 0.2,
    "linear_fit_threshold": 0.05,
}

VALIDATION_RULES = {
    "seq_missing": {"desc": "序号缺失", "level": "error"},
    "seq_duplicate": {"desc": "序号重复", "level": "error"},
    "seq_not_monotonic": {"desc": "序号非递增", "level": "warning"},
    "seq_gap": {"desc": "序号跳号", "level": "warning"},
    "dist_negative": {"desc": "起点距为负", "level": "error"},
    "dist_not_monotonic": {"desc": "起点距回退", "level": "error"},
    "dist_duplicate": {"desc": "起点距重复", "level": "warning"},
    "dist_jump": {"desc": "起点距跳变", "level": "warning"},
    "lon_out_of_range": {"desc": "经度超范围", "level": "error"},
    "lon_zero": {"desc": "经度为零或缺失", "level": "error"},
    "lon_jump": {"desc": "经度跳变", "level": "warning"},
    "lat_out_of_range": {"desc": "纬度超范围", "level": "error"},
    "lat_zero": {"desc": "纬度为零或缺失", "level": "error"},
    "lat_jump": {"desc": "纬度跳变", "level": "warning"},
    "lonlat_swapped": {"desc": "经纬度互换", "level": "error"},
    "coord_missing": {"desc": "经纬度缺失", "level": "warning"},
    "all_same_coord": {"desc": "所有点坐标相同", "level": "error"},
    "linear_fit_deviation": {"desc": "线性拟合偏差", "level": "warning"},
    "direction_reversal": {"desc": "连线方向折返", "level": "error"},
    "adjacent_reversal": {"desc": "相邻线段折返", "level": "error"},
}


def _ensure_db_dir():
    os.makedirs(DB_DIR, exist_ok=True)


def get_feature_keywords() -> List[str]:
    global FEATURE_KEYWORDS
    _ensure_db_dir()
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("CREATE TABLE IF NOT EXISTS config (key TEXT PRIMARY KEY, value TEXT)")
    c.execute("SELECT value FROM config WHERE key = 'feature_keywords'")
    row = c.fetchone()
    conn.close()
    if row:
        try:
            keywords = json.loads(row[0])
            FEATURE_KEYWORDS = keywords
            return keywords
        except:
            pass
    return FEATURE_KEYWORDS


def save_feature_keywords(keywords: List[str]):
    global FEATURE_KEYWORDS
    _ensure_db_dir()
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("CREATE TABLE IF NOT EXISTS config (key TEXT PRIMARY KEY, value TEXT)")
    c.execute("INSERT OR REPLACE INTO config (key, value) VALUES ('feature_keywords', ?)", (json.dumps(keywords),))
    conn.commit()
    conn.close()
    FEATURE_KEYWORDS = keywords


def safe_float(val: Any) -> Optional[float]:
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def haversine(lon1, lat1, lon2, lat2):
    lon1, lat1, lon2, lat2 = map(math.radians, [lon1, lat1, lon2, lat2])
    dlon = lon2 - lon1
    dlat = lat2 - lat1
    a = math.sin(dlat / 2) ** 2 + math.cos(lat1) * math.cos(lat2) * math.sin(dlon / 2) ** 2
    c = 2 * math.asin(math.sqrt(a))
    r = 6371000
    return c * r


def lonlat_to_local_xy(lon, lat, ref_lon, ref_lat):
    x = (lon - ref_lon) * math.cos(math.radians(ref_lat)) * 111320.0
    y = (lat - ref_lat) * 110540.0
    return x, y


def cross_product(o, a, b):
    return (a[0] - o[0]) * (b[1] - o[1]) - (a[1] - o[1]) * (b[0] - o[0])


def segments_intersect(p1, p2, p3, p4):
    d1 = cross_product(p3, p4, p1)
    d2 = cross_product(p3, p4, p2)
    d3 = cross_product(p1, p2, p3)
    d4 = cross_product(p1, p2, p4)
    return ((d1 > 0 and d2 < 0) or (d1 < 0 and d2 > 0)) and \
           ((d3 > 0 and d4 < 0) or (d3 < 0 and d4 > 0))


class SpatialContinuityChecker:
    def __init__(self, df):
        self.df = df.copy().reset_index(drop=True)
        self.n = len(df)
        self._pre_calculate()

    def _pre_calculate(self):
        if self.n < 2:
            return
        self.chain_dists = np.abs(np.diff(self.df['distance'].values))
        coord_dists = []
        for i in range(self.n - 1):
            d = haversine(self.df.loc[i, 'lon'], self.df.loc[i, 'lat'],
                          self.df.loc[i + 1, 'lon'], self.df.loc[i + 1, 'lat'])
            coord_dists.append(d)
        self.coord_dists = np.array(coord_dists)
        ref_lon = self.df.loc[0, 'lon']
        ref_lat = self.df.loc[0, 'lat']
        self.local_coords = []
        for i in range(self.n):
            x, y = lonlat_to_local_xy(self.df.loc[i, 'lon'], self.df.loc[i, 'lat'], ref_lon, ref_lat)
            self.local_coords.append((x, y))
        if self.n >= 2:
            self.main_direction = self._compute_main_direction()

    def _compute_main_direction(self):
        start = self.local_coords[0]
        end = self.local_coords[-1]
        dx = end[0] - start[0]
        dy = end[1] - start[1]
        length = math.sqrt(dx ** 2 + dy ** 2)
        if length < 1:
            return None
        return (dx / length, dy / length)

    def check_linear_fit_deviation(self, threshold=0.05):
        flagged_indices = set()
        if self.n < 3:
            return flagged_indices
        ref_lon = self.df.loc[0, 'lon']
        ref_lat = self.df.loc[0, 'lat']
        x_vals = np.array([lonlat_to_local_xy(row['lon'], row['lat'], ref_lon, ref_lat)[0] for _, row in self.df.iterrows()])
        y_vals = np.array([lonlat_to_local_xy(row['lon'], row['lat'], ref_lon, ref_lat)[1] for _, row in self.df.iterrows()])
        L = np.max(self.df['distance']) - np.min(self.df['distance'])
        if L <= 0:
            L = np.max(self.coord_dists)
        X = np.vstack([x_vals, y_vals]).T
        X_centered = X - np.mean(X, axis=0)
        cov_matrix = np.cov(X_centered.T)
        eig_vals, eig_vecs = np.linalg.eig(cov_matrix)
        max_idx = np.argmax(eig_vals)
        direction = eig_vecs[:, max_idx]
        A = direction[1]
        B = -direction[0]
        C = -(A * np.mean(x_vals) + B * np.mean(y_vals))
        norm = np.sqrt(A ** 2 + B ** 2)
        A, B, C = A / norm, B / norm, C / norm
        distances = np.abs(A * x_vals + B * y_vals + C)
        for i in range(self.n):
            if distances[i] > threshold * L:
                flagged_indices.add(i)
        return flagged_indices

    def check_direction_consistency(self, reversal_threshold=90):
        anomalies = []
        if self.n < 3 or self.main_direction is None:
            return anomalies
        for i in range(self.n - 1):
            p1 = self.local_coords[i]
            p2 = self.local_coords[i + 1]
            dx = p2[0] - p1[0]
            dy = p2[1] - p1[1]
            seg_length = math.sqrt(dx ** 2 + dy ** 2)
            if seg_length < 0.5:
                continue
            seg_unit = (dx / seg_length, dy / seg_length)
            dot = seg_unit[0] * self.main_direction[0] + seg_unit[1] * self.main_direction[1]
            dot_clamped = max(-1.0, min(1.0, dot))
            angle_deg = math.degrees(math.acos(dot_clamped))
            if angle_deg > reversal_threshold:
                anomalies.append({
                    'type': 'reversal',
                    'point_index': i + 1,
                    'point_seq': self.df.loc[i + 1, 'seq'] if 'seq' in self.df.columns else i + 1,
                    'angle': angle_deg,
                    'lon': self.df.loc[i + 1, 'lon'],
                    'lat': self.df.loc[i + 1, 'lat'],
                    'desc': f'线段与主方向夹角{angle_deg:.1f}°，超过{reversal_threshold}°阈值'
                })
        return anomalies

    def check_adjacent_segment_angle(self, sharp_turn_threshold=90):
        anomalies = []
        if self.n < 3:
            return anomalies
        prev_seg_unit = None
        for i in range(self.n - 1):
            p1 = self.local_coords[i]
            p2 = self.local_coords[i + 1]
            dx = p2[0] - p1[0]
            dy = p2[1] - p1[1]
            seg_length = math.sqrt(dx ** 2 + dy ** 2)
            if seg_length < 0.5:
                continue
            seg_unit = (dx / seg_length, dy / seg_length)
            if prev_seg_unit is not None:
                dot = seg_unit[0] * prev_seg_unit[0] + seg_unit[1] * prev_seg_unit[1]
                dot_clamped = max(-1.0, min(1.0, dot))
                angle_deg = math.degrees(math.acos(dot_clamped))
                if angle_deg > sharp_turn_threshold:
                    anomalies.append({
                        'type': 'sharp_turn',
                        'point_index': i + 1,
                        'point_seq': self.df.loc[i + 1, 'seq'] if 'seq' in self.df.columns else i + 1,
                        'angle': angle_deg,
                        'lon': self.df.loc[i + 1, 'lon'],
                        'lat': self.df.loc[i + 1, 'lat'],
                        'desc': f'相邻线段夹角{angle_deg:.1f}°，超过{sharp_turn_threshold}°阈值'
                    })
            prev_seg_unit = seg_unit
        return anomalies


def detect_anomaly(sec: Dict, reversal_threshold: float = 90.0) -> Tuple[bool, List[Dict]]:
    anomalies = []
    points = [p for p in sec["points"] if p.get("lon") is not None and p.get("lat") is not None]
    if len(points) < 3:
        return False, anomalies
    points.sort(key=lambda p: p.get("seq", 0) if p.get("seq") is not None else 0)
    ref_lon = points[0]["lon"]
    ref_lat = points[0]["lat"]
    local_coords = []
    for p in points:
        x, y = lonlat_to_local_xy(p["lon"], p["lat"], ref_lon, ref_lat)
        local_coords.append((x, y))
    start = local_coords[0]
    end = local_coords[-1]
    dx_main = end[0] - start[0]
    dy_main = end[1] - start[1]
    main_length = math.sqrt(dx_main ** 2 + dy_main ** 2)
    if main_length < 1:
        coords = [(p["lon"], p["lat"]) for p in points]
        for i in range(len(coords) - 1):
            for j in range(i + 2, len(coords) - 1):
                if segments_intersect(coords[i], coords[i + 1], coords[j], coords[j + 1]):
                    anomalies.append({
                        "type": "cross", "desc": "线段交叉",
                        "segment1_start_seq": points[i]["seq"],
                        "segment1_end_seq": points[i + 1]["seq"],
                        "segment2_start_seq": points[j]["seq"],
                        "segment2_end_seq": points[j + 1]["seq"],
                    })
        return len(anomalies) > 0, anomalies
    main_unit = (dx_main / main_length, dy_main / main_length)
    for i in range(len(local_coords) - 1):
        p1 = local_coords[i]
        p2 = local_coords[i + 1]
        dx = p2[0] - p1[0]
        dy = p2[1] - p1[1]
        seg_length = math.sqrt(dx ** 2 + dy ** 2)
        if seg_length < 0.5:
            continue
        seg_unit = (dx / seg_length, dy / seg_length)
        dot = seg_unit[0] * main_unit[0] + seg_unit[1] * main_unit[1]
        dot_clamped = max(-1.0, min(1.0, dot))
        angle_deg = math.degrees(math.acos(dot_clamped))
        if angle_deg > reversal_threshold:
            anomalies.append({
                "type": "reversal", "desc": "方向反转",
                "point_seq": points[i + 1]["seq"],
                "lon": points[i + 1]["lon"], "lat": points[i + 1]["lat"],
                "prev_lon": points[i]["lon"], "prev_lat": points[i]["lat"],
                "angle": angle_deg,
                "detail": f"线段与主方向夹角{angle_deg:.1f}°"
            })
    prev_seg_unit = None
    for i in range(len(local_coords) - 1):
        p1 = local_coords[i]
        p2 = local_coords[i + 1]
        dx = p2[0] - p1[0]
        dy = p2[1] - p1[1]
        seg_length = math.sqrt(dx ** 2 + dy ** 2)
        if seg_length < 0.5:
            continue
        seg_unit = (dx / seg_length, dy / seg_length)
        if prev_seg_unit is not None:
            dot = seg_unit[0] * prev_seg_unit[0] + seg_unit[1] * prev_seg_unit[1]
            dot_clamped = max(-1.0, min(1.0, dot))
            angle_deg = math.degrees(math.acos(dot_clamped))
            if angle_deg > reversal_threshold:
                anomalies.append({
                    "type": "sharp_turn", "desc": "相邻线段折返",
                    "point_seq": points[i + 1]["seq"],
                    "lon": points[i + 1]["lon"], "lat": points[i + 1]["lat"],
                    "prev_lon": points[i]["lon"], "prev_lat": points[i]["lat"],
                    "angle": angle_deg,
                    "detail": f"相邻线段夹角{angle_deg:.1f}°"
                })
        prev_seg_unit = seg_unit
    coords = [(p["lon"], p["lat"]) for p in points]
    for i in range(len(coords) - 1):
        for j in range(i + 2, len(coords) - 1):
            if segments_intersect(coords[i], coords[i + 1], coords[j], coords[j + 1]):
                anomalies.append({
                    "type": "cross", "desc": "线段交叉",
                    "segment1_start_seq": points[i]["seq"],
                    "segment1_end_seq": points[i + 1]["seq"],
                    "segment2_start_seq": points[j]["seq"],
                    "segment2_end_seq": points[j + 1]["seq"],
                })
    return len(anomalies) > 0, anomalies


def validate_section_data(sec: Dict) -> List[Dict]:
    issues = []
    points = sec.get("points", [])
    table_type = sec.get("table_type", "")
    is_chengtu = table_type == "chengtu"
    if len(points) == 0:
        return issues
    seq_values = []
    seq_missing_indices = []
    for i, p in enumerate(points):
        seq = p.get("seq")
        if seq is None:
            seq_missing_indices.append(i)
        else:
            try:
                seq_int = int(seq)
                seq_values.append((seq_int, i))
            except (ValueError, TypeError):
                seq_missing_indices.append(i)
    if seq_missing_indices:
        issues.append({"rule": "seq_missing", "level": VALIDATION_RULES["seq_missing"]["level"],
                        "desc": VALIDATION_RULES["seq_missing"]["desc"],
                        "details": f"第 {', '.join(str(i + 1) for i in seq_missing_indices)} 个点序号缺失",
                        "point_index": seq_missing_indices})
    if seq_values:
        seq_counts = {}
        for seq, idx in seq_values:
            if seq in seq_counts:
                seq_counts[seq].append(idx)
            else:
                seq_counts[seq] = [idx]
        for seq, indices in seq_counts.items():
            if len(indices) > 1:
                issues.append({"rule": "seq_duplicate", "level": VALIDATION_RULES["seq_duplicate"]["level"],
                                "desc": VALIDATION_RULES["seq_duplicate"]["desc"],
                                "details": f"序号 {seq} 出现 {len(indices)} 次", "point_index": indices})
        sorted_seq = sorted(seq_values, key=lambda x: x[0])
        not_monotonic_indices = []
        for i in range(1, len(sorted_seq)):
            if sorted_seq[i][0] <= sorted_seq[i - 1][0]:
                not_monotonic_indices.append(sorted_seq[i][1])
        if not_monotonic_indices:
            issues.append({"rule": "seq_not_monotonic", "level": VALIDATION_RULES["seq_not_monotonic"]["level"],
                            "desc": VALIDATION_RULES["seq_not_monotonic"]["desc"],
                            "details": f"序号非递增，涉及第 {', '.join(str(i + 1) for i in not_monotonic_indices)} 个点",
                            "point_index": not_monotonic_indices})
        gap_indices = []
        for i in range(1, len(sorted_seq)):
            gap = sorted_seq[i][0] - sorted_seq[i - 1][0]
            if gap > 1:
                gap_indices.append(sorted_seq[i][1])
        if gap_indices:
            issues.append({"rule": "seq_gap", "level": VALIDATION_RULES["seq_gap"]["level"],
                            "desc": VALIDATION_RULES["seq_gap"]["desc"],
                            "details": f"序号跳号，涉及第 {', '.join(str(i + 1) for i in gap_indices)} 个点",
                            "point_index": gap_indices})
    dist_values = [(p.get("distance"), i) for i, p in enumerate(points)]
    negative_indices = [idx for dist, idx in dist_values if dist is not None and dist < 0]
    if negative_indices:
        issues.append({"rule": "dist_negative", "level": VALIDATION_RULES["dist_negative"]["level"],
                        "desc": VALIDATION_RULES["dist_negative"]["desc"],
                        "details": f"起点距为负，涉及第 {', '.join(str(i + 1) for i in negative_indices)} 个点",
                        "point_index": negative_indices})
    valid_dist = [(dist, idx) for dist, idx in dist_values if dist is not None]
    if valid_dist:
        sorted_dist = sorted(valid_dist, key=lambda x: x[0])
        not_monotonic_indices = []
        for i in range(1, len(sorted_dist)):
            if sorted_dist[i][0] < sorted_dist[i - 1][0]:
                not_monotonic_indices.append(sorted_dist[i][1])
        if not_monotonic_indices:
            issues.append({"rule": "dist_not_monotonic", "level": VALIDATION_RULES["dist_not_monotonic"]["level"],
                            "desc": VALIDATION_RULES["dist_not_monotonic"]["desc"],
                            "details": f"起点距回退，涉及第 {', '.join(str(i + 1) for i in not_monotonic_indices)} 个点",
                            "point_index": not_monotonic_indices})
        dist_counts = {}
        for dist, idx in valid_dist:
            key = round(dist, 3)
            if key in dist_counts:
                dist_counts[key].append(idx)
            else:
                dist_counts[key] = [idx]
        for dist_key, indices in dist_counts.items():
            if len(indices) > 1:
                issues.append({"rule": "dist_duplicate", "level": VALIDATION_RULES["dist_duplicate"]["level"],
                                "desc": VALIDATION_RULES["dist_duplicate"]["desc"],
                                "details": f"起点距 {dist_key:.3f}m 重复出现 {len(indices)} 次",
                                "point_index": indices})
        if len(valid_dist) > 1:
            dist_diffs = [abs(valid_dist[i + 1][0] - valid_dist[i][0]) for i in range(len(valid_dist) - 1)]
            avg_diff = sum(dist_diffs) / len(dist_diffs) if dist_diffs else 0
            jump_indices = []
            for i, diff in enumerate(dist_diffs):
                if avg_diff > 0 and diff > avg_diff * VALIDATION_CONFIG["dist_jump_factor"]:
                    jump_indices.append(valid_dist[i + 1][1])
            if jump_indices:
                issues.append({"rule": "dist_jump", "level": VALIDATION_RULES["dist_jump"]["level"],
                                "desc": VALIDATION_RULES["dist_jump"]["desc"],
                                "details": f"起点距跳变(平均间距{avg_diff:.1f}m)，涉及第 {', '.join(str(i + 1) for i in jump_indices)} 个点",
                                "point_index": jump_indices})
    lon_values = [(p.get("lon"), i) for i, p in enumerate(points)]
    lat_values = [(p.get("lat"), i) for i, p in enumerate(points)]
    lon_range = VALIDATION_CONFIG["lon_range"]
    lat_range = VALIDATION_CONFIG["lat_range"]
    out_of_range_indices = []
    for lon, idx in lon_values:
        if lon is not None and (lon < lon_range[0] or lon > lon_range[1]):
            out_of_range_indices.append(idx)
    if out_of_range_indices:
        issues.append({"rule": "lon_out_of_range", "level": VALIDATION_RULES["lon_out_of_range"]["level"],
                        "desc": VALIDATION_RULES["lon_out_of_range"]["desc"],
                        "details": f"经度超出{lon_range[0]}~{lon_range[1]}范围", "point_index": out_of_range_indices})
    out_of_range_indices = []
    for lat, idx in lat_values:
        if lat is not None and (lat < lat_range[0] or lat > lat_range[1]):
            out_of_range_indices.append(idx)
    if out_of_range_indices:
        issues.append({"rule": "lat_out_of_range", "level": VALIDATION_RULES["lat_out_of_range"]["level"],
                        "desc": VALIDATION_RULES["lat_out_of_range"]["desc"],
                        "details": f"纬度超出{lat_range[0]}~{lat_range[1]}范围", "point_index": out_of_range_indices})
    zero_lon_indices = [idx for lon, idx in lon_values if lon is None or lon == 0]
    if zero_lon_indices and not is_chengtu:
        issues.append({"rule": "lon_zero", "level": VALIDATION_RULES["lon_zero"]["level"],
                        "desc": VALIDATION_RULES["lon_zero"]["desc"],
                        "details": f"经度为零或缺失", "point_index": zero_lon_indices})
    zero_lat_indices = [idx for lat, idx in lat_values if lat is None or lat == 0]
    if zero_lat_indices and not is_chengtu:
        issues.append({"rule": "lat_zero", "level": VALIDATION_RULES["lat_zero"]["level"],
                        "desc": VALIDATION_RULES["lat_zero"]["desc"],
                        "details": f"纬度为零或缺失", "point_index": zero_lat_indices})
    valid_lon = [(lon, idx) for lon, idx in lon_values if lon is not None and lon != 0]
    valid_lat = [(lat, idx) for lat, idx in lat_values if lat is not None and lat != 0]
    if len(valid_lon) > 1:
        lon_data = [lon for lon, idx in valid_lon]
        lon_span = max(lon_data) - min(lon_data)
        threshold = max(VALIDATION_CONFIG["coord_jump_threshold"], lon_span * VALIDATION_CONFIG["coord_jump_ratio"])
        jump_indices = []
        for i in range(1, len(valid_lon)):
            diff = abs(valid_lon[i][0] - valid_lon[i - 1][0])
            if diff > threshold:
                jump_indices.append(valid_lon[i][1])
        if jump_indices:
            issues.append({"rule": "lon_jump", "level": VALIDATION_RULES["lon_jump"]["level"],
                            "desc": VALIDATION_RULES["lon_jump"]["desc"],
                            "details": f"经度跳变(跨度{lon_span:.4f}°)", "point_index": jump_indices})
    if len(valid_lat) > 1:
        lat_data = [lat for lat, idx in valid_lat]
        lat_span = max(lat_data) - min(lat_data)
        threshold = max(VALIDATION_CONFIG["coord_jump_threshold"], lat_span * VALIDATION_CONFIG["coord_jump_ratio"])
        jump_indices = []
        for i in range(1, len(valid_lat)):
            diff = abs(valid_lat[i][0] - valid_lat[i - 1][0])
            if diff > threshold:
                jump_indices.append(valid_lat[i][1])
        if jump_indices:
            issues.append({"rule": "lat_jump", "level": VALIDATION_RULES["lat_jump"]["level"],
                            "desc": VALIDATION_RULES["lat_jump"]["desc"],
                            "details": f"纬度跳变(跨度{lat_span:.4f}°)", "point_index": jump_indices})
    swapped_indices = []
    for i, p in enumerate(points):
        lon = p.get("lon")
        lat = p.get("lat")
        if lon is not None and lat is not None and lon != 0 and lat != 0:
            if lat_range[0] <= lon <= lat_range[1] and lon_range[0] <= lat <= lon_range[1]:
                swapped_indices.append(i)
    if swapped_indices:
        issues.append({"rule": "lonlat_swapped", "level": VALIDATION_RULES["lonlat_swapped"]["level"],
                        "desc": VALIDATION_RULES["lonlat_swapped"]["desc"],
                        "details": f"经纬度可能互换", "point_index": swapped_indices})
    coord_missing_indices = []
    for i, p in enumerate(points):
        if p.get("distance") is not None and p.get("elevation") is not None:
            lon = p.get("lon")
            lat = p.get("lat")
            if lon is None or lat is None or lon == 0 or lat == 0:
                coord_missing_indices.append(i)
    if coord_missing_indices and not is_chengtu:
        issues.append({"rule": "coord_missing", "level": VALIDATION_RULES["coord_missing"]["level"],
                        "desc": VALIDATION_RULES["coord_missing"]["desc"],
                        "details": f"有起点距高程但缺经纬度", "point_index": coord_missing_indices})
    valid_coords = [(p.get("lon"), p.get("lat")) for p in points if
                    p.get("lon") and p.get("lat") and p.get("lon") != 0 and p.get("lat") != 0]
    if len(valid_coords) >= 2:
        first_coord = valid_coords[0]
        all_same = all(c == first_coord for c in valid_coords)
        if all_same:
            issues.append({"rule": "all_same_coord", "level": VALIDATION_RULES["all_same_coord"]["level"],
                            "desc": VALIDATION_RULES["all_same_coord"]["desc"],
                            "details": f"所有{len(valid_coords)}个点的经纬度完全相同",
                            "point_index": list(range(len(points)))})
    if len(valid_coords) >= 3:
        valid_points_data = []
        valid_points_indices = []
        for i, p in enumerate(points):
            lon = p.get("lon")
            lat = p.get("lat")
            dist = p.get("distance")
            seq = p.get("seq")
            if lon and lat and lon != 0 and lat != 0 and dist is not None:
                valid_points_data.append(
                    {"distance": dist, "lon": lon, "lat": lat, "seq": seq if seq is not None else i + 1})
                valid_points_indices.append(i)
        if len(valid_points_data) >= 3:
            df = pd.DataFrame(valid_points_data)
            checker = SpatialContinuityChecker(df)
            flagged_indices = checker.check_linear_fit_deviation(threshold=VALIDATION_CONFIG["linear_fit_threshold"])
            if flagged_indices:
                deviation_indices = [valid_points_indices[i] for i in flagged_indices if i < len(valid_points_indices)]
                if deviation_indices:
                    L = np.max(df['distance']) - np.min(df['distance'])
                    issues.append({"rule": "linear_fit_deviation",
                                    "level": VALIDATION_RULES["linear_fit_deviation"]["level"],
                                    "desc": VALIDATION_RULES["linear_fit_deviation"]["desc"],
                                    "details": f"偏离PCA拟合直线超过断面长度{L:.1f}m的{VALIDATION_CONFIG['linear_fit_threshold'] * 100}%",
                                    "point_index": deviation_indices})
            direction_anomalies = checker.check_direction_consistency(reversal_threshold=90)
            for anomaly in direction_anomalies:
                idx = anomaly.get('point_index', 0)
                if idx < len(valid_points_indices):
                    point_idx = valid_points_indices[idx]
                    issues.append({"rule": "direction_reversal",
                                    "level": VALIDATION_RULES["direction_reversal"]["level"],
                                    "desc": VALIDATION_RULES["direction_reversal"]["desc"],
                                    "details": f"第{point_idx + 1}个点连线方向折返，{anomaly.get('desc', '')}",
                                    "point_index": [point_idx]})
            adjacent_anomalies = checker.check_adjacent_segment_angle(sharp_turn_threshold=90)
            for anomaly in adjacent_anomalies:
                idx = anomaly.get('point_index', 0)
                if idx < len(valid_points_indices):
                    point_idx = valid_points_indices[idx]
                    issues.append({"rule": "adjacent_reversal",
                                    "level": VALIDATION_RULES["adjacent_reversal"]["level"],
                                    "desc": VALIDATION_RULES["adjacent_reversal"]["desc"],
                                    "details": f"第{point_idx + 1}个点相邻线段折返，{anomaly.get('desc', '')}",
                                    "point_index": [point_idx]})
    return issues


class SectionChartService:
    def __init__(self):
        _ensure_db_dir()
        self._init_db()

    def _init_db(self):
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute('''CREATE TABLE IF NOT EXISTS sections (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            section_key TEXT UNIQUE,
            name TEXT,
            category TEXT,
            location TEXT,
            district_code TEXT,
            river_code TEXT,
            is_control_section TEXT,
            section_shape TEXT,
            is_cross_county TEXT,
            river_bed_material TEXT,
            survey_method TEXT,
            base_lon REAL,
            base_lat REAL,
            base_elevation REAL,
            azimuth REAL,
            hmz REAL,
            czz REAL,
            anomaly INTEGER DEFAULT 0,
            anomaly_details TEXT,
            validation_error INTEGER DEFAULT 0,
            validation_warning INTEGER DEFAULT 0,
            validation_issues TEXT,
            source_file TEXT,
            file_name TEXT,
            file_path TEXT,
            sheet_name TEXT,
            table_type TEXT,
            created_at TEXT
        )''')
        c.execute("PRAGMA table_info(sections)")
        columns = [col[1] for col in c.fetchall()]
        if "file_name" not in columns:
            c.execute("ALTER TABLE sections ADD COLUMN file_name TEXT")
        if "file_path" not in columns:
            c.execute("ALTER TABLE sections ADD COLUMN file_path TEXT")
        if "table_type" not in columns:
            c.execute("ALTER TABLE sections ADD COLUMN table_type TEXT")
        c.execute('''CREATE TABLE IF NOT EXISTS section_points (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            section_key TEXT,
            seq REAL,
            feature_desc TEXT,
            distance REAL,
            elevation REAL,
            lon REAL,
            lat REAL,
            roughness TEXT,
            is_feature INTEGER DEFAULT 0,
            FOREIGN KEY (section_key) REFERENCES sections(section_key)
        )''')
        c.execute('''CREATE INDEX IF NOT EXISTS idx_sections_category ON sections(category)''')
        c.execute('''CREATE INDEX IF NOT EXISTS idx_sections_anomaly ON sections(anomaly)''')
        c.execute('''CREATE INDEX IF NOT EXISTS idx_sections_validation_error ON sections(validation_error)''')
        c.execute('''CREATE INDEX IF NOT EXISTS idx_points_section_key ON section_points(section_key)''')
        conn.commit()
        conn.close()

    def clear_data(self):
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute("DELETE FROM section_points")
        c.execute("DELETE FROM sections")
        conn.commit()
        conn.close()

    def load_from_directory(self, base_dir: str, progress_callback=None) -> Dict:
        self.clear_data()
        all_files = []
        for root, dirs, files in os.walk(base_dir):
            for f in files:
                if f.endswith('.xlsx') and not f.startswith('~'):
                    all_files.append(os.path.join(root, f))
        total = len(all_files)
        all_sections = {}
        for idx, filepath in enumerate(all_files):
            if progress_callback:
                progress_callback(idx + 1, total, os.path.basename(filepath))
            category = os.path.basename(os.path.dirname(filepath))
            sections = self._read_excel_file(filepath, category)
            all_sections.update(sections)
        self._validate_and_store(all_sections)
        return {
            "total_files": total,
            "total_sections": len(all_sections),
            "total_points": sum(len(s["points"]) for s in all_sections.values()),
            "anomaly_count": sum(1 for s in all_sections.values() if s.get("anomaly")),
            "validation_error_count": sum(1 for s in all_sections.values() if s.get("validation_error")),
        }

    def _read_excel_file(self, filepath: str, category: str) -> Dict:
        sections = {}
        basename = os.path.basename(filepath)
        file_path = os.path.dirname(filepath)
        wb = None
        try:
            wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(min_row=1, max_row=10, values_only=True))
                if len(rows) < 3:
                    continue
                is_guifan = False
                for row in rows:
                    row_str = str(row[0]) if row[0] else ""
                    if "沟道横断面测量成果表" in row_str or "所在位置" in row_str:
                        is_guifan = True
                        break
                if is_guifan:
                    sec = self._parse_guifan_sheet(ws, filepath, sheet_name, category, basename, file_path)
                    if sec:
                        sections[sec["section_key"]] = sec
                else:
                    sec = self._parse_chengtu_sheet(ws, filepath, sheet_name, category, basename, file_path)
                    if sec:
                        sections[sec["section_key"]] = sec
        except Exception as e:
            print(f"读取失败: {filepath}, 错误: {e}")
        finally:
            if wb:
                wb.close()
        return sections

    def _parse_guifan_sheet(self, ws, filepath: str, sheet_name: str, category: str, basename: str, file_path: str) -> Optional[Dict]:
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 9:
            return None
        location = str(rows[1][1]) if rows[1][1] else ""
        district_code = str(rows[1][4]) if rows[1][4] else ""
        river_code = str(rows[2][1]) if rows[2][1] else ""
        is_control = str(rows[2][4]) if rows[2][4] else ""
        section_shape = str(rows[3][1]) if rows[3][1] else ""
        is_cross_county = str(rows[3][4]) if rows[3][4] else ""
        river_bed = str(rows[4][1]) if rows[4][1] else ""
        survey_method = str(rows[4][4]) if rows[4][4] else ""
        base_lon = safe_float(rows[5][1])
        base_lat = safe_float(rows[5][4])
        base_elev = safe_float(rows[6][1])
        azimuth = safe_float(rows[6][4])
        hmz = safe_float(rows[7][1])
        czz = safe_float(rows[7][4])
        section_name = f"{location}_{sheet_name}"
        section_key = f"{category}|{basename}|{sheet_name}"
        points = []
        for row in rows[9:]:
            if row[0] is None and row[2] is None:
                break
            seq = row[0]
            feature_desc = str(row[1]) if row[1] else ""
            dist = safe_float(row[2])
            elev = safe_float(row[3])
            lon = safe_float(row[4])
            lat = safe_float(row[5])
            roughness = str(row[6]) if row[6] else ""
            is_feature = any(kw in feature_desc for kw in FEATURE_KEYWORDS) if feature_desc else False
            if dist is None or elev is None:
                continue
            points.append({
                "seq": int(seq) if seq is not None else len(points) + 1,
                "feature_desc": feature_desc,
                "distance": round(dist, 3),
                "elevation": round(elev, 3),
                "lon": lon,
                "lat": lat,
                "roughness": roughness,
                "isFeature": is_feature,
            })
        return {
            "section_key": section_key,
            "name": section_name,
            "category": category,
            "location": location,
            "district_code": district_code,
            "river_code": river_code,
            "is_control_section": is_control,
            "section_shape": section_shape,
            "is_cross_county": is_cross_county,
            "river_bed_material": river_bed,
            "survey_method": survey_method,
            "base_lon": base_lon,
            "base_lat": base_lat,
            "base_elevation": base_elev,
            "azimuth": azimuth,
            "hmz": hmz,
            "czz": czz,
            "points": points,
            "source_file": filepath,
            "file_name": basename,
            "file_path": file_path,
            "sheet_name": sheet_name,
            "table_type": "guifan",
        }

    def _parse_chengtu_sheet(self, ws, filepath: str, sheet_name: str, category: str, basename: str, file_path: str) -> Optional[Dict]:
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 3:
            return None
        section_name_raw = str(rows[0][1]) if rows[0][1] else sheet_name
        name_from_file = basename.split("_")[-1].replace(".xlsx", "") if "_" in basename else ""
        section_name = f"{name_from_file}_{section_name_raw}" if name_from_file else section_name_raw
        section_key = f"{category}|{basename}|{sheet_name}"
        points = []
        for row in rows[2:]:
            if row[0] is None and row[1] is None:
                break
            dist = safe_float(row[0])
            elev = safe_float(row[1])
            if dist is None or elev is None:
                continue
            points.append({
                "seq": len(points) + 1,
                "feature_desc": "",
                "distance": round(dist, 3),
                "elevation": round(elev, 3),
                "lon": None,
                "lat": None,
                "roughness": "",
                "isFeature": False,
            })
        return {
            "section_key": section_key,
            "name": section_name,
            "category": category,
            "location": name_from_file,
            "district_code": "",
            "river_code": "",
            "is_control_section": "",
            "section_shape": "",
            "is_cross_county": "",
            "river_bed_material": "",
            "survey_method": "",
            "base_lon": None,
            "base_lat": None,
            "base_elevation": None,
            "azimuth": None,
            "hmz": None,
            "czz": None,
            "points": points,
            "source_file": filepath,
            "file_name": basename,
            "file_path": file_path,
            "sheet_name": sheet_name,
            "table_type": "chengtu",
        }

    def _validate_and_store(self, all_sections: Dict):
        for key, sec in all_sections.items():
            has_coords = any(p.get("lon") and p.get("lat") for p in sec.get("points", []))
            if has_coords:
                is_anomaly, anomaly_details = detect_anomaly(sec)
                sec["anomaly"] = is_anomaly
                sec["anomaly_details"] = anomaly_details
            else:
                sec["anomaly"] = False
                sec["anomaly_details"] = []
            issues = validate_section_data(sec)
            sec["validation_issues"] = issues
            sec["validation_error"] = any(i["level"] == "error" for i in issues)
            sec["validation_warning"] = any(i["level"] == "warning" for i in issues)
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        for key, sec in all_sections.items():
            c.execute("""INSERT OR REPLACE INTO sections 
                (section_key, name, category, location, district_code, river_code,
                 is_control_section, section_shape, is_cross_county, river_bed_material,
                 survey_method, base_lon, base_lat, base_elevation, azimuth, hmz, czz,
                 anomaly, anomaly_details, validation_error, validation_warning,
                 validation_issues, source_file, file_name, file_path, sheet_name, table_type, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                      (key, sec["name"], sec["category"], sec["location"], sec["district_code"],
                       sec["river_code"], sec["is_control_section"], sec["section_shape"],
                       sec["is_cross_county"], sec["river_bed_material"], sec["survey_method"],
                       sec["base_lon"], sec["base_lat"], sec["base_elevation"], sec["azimuth"],
                       sec["hmz"], sec["czz"],
                       1 if sec["anomaly"] else 0,
                       json.dumps(sec["anomaly_details"], ensure_ascii=False),
                       1 if sec["validation_error"] else 0,
                       1 if sec["validation_warning"] else 0,
                       json.dumps(sec["validation_issues"], ensure_ascii=False),
                       sec["source_file"], sec.get("file_name", ""), sec.get("file_path", ""),
                       sec["sheet_name"], sec.get("table_type", ""),
                       datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
            c.execute("DELETE FROM section_points WHERE section_key = ?", (key,))
            for p in sec["points"]:
                c.execute("""INSERT INTO section_points 
                    (section_key, seq, feature_desc, distance, elevation, lon, lat, roughness, is_feature)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                          (key, p["seq"], p["feature_desc"], p["distance"], p["elevation"],
                           p.get("lon"), p.get("lat"), p.get("roughness", ""),
                           1 if p.get("isFeature") else 0))
        conn.commit()
        conn.close()

    def get_tree_data(self) -> Dict:
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        c = conn.cursor()
        c.execute("SELECT file_path, file_name, section_key, name, anomaly, validation_error, validation_warning, table_type FROM sections ORDER BY file_path, file_name, name")
        rows = c.fetchall()
        tree = {}
        for row in rows:
            file_path = row["file_path"] or "未知路径"
            file_name = row["file_name"] or "未知文件"
            if file_path not in tree:
                tree[file_path] = {}
            if file_name not in tree[file_path]:
                tree[file_path][file_name] = []
            c2 = conn.cursor()
            c2.execute("SELECT COUNT(*) as cnt FROM section_points WHERE section_key = ?", (row["section_key"],))
            pt_count = c2.fetchone()["cnt"]
            c2.execute("SELECT COUNT(*) as cnt FROM section_points WHERE section_key = ? AND is_feature = 1",
                       (row["section_key"],))
            feat_count = c2.fetchone()["cnt"]
            tree[file_path][file_name].append({
                "key": row["section_key"],
                "name": row["name"],
                "points": pt_count,
                "features": feat_count,
                "anomaly": bool(row["anomaly"]),
                "validation_error": bool(row["validation_error"]),
                "validation_warning": bool(row["validation_warning"]),
                "table_type": row["table_type"] or "",
            })
        conn.close()
        return tree

    def get_section_detail(self, section_key: str) -> Optional[Dict]:
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        c = conn.cursor()
        c.execute("SELECT * FROM sections WHERE section_key = ?", (section_key,))
        row = c.fetchone()
        if not row:
            conn.close()
            return None
        sec = dict(row)
        sec["anomaly"] = bool(sec["anomaly"])
        sec["validation_error"] = bool(sec["validation_error"])
        sec["validation_warning"] = bool(sec["validation_warning"])
        sec["anomaly_details"] = json.loads(sec["anomaly_details"] or "[]")
        sec["validation_issues"] = json.loads(sec["validation_issues"] or "[]")
        c.execute("SELECT * FROM section_points WHERE section_key = ? ORDER BY seq", (section_key,))
        points = []
        for pr in c.fetchall():
            p = dict(pr)
            p["isFeature"] = bool(p["is_feature"])
            points.append(p)
        sec["points"] = points
        conn.close()
        return sec

    def recalculate_sections(self):
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        c = conn.cursor()
        c.execute("SELECT section_key FROM sections")
        keys = [row["section_key"] for row in c.fetchall()]
        for key in keys:
            c2 = conn.cursor()
            c2.execute("SELECT * FROM section_points WHERE section_key = ? ORDER BY seq", (key,))
            points = []
            for pr in c2.fetchall():
                p = dict(pr)
                feature_desc = p.get("feature_desc", "") or ""
                p["isFeature"] = any(kw in feature_desc for kw in FEATURE_KEYWORDS) if feature_desc else False
                points.append(p)
            sec = {"points": points}
            has_coords = any(p.get("lon") and p.get("lat") for p in points)
            if has_coords:
                is_anomaly, anomaly_details = detect_anomaly(sec)
            else:
                is_anomaly, anomaly_details = False, []
            issues = validate_section_data(sec)
            validation_error = any(i["level"] == "error" for i in issues)
            validation_warning = any(i["level"] == "warning" for i in issues)
            c.execute("UPDATE sections SET anomaly=?, anomaly_details=?, validation_error=?, validation_warning=?, validation_issues=? WHERE section_key=?",
                      (1 if is_anomaly else 0,
                       json.dumps(anomaly_details, ensure_ascii=False),
                       1 if validation_error else 0,
                       1 if validation_warning else 0,
                       json.dumps(issues, ensure_ascii=False),
                       key))
            for p in points:
                c.execute("UPDATE section_points SET is_feature=? WHERE id=?", (1 if p["isFeature"] else 0, p["id"]))
        conn.commit()
        conn.close()

    def get_stats(self) -> Dict:
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute("SELECT COUNT(*) FROM sections")
        total_sections = c.fetchone()[0]
        c.execute("SELECT COUNT(*) FROM section_points")
        total_points = c.fetchone()[0]
        c.execute("SELECT COUNT(*) FROM sections WHERE anomaly = 1")
        anomaly_count = c.fetchone()[0]
        c.execute("SELECT COUNT(*) FROM sections WHERE validation_error = 1")
        error_count = c.fetchone()[0]
        c.execute("SELECT COUNT(DISTINCT category) FROM sections")
        category_count = c.fetchone()[0]
        c.execute("SELECT COUNT(*) FROM section_points WHERE lon IS NOT NULL AND lon != 0")
        coords_count = c.fetchone()[0]
        conn.close()
        return {
            "total_sections": total_sections,
            "total_points": total_points,
            "anomaly_count": anomaly_count,
            "validation_error_count": error_count,
            "category_count": category_count,
            "coords_count": coords_count,
        }

    def get_sections_by_filter(self, category: str = None, anomaly_only: bool = False,
                                error_only: bool = False, search: str = "") -> List[Dict]:
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        c = conn.cursor()
        query = "SELECT s.section_key, s.name, s.category, s.location, s.anomaly, s.validation_error, s.validation_warning, COUNT(p.id) as point_count FROM sections s LEFT JOIN section_points p ON s.section_key = p.section_key WHERE 1=1"
        params = []
        if category:
            query += " AND s.category = ?"
            params.append(category)
        if anomaly_only:
            query += " AND s.anomaly = 1"
        if error_only:
            query += " AND s.validation_error = 1"
        if search:
            query += " AND (s.name LIKE ? OR s.location LIKE ? OR s.sheet_name LIKE ?)"
            params.extend([f"%{search}%", f"%{search}%", f"%{search}%"])
        query += " GROUP BY s.section_key ORDER BY s.category, s.location, s.name"
        c.execute(query, params)
        results = []
        for row in c.fetchall():
            results.append(dict(row))
        conn.close()
        return results

    def export_validation_report(self, output_path: str) -> str:
        self.recalculate_sections()
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        c = conn.cursor()
        c.execute("SELECT * FROM sections WHERE anomaly = 1 OR validation_error = 1 OR validation_warning = 1 ORDER BY category, name")
        rows = c.fetchall()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "异常断面汇总"
        headers = ["断面名称", "分类", "位置", "异常", "校验错误", "校验警告", "问题详情"]
        ws.append(headers)
        for row in rows:
            issues = json.loads(row["validation_issues"] or "[]")
            issue_strs = []
            for iss in issues:
                issue_strs.append(f"[{iss['level']}]{iss['desc']}: {iss['details']}")
            anomaly_details = json.loads(row["anomaly_details"] or "[]")
            for ad in anomaly_details:
                issue_strs.append(f"[异常]{ad.get('desc', ad.get('type', ''))}")
            ws.append([
                row["name"], row["category"], row["location"],
                "是" if row["anomaly"] else "否",
                "是" if row["validation_error"] else "否",
                "是" if row["validation_warning"] else "否",
                "\n".join(issue_strs) if issue_strs else ""
            ])
        for col in range(1, 8):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 25
        wb.save(output_path)
        wb.close()
        conn.close()
        return output_path

    def get_anomaly_sections(self) -> Dict:
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        c = conn.cursor()
        c.execute("SELECT * FROM sections WHERE anomaly = 1 OR validation_error = 1")
        rows = c.fetchall()
        sections = {}
        for row in rows:
            sec = dict(row)
            sec["anomaly"] = bool(sec["anomaly"])
            sec["validation_error"] = bool(sec["validation_error"])
            sec["validation_warning"] = bool(sec["validation_warning"])
            sec["anomaly_details"] = json.loads(sec["anomaly_details"] or "[]")
            sec["validation_issues"] = json.loads(sec["validation_issues"] or "[]")
            c2 = conn.cursor()
            c2.execute("SELECT * FROM section_points WHERE section_key = ? ORDER BY seq", (sec["section_key"],))
            points = []
            for pr in c2.fetchall():
                p = dict(pr)
                p["isFeature"] = bool(p["is_feature"])
                p["feature"] = p.get("feature_desc", "")
                points.append(p)
            sec["points"] = points
            sec["river"] = sec.get("category", "")
            sec["district"] = sec.get("location", "")
            sections[sec["section_key"]] = sec
        conn.close()
        return sections

    def export_anomaly_html(self, output_path: str) -> str:
        from services.generate_section_chart_with_coords_青海 import generate_html, build_tree_data

        self.recalculate_sections()
        sections = self.get_anomaly_sections()
        if not sections:
            return ""

        for key, sec in sections.items():
            sec["river"] = sec.get("category", "")
            sec["district"] = sec.get("file_name", "")

        tree = build_tree_data(sections)
        html = generate_html(sections, tree, "异常断面汇总", default_filter_anomaly=True)

        with open(output_path, "w", encoding="utf-8") as f:
            f.write(html)
        return output_path

    def get_all_sections(self) -> Dict:
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        c = conn.cursor()
        c.execute("SELECT * FROM sections")
        rows = c.fetchall()
        sections = {}
        for row in rows:
            sec = dict(row)
            sec["anomaly"] = bool(sec["anomaly"])
            sec["validation_error"] = bool(sec["validation_error"])
            sec["validation_warning"] = bool(sec["validation_warning"])
            sec["anomaly_details"] = json.loads(sec["anomaly_details"] or "[]")
            sec["validation_issues"] = json.loads(sec["validation_issues"] or "[]")
            c2 = conn.cursor()
            c2.execute("SELECT * FROM section_points WHERE section_key = ? ORDER BY seq", (sec["section_key"],))
            points = []
            for pr in c2.fetchall():
                p = dict(pr)
                p["isFeature"] = bool(p["is_feature"])
                p["feature"] = p.get("feature_desc", "")
                points.append(p)
            sec["points"] = points
            sec["river"] = sec.get("category", "")
            sec["district"] = sec.get("location", "")
            sections[sec["section_key"]] = sec
        conn.close()
        return sections

    def export_all_html(self, output_path: str) -> str:
        from services.generate_section_chart_with_coords_青海 import generate_html, build_tree_data

        self.recalculate_sections()
        sections = self.get_all_sections()
        if not sections:
            return ""

        for key, sec in sections.items():
            sec["river"] = sec.get("category", "")
            sec["district"] = sec.get("file_name", "")

        tree = build_tree_data(sections)
        html = generate_html(sections, tree, "全部断面成图", default_filter_anomaly=False)

        with open(output_path, "w", encoding="utf-8") as f:
            f.write(html)
        return output_path
