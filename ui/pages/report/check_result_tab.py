# -*- coding: utf-8 -*-
import os
from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QPushButton,
    QLabel, QTableWidget, QTableWidgetItem,
    QHeaderView, QAbstractItemView, QComboBox,
    QFileDialog, QMessageBox
)
from PySide6.QtCore import Qt
from PySide6.QtGui import QColor


class CheckResultTab(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.check_errors = []
        self.check_original_errors = []
        self.report_data = None
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout()
        layout.setContentsMargins(0, 4, 0, 4)

        self.stats_label = QLabel('点击"报表成果检查"开始检查...')
        self.stats_label.setObjectName("statsLabel")
        layout.addWidget(self.stats_label)

        filter_layout = QHBoxLayout()
        filter_layout.setSpacing(10)

        filter_label = QLabel("筛选:")
        filter_label.setObjectName("boldLabel")

        self.table_combo = QComboBox()
        self.table_combo.addItems(["全部", "附表1", "附表2", "附表3"])
        self.table_combo.setFixedWidth(80)
        self.table_combo.currentTextChanged.connect(self.apply_filter)

        self.type_combo = QComboBox()
        self.type_combo.addItems(["全部", "格式错误", "一致性错误"])
        self.type_combo.setFixedWidth(100)
        self.type_combo.currentTextChanged.connect(self.apply_filter)

        filter_layout.addWidget(filter_label)
        filter_layout.addWidget(QLabel("表名:"))
        filter_layout.addWidget(self.table_combo)
        filter_layout.addWidget(QLabel("错误类型:"))
        filter_layout.addWidget(self.type_combo)
        filter_layout.addStretch()
        layout.addLayout(filter_layout)

        self.table = QTableWidget()
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setAlternatingRowColors(False)
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.verticalHeader().setVisible(False)
        layout.addWidget(self.table)

        export_btn = QPushButton("导出检查报告")
        export_btn.setFixedWidth(120)
        export_btn.clicked.connect(self.export_report)
        layout.addWidget(export_btn, alignment=Qt.AlignRight)

        self.setLayout(layout)

    def set_report_data(self, data):
        self.report_data = data

    def set_check_errors(self, errors):
        self.check_errors = errors.copy()
        self.check_original_errors = errors.copy()

        format_count = sum(1 for e in self.check_errors if e['错误类型'] == '格式错误')
        consist_count = sum(1 for e in self.check_errors if e['错误类型'] == '一致性错误')
        self.stats_label.setText(
            f"共发现 {len(self.check_errors)} 个问题 | 格式错误: {format_count} | 一致性错误: {consist_count}"
        )
        self.display_results(self.check_errors)

    def display_results(self, errors):
        headers = ['序号', '表名', '字段名', '错误类型', '错误描述', '当前值']
        self.table.clear()
        self.table.setColumnCount(len(headers))
        self.table.setHorizontalHeaderLabels(headers)
        self.table.setRowCount(len(errors))

        for row, err in enumerate(errors):
            for col, header in enumerate(headers):
                value = err.get(header, '')
                item = QTableWidgetItem(str(value))
                if header == '错误类型':
                    if value == '格式错误':
                        item.setForeground(QColor('#e67e22'))
                    elif value == '一致性错误':
                        item.setForeground(QColor('#e74c3c'))
                self.table.setItem(row, col, item)

        self.table.resizeColumnsToContents()

    def apply_filter(self):
        table_filter = self.table_combo.currentText()
        type_filter = self.type_combo.currentText()

        filtered = []
        for err in self.check_original_errors:
            if table_filter != "全部" and err['表名'] != table_filter:
                continue
            if type_filter != "全部" and err['错误类型'] != type_filter:
                continue
            filtered.append(err)

        self.check_errors = filtered
        self.display_results(filtered)

    def clear_data(self):
        self.check_errors = []
        self.check_original_errors = []
        self.report_data = None
        self.table.clear()
        self.table.setRowCount(0)
        self.stats_label.setText('点击"报表成果检查"开始检查...')
        self.table_combo.setCurrentText("全部")
        self.type_combo.setCurrentText("全部")

    def export_report(self):
        if not self.report_data:
            QMessageBox.warning(self, "警告", "没有数据可导出")
            return

        file, _ = QFileDialog.getSaveFileName(
            self, "导出检查报告", "数据检查报告.xlsx", "Excel Files (*.xlsx)"
        )

        if file:
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

                wb = Workbook()

                header_font = Font(bold=True)
                header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
                error_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

                error_index = {'附表1': {}, '附表2': {}, '附表3': {}}
                for err in self.check_errors:
                    table = err['表名']
                    row_num = err['序号']
                    field = err['字段名']
                    desc = err['错误描述']
                    if table not in error_index:
                        error_index[table] = {}
                    if row_num not in error_index[table]:
                        error_index[table][row_num] = []
                    error_index[table][row_num].append((field, desc))

                ws1 = wb.active
                ws1.title = '附表1-防治对象名录'

                records = self.report_data['fubiao1']['records']
                merge_cells = self.report_data['fubiao1'].get('merge_info', {}).get('merge_cells', [])

                header_row1 = [
                    '序号', '1.县（区、市、旗）名称', '2.县（区、市、旗）代码',
                    '3.乡镇名称', '4.乡镇代码', '5.名称', '6.代码',
                    '7.类型', '8.人口', '9.河流名称', '10.河流代码',
                    '风险隐患要素类别', '', '', '', '', '', '', '', '', '', '', '',
                    '风险隐患影响类型', '', '', '', '',
                    '28.备注', '检查结果'
                ]
                header_row2 = [
                    '', '', '', '', '', '', '', '', '', '', '',
                    '跨沟道路、桥涵', '',
                    '塘（堰）坝', '',
                    '多支齐汇', '',
                    '局地河势与微地形', '', '', '', '',
                    '22.沟滩占地',
                    '23.溃决', '24.壅水', '25.顶托', '26.改道', '27.漫流',
                    '', ''
                ]
                header_row3 = [
                    '', '', '', '', '', '', '', '', '', '', '',
                    '11.名称', '12.编码',
                    '13.名称', '14.编码',
                    '15.河流名称', '16.河流代码',
                    '17.束窄', '18.急弯', '19.低洼地', '20.临河滑坡', '21.泥石流',
                    '',
                    '', '', '', '', '',
                    '', ''
                ]
                data_headers = [
                    '序号', '1.县（区、市、旗）名称', '2.县（区、市、旗）代码',
                    '3.乡镇名称', '4.乡镇代码', '5.名称', '6.代码',
                    '7.类型', '8.人口', '9.河流名称', '10.河流代码',
                    '11.名称', '12.编码', '13.名称', '14.编码',
                    '15.河流名称', '16.河流代码',
                    '17.束窄', '18.急弯', '19.低洼地', '20.临河滑坡', '21.泥石流',
                    '22.沟滩占地', '23.溃决', '24.壅水', '25.顶托', '26.改道', '27.漫流',
                    '28.备注'
                ]

                ws1.append(header_row1)
                ws1.append(header_row2)
                ws1.append(header_row3)

                for i, record in enumerate(records):
                    row_num = i + 1
                    row_data = []
                    for header in data_headers:
                        val = record.get(header, '')
                        if val == '☑' or val == 'þ':
                            val = 'R'
                        elif val == '☒' or val == 'ý':
                            val = 'S'
                        row_data.append(val)

                    errors = error_index.get('附表1', {}).get(row_num, [])
                    check_result = '；'.join([f"{f}:{d}" for f, d in errors]) if errors else ''
                    row_data.append(check_result)

                    ws1.append(row_data)

                    if errors:
                        excel_row = i + 4
                        for field, desc in errors:
                            if field == '县代码':
                                ws1.cell(row=excel_row, column=3).fill = error_fill
                            elif field == '乡镇代码':
                                ws1.cell(row=excel_row, column=5).fill = error_fill
                            elif field == '跨沟道路桥涵':
                                ws1.cell(row=excel_row, column=12).fill = error_fill
                                ws1.cell(row=excel_row, column=13).fill = error_fill
                        if check_result:
                            ws1.cell(row=excel_row, column=30).fill = error_fill

                for col in range(1, 12):
                    ws1.merge_cells(start_row=1, start_column=col, end_row=3, end_column=col)
                ws1.merge_cells(start_row=1, start_column=12, end_row=1, end_column=23)
                ws1.merge_cells(start_row=1, start_column=24, end_row=1, end_column=28)
                ws1.merge_cells(start_row=1, start_column=29, end_row=3, end_column=29)
                ws1.merge_cells(start_row=1, start_column=30, end_row=3, end_column=30)
                ws1.merge_cells(start_row=2, start_column=12, end_row=2, end_column=13)
                ws1.merge_cells(start_row=2, start_column=14, end_row=2, end_column=15)
                ws1.merge_cells(start_row=2, start_column=16, end_row=2, end_column=17)
                ws1.merge_cells(start_row=2, start_column=18, end_row=2, end_column=22)
                for col in range(23, 29):
                    ws1.merge_cells(start_row=2, start_column=col, end_row=3, end_column=col)

                for merge_info in merge_cells:
                    start_row, start_col, row_span, col_span = merge_info
                    ws1.merge_cells(
                        start_row=start_row + 4,
                        start_column=start_col + 1,
                        end_row=start_row + 4 + row_span - 1,
                        end_column=start_col + col_span
                    )

                for row in ws1.iter_rows(min_row=1, max_row=3, max_col=30):
                    for cell in row:
                        cell.font = header_font
                        cell.alignment = header_alignment
                        cell.border = thin_border

                wingdings2_font = Font(name='Wingdings 2')
                checkbox_cols = list(range(18, 29))
                for row in ws1.iter_rows(min_row=4, max_row=3 + len(records), max_col=30):
                    for cell in row:
                        cell.border = thin_border
                        cell.alignment = Alignment(vertical='center')
                        if cell.column in checkbox_cols:
                            cell.font = wingdings2_font
                            cell.alignment = Alignment(horizontal='center', vertical='center')

                ws2 = wb.create_sheet('附表2-跨沟道路桥涵')
                records2 = self.report_data['fubiao2']['records']
                headers2 = self.report_data['fubiao2']['headers'] + ['检查结果']

                ws2.append(headers2)
                for i, record in enumerate(records2):
                    row_num = i + 1
                    row_data = [record.get(h, '') for h in self.report_data['fubiao2']['headers']]

                    errors = error_index.get('附表2', {}).get(row_num, [])
                    check_result = '；'.join([f"{f}:{d}" for f, d in errors]) if errors else ''
                    row_data.append(check_result)

                    ws2.append(row_data)

                    if errors:
                        excel_row = i + 2
                        error_fields = [f for f, d in errors]
                        for col_idx, h in enumerate(self.report_data['fubiao2']['headers'], 1):
                            for ef in error_fields:
                                if ef in h or h in ef:
                                    ws2.cell(row=excel_row, column=col_idx).fill = error_fill
                                    break
                        if check_result:
                            ws2.cell(row=excel_row, column=len(headers2)).fill = error_fill

                for row in ws2.iter_rows(min_row=1, max_row=1 + len(records2), max_col=len(headers2)):
                    for cell in row:
                        cell.border = thin_border
                        if cell.row == 1:
                            cell.font = header_font
                            cell.alignment = header_alignment

                ws3 = wb.create_sheet('附表3-沟滩占地')
                records3 = self.report_data['fubiao3']['records']
                headers3 = self.report_data['fubiao3']['headers'] + ['检查结果']

                ws3.append(headers3)
                for i, record in enumerate(records3):
                    row_num = i + 1
                    row_data = [record.get(h, '') for h in self.report_data['fubiao3']['headers']]

                    errors = error_index.get('附表3', {}).get(row_num, [])
                    check_result = '；'.join([f"{f}:{d}" for f, d in errors]) if errors else ''
                    row_data.append(check_result)

                    ws3.append(row_data)

                    if errors:
                        excel_row = i + 2
                        error_fields = [f for f, d in errors]
                        for col_idx, h in enumerate(self.report_data['fubiao3']['headers'], 1):
                            for ef in error_fields:
                                if ef in h or h in ef:
                                    ws3.cell(row=excel_row, column=col_idx).fill = error_fill
                                    break
                        if check_result:
                            ws3.cell(row=excel_row, column=len(headers3)).fill = error_fill

                for row in ws3.iter_rows(min_row=1, max_row=1 + len(records3), max_col=len(headers3)):
                    for cell in row:
                        cell.border = thin_border
                        if cell.row == 1:
                            cell.font = header_font
                            cell.alignment = header_alignment

                ws_check = wb.create_sheet("检查结果汇总")
                check_headers = ['序号', '表名', '字段名', '错误类型', '错误描述', '当前值']
                ws_check.append(check_headers)
                for cell in ws_check[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = thin_border
                    cell.alignment = header_alignment

                for err in self.check_errors:
                    ws_check.append([
                        err['序号'],
                        err['表名'],
                        err['字段名'],
                        err['错误类型'],
                        err['错误描述'],
                        err['当前值']
                    ])

                for row in ws_check.iter_rows(min_row=2, max_row=ws_check.max_row, max_col=len(check_headers)):
                    for cell in row:
                        cell.border = thin_border

                wb.save(file)
                QMessageBox.information(self, "完成", f"已导出到:\n{file}")

            except Exception as e:
                import traceback
                traceback.print_exc()
                QMessageBox.critical(self, "错误", f"导出失败:\n{str(e)}")
