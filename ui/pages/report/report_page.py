# -*- coding: utf-8 -*-
import os
from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QFrame,
    QLabel, QLineEdit, QFileDialog, QTabWidget, QMessageBox
)
from PySide6.QtCore import Qt
from core.report_reader import load_all_reports
from core.effects_manager import StaggerEntrance, TabFadeTransition, ButtonGlowHelper
from .fubiao1_tab import Fubiao1Tab
from .fubiao2_tab import Fubiao2Tab
from .fubiao3_tab import Fubiao3Tab
from .check_result_tab import CheckResultTab
from .validation_dialog import ValidationDialog


class ReportPage(QWidget):

    def __init__(self, parent=None):
        super().__init__(parent)
        self.report_data = None
        self.result_dialog = None
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout()
        layout.setSpacing(12)
        layout.setContentsMargins(0, 0, 0, 0)

        header_card = QFrame()
        header_card.setObjectName("pageHeader")
        header_layout = QHBoxLayout(header_card)
        header_layout.setSpacing(16)
        header_layout.setContentsMargins(20, 16, 20, 16)

        accent_bar = QFrame()
        accent_bar.setObjectName("accentBar")
        accent_bar.setFixedWidth(4)
        header_layout.addWidget(accent_bar)

        title_layout = QVBoxLayout()
        title_layout.setSpacing(4)

        page_title = QLabel("成果报表展示")
        page_title.setObjectName("sectionHeaderLg")
        title_layout.addWidget(page_title)

        page_subtitle = QLabel("加载、检查和校验空间数据报表成果")
        page_subtitle.setObjectName("pageSubtitle")
        title_layout.addWidget(page_subtitle)

        header_layout.addLayout(title_layout, 1)
        layout.addWidget(header_card)

        card = QFrame()
        card.setObjectName("card")
        group_layout = QVBoxLayout(card)
        group_layout.setSpacing(10)
        group_layout.setContentsMargins(16, 16, 16, 16)

        file_row = QHBoxLayout()
        file_row.setSpacing(10)

        folder_label = QLabel("成果报表文件夹:")
        folder_label.setObjectName("boldLabel")
        self.folder_edit = QLineEdit()
        self.folder_edit.setPlaceholderText("选择包含附表Excel文件的文件夹...")
        self.folder_edit.setReadOnly(True)

        default_folder = r"D:\github\空间数据检查桌面版-主题-design-2026\青海24示范小流域-药草沟-20260313"
        if os.path.exists(default_folder):
            self.folder_edit.setText(default_folder)
            self.load_btn = QPushButton("加载数据")
            self.load_btn.setFixedWidth(100)
            self.load_btn.clicked.connect(self.load_data)
            self.load_btn.setEnabled(True)
        else:
            self.load_btn = QPushButton("加载数据")
            self.load_btn.setFixedWidth(100)
            self.load_btn.clicked.connect(self.load_data)
            self.load_btn.setEnabled(False)

        self.browse_btn = QPushButton("浏览...")
        self.browse_btn.setFixedWidth(80)
        self.browse_btn.clicked.connect(self.select_folder)

        file_row.addWidget(folder_label)
        file_row.addWidget(self.folder_edit, 1)
        file_row.addWidget(self.browse_btn)
        group_layout.addLayout(file_row)

        btn_row = QHBoxLayout()
        btn_row.setSpacing(15)
        btn_row.addStretch()

        self.check_btn = QPushButton("报表成果检查")
        self.check_btn.setObjectName("checkBtn")
        self.check_btn.setFixedWidth(120)
        self.check_btn.clicked.connect(self.run_data_check)
        self.check_btn.setEnabled(False)

        self.export_btn = QPushButton("导出Excel")
        self.export_btn.setObjectName("exportBtn")
        self.export_btn.setFixedWidth(100)
        self.export_btn.clicked.connect(self.export_excel)
        self.export_btn.setEnabled(False)

        self.validate_btn = QPushButton("空间数据校验")
        self.validate_btn.setObjectName("validateBtn")
        self.validate_btn.setFixedWidth(120)
        self.validate_btn.clicked.connect(self.show_validation_dialog)
        self.validate_btn.setEnabled(False)

        self.clear_btn = QPushButton("清空")
        self.clear_btn.setObjectName("clearBtn")
        self.clear_btn.setFixedWidth(80)
        self.clear_btn.clicked.connect(self.clear_data)

        btn_row.addWidget(self.load_btn)
        btn_row.addWidget(self.clear_btn)
        btn_row.addSpacing(10)

        separator1 = QFrame()
        separator1.setFrameShape(QFrame.VLine)
        separator1.setFrameShadow(QFrame.Sunken)
        btn_row.addWidget(separator1)
        btn_row.addSpacing(10)

        btn_row.addWidget(self.check_btn)
        btn_row.addWidget(self.export_btn)
        btn_row.addSpacing(10)

        separator2 = QFrame()
        separator2.setFrameShape(QFrame.VLine)
        separator2.setFrameShadow(QFrame.Sunken)
        btn_row.addWidget(separator2)
        btn_row.addSpacing(10)

        btn_row.addWidget(self.validate_btn)
        btn_row.addStretch()
        group_layout.addLayout(btn_row)
        layout.addWidget(card)

        self.tabs = QTabWidget()

        self.fubiao1_tab = Fubiao1Tab()
        self.tabs.addTab(self.fubiao1_tab, "附表1-防治对象名录")

        self.fubiao2_tab = Fubiao2Tab()
        self.tabs.addTab(self.fubiao2_tab, "附表2-跨沟道路桥涵")

        self.fubiao3_tab = Fubiao3Tab()
        self.tabs.addTab(self.fubiao3_tab, "附表3-沟滩占地")

        self.check_result_tab = CheckResultTab()
        self.tabs.addTab(self.check_result_tab, "检查结果")

        layout.addWidget(self.tabs, 1)

        self.status_label = QLabel("")
        self.status_label.setObjectName("secondaryLabel")
        layout.addWidget(self.status_label)

        self.setLayout(layout)

        self._tab_fade = TabFadeTransition(self.tabs, duration=180)
        self._tab_fade.attach()

        for btn in self.findChildren(QPushButton):
            if btn.objectName() in ('loadBtn', 'checkBtn', 'validateBtn'):
                ButtonGlowHelper.install(btn)

    def showEvent(self, event):
        super().showEvent(event)
        if not hasattr(self, '_entrance_played'):
            self._entrance_played = True
            StaggerEntrance.play(self, stagger_ms=60, duration=320)

    def select_folder(self):
        folder = QFileDialog.getExistingDirectory(self, "选择成果报表文件夹")
        if folder:
            self.folder_edit.setText(folder)
            self.load_btn.setEnabled(True)

    def clear_data(self):
        self.folder_edit.clear()
        self.report_data = None

        self.fubiao1_tab.clear_data()
        self.fubiao2_tab.clear_data()
        self.fubiao3_tab.clear_data()
        self.check_result_tab.clear_data()

        self.load_btn.setEnabled(False)
        self.check_btn.setEnabled(False)
        self.export_btn.setEnabled(False)
        self.validate_btn.setEnabled(False)
        self.status_label.setText("")

    def load_data(self):
        folder = self.folder_edit.text()
        if not folder:
            QMessageBox.warning(self, "警告", "请先选择文件夹")
            return

        if not os.path.exists(folder):
            QMessageBox.warning(self, "警告", "文件夹不存在")
            return

        try:
            self.report_data = load_all_reports(folder)
        except Exception as e:
            QMessageBox.critical(self, "错误", f"加载数据失败:\n{e}")
            return

        if not self.report_data:
            QMessageBox.warning(self, "警告", "未能加载到任何数据")
            return

        if self.report_data['missing']:
            QMessageBox.warning(
                self, "警告",
                f"以下文件未找到:\n" + "\n".join(self.report_data['missing'])
            )

        self.fubiao1_tab.update_table(
            self.report_data['fubiao1']['records'],
            self.report_data['fubiao1'].get('merge_info', {}).get('merge_cells', [])
        )
        self.fubiao1_tab.set_original_data(self.report_data['fubiao1']['records'])

        self.fubiao2_tab.update_table(
            self.report_data['fubiao2']['records'],
            self.report_data['fubiao2']['headers']
        )
        self.fubiao2_tab.set_original_data(self.report_data['fubiao2']['records'])

        self.fubiao3_tab.update_table(
            self.report_data['fubiao3']['records'],
            self.report_data['fubiao3']['headers']
        )
        self.fubiao3_tab.set_original_data(self.report_data['fubiao3']['records'])

        self.check_result_tab.set_report_data(self.report_data)

        total = (
            len(self.report_data['fubiao1']['records']) +
            len(self.report_data['fubiao2']['records']) +
            len(self.report_data['fubiao3']['records'])
        )
        self.status_label.setText(f"共 {total} 条记录")
        self.load_btn.setEnabled(False)
        self.check_btn.setEnabled(total > 0)
        self.export_btn.setEnabled(total > 0)
        self.validate_btn.setEnabled(total > 0)

        QMessageBox.information(self, "完成", f"数据加载完成，共 {total} 条记录")

    def run_data_check(self):
        if not self.report_data:
            QMessageBox.warning(self, "警告", "请先加载数据")
            return

        try:
            from utils.data_checker import DataChecker
            checker = DataChecker(self.report_data)
            check_errors = checker.check_all()
        except Exception as e:
            QMessageBox.critical(self, "错误", f"数据检查失败:\n{e}")
            return

        self.check_result_tab.set_check_errors(check_errors)

    def export_excel(self):
        if not self.report_data:
            QMessageBox.warning(self, "警告", "没有可导出的数据")
            return

        file, _ = QFileDialog.getSaveFileName(
            self, "导出Excel", "成果报表.xlsx", "Excel Files (*.xlsx)"
        )

        if file:
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Alignment, Font, Border, Side
                from openpyxl.utils import get_column_letter

                wb = Workbook()

                ws1 = wb.active
                ws1.title = '附表1-防治对象名录'

                records = self.report_data['fubiao1']['records']
                merge_cells = self.report_data['fubiao1'].get('merge_info', {}).get('merge_cells', [])

                header_row1 = [
                    '序号', '1.县（区、市、旗）名称', '2.县（区、市、旗）代码',
                    '3.乡镇名称', '4.乡镇代码', '5.名称', '6.代码',
                    '7.类型', '8.人口', '9.河流名称', '10.河流代码',
                    '风险隐患要素类别', '', '', '', '', '', '', '', '', '', '', '',
                    '风险隐患影响类型', '', '', '', '',
                    '28.备注'
                ]

                header_row2 = [
                    '', '', '', '', '', '', '', '', '', '', '',
                    '跨沟道路、桥涵', '',
                    '塘（堰）坝', '',
                    '多支齐汇', '',
                    '局地河势与微地形', '', '', '', '',
                    '22.沟滩占地',
                    '23.溃决', '24.壅水', '25.顶托', '26.改道', '27.漫流',
                    ''
                ]

                header_row3 = [
                    '', '', '', '', '', '', '', '', '', '', '',
                    '11.名称', '12.编码',
                    '13.名称', '14.编码',
                    '15.河流名称', '16.河流代码',
                    '17.束窄', '18.急弯', '19.低洼地', '20.临河滑坡', '21.泥石流',
                    '',
                    '', '', '', '', '',
                    ''
                ]

                data_headers = [
                    '序号', '1.县（区、市、旗）名称', '2.县（区、市、旗）代码',
                    '3.乡镇名称', '4.乡镇代码', '5.名称', '6.代码',
                    '7.类型', '8.人口', '9.河流名称', '10.河流代码',
                    '11.名称', '12.编码', '13.名称', '14.编码',
                    '15.河流名称', '16.河流代码',
                    '17.束窄', '18.急弯', '19.低洼地', '20.临河滑坡', '21.泥石流',
                    '22.沟滩占地', '23.溃决', '24.壅水', '25.顶托', '26.改道', '27.漫流',
                    '28.备注'
                ]

                ws1.append(header_row1)
                ws1.append(header_row2)
                ws1.append(header_row3)

                for record in records:
                    row_data = []
                    for header in data_headers:
                        val = record.get(header, '')
                        if val == '☑' or val == 'þ':
                            val = 'R'
                        elif val == '☒' or val == 'ý':
                            val = 'S'
                        row_data.append(val)
                    ws1.append(row_data)

                for col in range(1, 12):
                    ws1.merge_cells(start_row=1, start_column=col, end_row=3, end_column=col)
                ws1.merge_cells(start_row=1, start_column=12, end_row=1, end_column=23)
                ws1.merge_cells(start_row=1, start_column=24, end_row=1, end_column=28)
                ws1.merge_cells(start_row=1, start_column=29, end_row=3, end_column=29)

                ws1.merge_cells(start_row=2, start_column=12, end_row=2, end_column=13)
                ws1.merge_cells(start_row=2, start_column=14, end_row=2, end_column=15)
                ws1.merge_cells(start_row=2, start_column=16, end_row=2, end_column=17)
                ws1.merge_cells(start_row=2, start_column=18, end_row=2, end_column=22)
                for col in range(23, 29):
                    ws1.merge_cells(start_row=2, start_column=col, end_row=3, end_column=col)

                for merge_info in merge_cells:
                    start_row, start_col, row_span, col_span = merge_info
                    ws1.merge_cells(
                        start_row=start_row + 4,
                        start_column=start_col + 1,
                        end_row=start_row + 4 + row_span - 1,
                        end_column=start_col + col_span
                    )

                header_font = Font(bold=True)
                header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

                for row in ws1.iter_rows(min_row=1, max_row=3, max_col=29):
                    for cell in row:
                        cell.font = header_font
                        cell.alignment = header_alignment
                        cell.border = thin_border

                wingdings2_font = Font(name='Wingdings 2')
                checkbox_cols = list(range(18, 29))

                for row in ws1.iter_rows(min_row=4, max_row=3 + len(records), max_col=29):
                    for cell in row:
                        cell.border = thin_border
                        cell.alignment = Alignment(vertical='center')
                        if cell.column in checkbox_cols:
                            cell.font = wingdings2_font
                            cell.alignment = Alignment(horizontal='center', vertical='center')

                ws2 = wb.create_sheet('附表2-跨沟道路桥涵')
                records2 = self.report_data['fubiao2']['records']
                headers2 = self.report_data['fubiao2']['headers']

                ws2.append(headers2)
                for record in records2:
                    row_data = [record.get(h, '') for h in headers2]
                    ws2.append(row_data)

                for row in ws2.iter_rows(min_row=1, max_row=1 + len(records2), max_col=len(headers2)):
                    for cell in row:
                        cell.border = thin_border
                        if cell.row == 1:
                            cell.font = header_font
                            cell.alignment = header_alignment

                ws3 = wb.create_sheet('附表3-沟滩占地')
                records3 = self.report_data['fubiao3']['records']
                headers3 = self.report_data['fubiao3']['headers']

                ws3.append(headers3)
                for record in records3:
                    row_data = [record.get(h, '') for h in headers3]
                    ws3.append(row_data)

                for row in ws3.iter_rows(min_row=1, max_row=1 + len(records3), max_col=len(headers3)):
                    for cell in row:
                        cell.border = thin_border
                        if cell.row == 1:
                            cell.font = header_font
                            cell.alignment = header_alignment

                wb.save(file)
                QMessageBox.information(self, "完成", f"已导出到:\n{file}")

            except Exception as e:
                QMessageBox.critical(self, "错误", f"导出失败:\n{str(e)}")

    def show_validation_dialog(self):
        if not self.report_data:
            QMessageBox.warning(self, "警告", "请先加载附表数据")
            return

        if self.result_dialog is not None and self.result_dialog.isVisible():
            self.result_dialog.raise_()
            self.result_dialog.activateWindow()
            return

        self.result_dialog = ValidationDialog(self)
        self.result_dialog.set_report_data(self.report_data)
        self.result_dialog.show()
