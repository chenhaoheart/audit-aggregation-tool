# -*- coding: utf-8 -*-
import os
import copy
from PySide6.QtWidgets import (
    QDialog, QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QFrame,
    QLabel, QLineEdit, QTableWidget, QTableWidgetItem,
    QHeaderView, QTabWidget, QMessageBox, QAbstractItemView,
    QTextEdit, QSplitter, QFileDialog,
    QComboBox, QCheckBox, QGraphicsOpacityEffect
)
from PySide6.QtCore import Qt, QPropertyAnimation, QEasingCurve
from PySide6.QtGui import QColor, QPalette
from core.data_validator import DataValidator
from core.config_manager import get_validation_mapping_config
from services.check_service import scan_shp_files
from core.theme_manager import get_theme_manager


class ValidationDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("空间数据校验")
        self.setMinimumSize(900, 600)
        self.setWindowFlags(Qt.Window | Qt.WindowMinMaxButtonsHint | Qt.WindowCloseButtonHint)
        self.setWindowModality(Qt.NonModal)
        self.report_data = None
        self.validator = None
        self.validation_results = None
        self.fubiao1_detail_data = []
        self.fubiao23_detail_data = []
        self.fubiao1_original_data = []
        self.fubiao23_original_data = []
        self._fangzhi_path = ''
        self._yinhuan_path = ''
        self._spatial_folder = ''
        self._validation_mapping_cfg = get_validation_mapping_config()
        self.field_mapping = copy.deepcopy(self._validation_mapping_cfg.mapping)

        self.theme_manager = get_theme_manager()

        theme = self.theme_manager.get_current_theme()
        palette = self.palette()
        palette.setColor(QPalette.Window, QColor(theme['content_bg']))
        self.setPalette(palette)
        self.setAutoFillBackground(True)
        self.setStyleSheet(self.theme_manager.get_stylesheet())

        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout()

        self.theme_manager = get_theme_manager()

        load_header = QLabel("数据加载")
        load_header.setObjectName("sectionHeaderMd")
        layout.addWidget(load_header)

        load_card = QFrame()
        load_card.setObjectName("card")
        group_layout = QVBoxLayout(load_card)
        group_layout.setSpacing(10)

        file_row = QHBoxLayout()
        shp_label = QLabel("空间数据文件夹:")
        shp_label.setObjectName("boldLabel")
        self.shp_folder_edit = QLineEdit()
        self.shp_folder_edit.setPlaceholderText("加载附表后自动搜索...")
        self.shp_folder_edit.setReadOnly(True)
        self.rescan_btn = QPushButton("重新扫描")
        self.rescan_btn.setFixedWidth(80)
        self.rescan_btn.setObjectName("clearBtn")
        self.rescan_btn.clicked.connect(self._auto_search_shp)
        self.browse_btn = QPushButton("浏览...")
        self.browse_btn.setFixedWidth(80)
        self.browse_btn.clicked.connect(self.select_shp_folder)

        file_row.addWidget(shp_label)
        file_row.addWidget(self.shp_folder_edit, 1)
        file_row.addWidget(self.rescan_btn)
        file_row.addWidget(self.browse_btn)
        group_layout.addLayout(file_row)

        self._shp_collapsed = False
        shp_collapse_header = QHBoxLayout()
        self._shp_collapse_btn = QPushButton("▼ 已识别的SHP图层")
        self._shp_collapse_btn.setObjectName("collapseSectionBtn")
        self._shp_collapse_btn.setFlat(True)
        self._shp_collapse_btn.setCursor(Qt.PointingHandCursor)
        self._shp_collapse_btn.clicked.connect(self._toggle_shp_section)
        shp_collapse_header.addWidget(self._shp_collapse_btn)
        shp_collapse_header.addStretch()
        group_layout.addLayout(shp_collapse_header)

        self.shp_table = QTableWidget()
        self.shp_table.setObjectName("layerTable")
        self.shp_table.setColumnCount(5)
        self.shp_table.setHorizontalHeaderLabels([
            "图层类型", "匹配状态", "SHP文件路径", "启用", "操作"
        ])
        self.shp_table.horizontalHeader().setStretchLastSection(False)
        self.shp_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Fixed)
        self.shp_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents)
        self.shp_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.Stretch)
        self.shp_table.horizontalHeader().setSectionResizeMode(3, QHeaderView.Fixed)
        self.shp_table.horizontalHeader().setSectionResizeMode(4, QHeaderView.Fixed)
        self.shp_table.setColumnWidth(0, 130)
        self.shp_table.setColumnWidth(3, 55)
        self.shp_table.setColumnWidth(4, 90)
        self.shp_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.shp_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.shp_table.setAlternatingRowColors(False)
        self.shp_table.verticalHeader().setVisible(False)
        row_height = 36
        self.shp_table.verticalHeader().setDefaultSectionSize(row_height)
        self.shp_table.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        group_layout.addWidget(self.shp_table)

        self._init_shp_table()

        btn_row = QHBoxLayout()
        btn_row.addStretch()

        self.validate_btn = QPushButton("开始校验")
        self.validate_btn.setFixedWidth(100)
        self.validate_btn.clicked.connect(self.start_validation)
        self.validate_btn.setEnabled(False)

        self.clear_btn = QPushButton("清空")
        self.clear_btn.setObjectName("clearBtn")
        self.clear_btn.setFixedWidth(80)
        self.clear_btn.clicked.connect(self.clear_results)

        self.export_btn = QPushButton("导出报告")
        self.export_btn.setObjectName("exportBtn")
        self.export_btn.setFixedWidth(100)
        self.export_btn.clicked.connect(self.export_report)
        self.export_btn.setEnabled(False)

        self.log_toggle_btn = QPushButton("显示日志")
        self.log_toggle_btn.setObjectName("logToggleBtn")
        self.log_toggle_btn.setFixedWidth(80)
        self.log_toggle_btn.clicked.connect(self.toggle_log)

        btn_row.addWidget(self.validate_btn)
        btn_row.addWidget(self.clear_btn)
        btn_row.addSpacing(10)

        separator1 = QFrame()
        separator1.setFrameShape(QFrame.VLine)
        separator1.setFrameShadow(QFrame.Sunken)
        btn_row.addWidget(separator1)
        btn_row.addSpacing(10)

        btn_row.addWidget(self.export_btn)
        btn_row.addSpacing(10)

        separator2 = QFrame()
        separator2.setFrameShape(QFrame.VLine)
        separator2.setFrameShadow(QFrame.Sunken)
        btn_row.addWidget(separator2)
        btn_row.addSpacing(10)

        btn_row.addWidget(self.log_toggle_btn)
        btn_row.addStretch()
        group_layout.addLayout(btn_row)
        layout.addWidget(load_card)

        main_splitter = QSplitter(Qt.Horizontal)

        self.result_tabs = QTabWidget()

        self.fubiao1_result_tab = QWidget()
        self._setup_fubiao1_result_tab()
        self.result_tabs.addTab(self.fubiao1_result_tab, "附表1 ↔ 防治对象分布P")

        self.fubiao23_result_tab = QWidget()
        self._setup_fubiao23_result_tab()
        self.result_tabs.addTab(self.fubiao23_result_tab, "附表2/3 ↔ 隐患要素分布L")

        main_splitter.addWidget(self.result_tabs)

        log_widget = QWidget()
        log_layout = QVBoxLayout()
        log_layout.setContentsMargins(5, 5, 5, 5)

        log_title = QLabel("校验日志")
        log_title.setObjectName("boldLabel")
        log_layout.addWidget(log_title)

        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)
        self.log_text.setObjectName("logText")
        log_layout.addWidget(self.log_text)
        log_widget.setLayout(log_layout)

        self.log_panel = log_widget
        main_splitter.addWidget(self.log_panel)

        self.main_splitter = main_splitter

        self.log_panel.hide()
        self.log_visible = False

        layout.addWidget(main_splitter, 1)

        self.setLayout(layout)

    def _setup_fubiao1_result_tab(self):
        layout = QVBoxLayout()
        layout.setContentsMargins(0, 4, 0, 4)

        self.fubiao1_stats_label = QLabel("等待校验...")
        self.fubiao1_stats_label.setObjectName("statsLabel")
        layout.addWidget(self.fubiao1_stats_label)

        filter_layout = QHBoxLayout()
        filter_layout.setSpacing(10)

        filter_label = QLabel("筛选:")
        filter_label.setObjectName("boldLabel")

        self.fubiao1_status_combo = QComboBox()
        self.fubiao1_status_combo.addItems(["全部", "✓ 匹配", "⚠ 仅附表", "⚠ 仅shp"])
        self.fubiao1_status_combo.setFixedWidth(100)
        self.fubiao1_status_combo.currentTextChanged.connect(self.apply_fubiao1_result_filter)

        self.fubiao1_name_edit = QLineEdit()
        self.fubiao1_name_edit.setPlaceholderText("名称搜索...")
        self.fubiao1_name_edit.setFixedWidth(150)
        self.fubiao1_name_edit.textChanged.connect(self.apply_fubiao1_result_filter)

        match_label = QLabel("是否完全匹配:")
        match_label.setObjectName("boldLabel")
        self.fubiao1_match_combo = QComboBox()
        self.fubiao1_match_combo.addItems(["全部", "是", "否"])
        self.fubiao1_match_combo.setFixedWidth(80)
        self.fubiao1_match_combo.currentTextChanged.connect(self.apply_fubiao1_result_filter)

        clear_filter_btn = QPushButton("清除筛选")
        clear_filter_btn.setFixedWidth(80)
        clear_filter_btn.setObjectName("logToggleBtn")
        clear_filter_btn.clicked.connect(self.clear_fubiao1_result_filter)

        filter_layout.addWidget(filter_label)
        filter_layout.addWidget(self.fubiao1_status_combo)
        filter_layout.addWidget(match_label)
        filter_layout.addWidget(self.fubiao1_match_combo)
        filter_layout.addWidget(QLabel("名称:"))
        filter_layout.addWidget(self.fubiao1_name_edit)
        filter_layout.addWidget(clear_filter_btn)
        filter_layout.addStretch()
        layout.addLayout(filter_layout)

        splitter = QSplitter(Qt.Vertical)

        self.fubiao1_result_table = QTableWidget()
        self.fubiao1_result_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.fubiao1_result_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.fubiao1_result_table.setSelectionMode(QAbstractItemView.SingleSelection)
        self.fubiao1_result_table.setAlternatingRowColors(False)
        self.fubiao1_result_table.horizontalHeader().setStretchLastSection(True)
        self.fubiao1_result_table.verticalHeader().setVisible(False)
        self.fubiao1_result_table.itemSelectionChanged.connect(self.on_fubiao1_selection_changed)
        splitter.addWidget(self.fubiao1_result_table)

        detail_widget = QWidget()
        detail_layout = QVBoxLayout()
        detail_layout.setContentsMargins(5, 5, 5, 5)

        detail_title = QLabel("字段详情对比")
        detail_title.setObjectName("boldLabel")
        detail_layout.addWidget(detail_title)

        self.fubiao1_detail_table = QTableWidget()
        self.fubiao1_detail_table.setRowCount(2)
        self.fubiao1_detail_table.setVerticalHeaderLabels(['附表', 'shp'])
        self.fubiao1_detail_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.fubiao1_detail_table.setAlternatingRowColors(False)
        self.fubiao1_detail_table.horizontalHeader().setStretchLastSection(True)
        self.fubiao1_detail_table.verticalHeader().setDefaultAlignment(Qt.AlignCenter)
        detail_layout.addWidget(self.fubiao1_detail_table)

        detail_widget.setLayout(detail_layout)
        splitter.addWidget(detail_widget)

        splitter.setSizes([300, 150])
        layout.addWidget(splitter)

        self.fubiao1_result_tab.setLayout(layout)

    def _setup_fubiao23_result_tab(self):
        layout = QVBoxLayout()
        layout.setContentsMargins(0, 4, 0, 4)

        self.fubiao23_stats_label = QLabel("等待校验...")
        self.fubiao23_stats_label.setObjectName("statsLabel")
        layout.addWidget(self.fubiao23_stats_label)

        filter_layout = QHBoxLayout()
        filter_layout.setSpacing(10)

        filter_label = QLabel("筛选:")
        filter_label.setObjectName("boldLabel")

        self.fubiao23_status_combo = QComboBox()
        self.fubiao23_status_combo.addItems(["全部", "✓ 匹配", "⚠ 仅附表", "⚠ 仅shp"])
        self.fubiao23_status_combo.setFixedWidth(100)
        self.fubiao23_status_combo.currentTextChanged.connect(self.apply_fubiao23_result_filter)

        self.fubiao23_name_edit = QLineEdit()
        self.fubiao23_name_edit.setPlaceholderText("名称搜索...")
        self.fubiao23_name_edit.setFixedWidth(150)
        self.fubiao23_name_edit.textChanged.connect(self.apply_fubiao23_result_filter)

        match_label = QLabel("是否完全匹配:")
        match_label.setObjectName("boldLabel")
        self.fubiao23_match_combo = QComboBox()
        self.fubiao23_match_combo.addItems(["全部", "是", "否"])
        self.fubiao23_match_combo.setFixedWidth(80)
        self.fubiao23_match_combo.currentTextChanged.connect(self.apply_fubiao23_result_filter)

        clear_filter_btn = QPushButton("清除筛选")
        clear_filter_btn.setFixedWidth(80)
        clear_filter_btn.setObjectName("logToggleBtn")
        clear_filter_btn.clicked.connect(self.clear_fubiao23_result_filter)

        filter_layout.addWidget(filter_label)
        filter_layout.addWidget(self.fubiao23_status_combo)
        filter_layout.addWidget(match_label)
        filter_layout.addWidget(self.fubiao23_match_combo)
        filter_layout.addWidget(QLabel("名称:"))
        filter_layout.addWidget(self.fubiao23_name_edit)
        filter_layout.addWidget(clear_filter_btn)
        filter_layout.addStretch()
        layout.addLayout(filter_layout)

        splitter = QSplitter(Qt.Vertical)

        self.fubiao23_result_table = QTableWidget()
        self.fubiao23_result_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.fubiao23_result_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.fubiao23_result_table.setSelectionMode(QAbstractItemView.SingleSelection)
        self.fubiao23_result_table.setAlternatingRowColors(False)
        self.fubiao23_result_table.horizontalHeader().setStretchLastSection(True)
        self.fubiao23_result_table.verticalHeader().setVisible(False)
        self.fubiao23_result_table.itemSelectionChanged.connect(self.on_fubiao23_selection_changed)
        splitter.addWidget(self.fubiao23_result_table)

        detail_widget = QWidget()
        detail_layout = QVBoxLayout()
        detail_layout.setContentsMargins(5, 5, 5, 5)

        detail_title = QLabel("字段详情对比")
        detail_title.setObjectName("boldLabel")
        detail_layout.addWidget(detail_title)

        self.fubiao23_detail_table = QTableWidget()
        self.fubiao23_detail_table.setRowCount(2)
        self.fubiao23_detail_table.setVerticalHeaderLabels(['附表', 'shp'])
        self.fubiao23_detail_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.fubiao23_detail_table.setAlternatingRowColors(False)
        self.fubiao23_detail_table.horizontalHeader().setStretchLastSection(True)
        self.fubiao23_detail_table.verticalHeader().setDefaultAlignment(Qt.AlignCenter)
        detail_layout.addWidget(self.fubiao23_detail_table)

        detail_widget.setLayout(detail_layout)
        splitter.addWidget(detail_widget)

        splitter.setSizes([300, 150])
        layout.addWidget(splitter)

        self.fubiao23_result_tab.setLayout(layout)

    def set_report_data(self, report_data):
        self.report_data = report_data
        self._auto_search_shp()

    def _auto_search_shp(self):
        if not self.report_data:
            return

        fubiao_file = None
        for key in ['fubiao1', 'fubiao2', 'fubiao3']:
            f = self.report_data.get(key, {}).get('file')
            if f:
                fubiao_file = f
                break

        if not fubiao_file:
            self.shp_folder_edit.setPlaceholderText("未找到附表文件路径，请手动选择...")
            return

        search_folder = os.path.dirname(fubiao_file)
        parent_folder = os.path.dirname(search_folder)

        scan_result = scan_shp_files(parent_folder)
        layers = scan_result.get('layers', {})
        spatial_folder = scan_result.get('spatial_folder')

        fangzhi_info = layers.get('fangzhi', {})
        yinhuan_info = layers.get('yinhuan', {})

        self._fangzhi_path = fangzhi_info.get('path', '') if fangzhi_info.get('matched') else ''
        self._yinhuan_path = yinhuan_info.get('path', '') if yinhuan_info.get('matched') else ''
        self._spatial_folder = spatial_folder or parent_folder

        self.shp_folder_edit.setText(self._spatial_folder)
        self._update_shp_table(scan_result)

        has_folder = bool(self._spatial_folder)
        self.validate_btn.setEnabled(has_folder)

    def _init_shp_table(self):
        shp_layer_types = [
            {'key': 'fangzhi', 'name': '防治对象分布'},
            {'key': 'yinhuan', 'name': '隐患要素分布'},
        ]
        self.shp_table.setRowCount(len(shp_layer_types))
        self._shp_checkboxes = {}
        self._shp_path_edits = {}
        self._shp_select_btns = {}

        for row, lt in enumerate(shp_layer_types):
            key = lt['key']
            name_item = QTableWidgetItem(lt['name'])
            name_item.setTextAlignment(Qt.AlignCenter)
            self.shp_table.setItem(row, 0, name_item)

            status_label = QLabel("未匹配")
            status_label.setObjectName("layerBadgeFail")
            status_label.setAlignment(Qt.AlignCenter)
            status_label.setFixedHeight(22)
            self.shp_table.setCellWidget(row, 1, status_label)

            path_edit = QLineEdit()
            path_edit.setReadOnly(True)
            path_edit.setPlaceholderText("未找到匹配文件...")
            path_edit.setObjectName("layerPathEdit")
            self.shp_table.setCellWidget(row, 2, path_edit)
            self._shp_path_edits[key] = path_edit

            checkbox = QCheckBox()
            checkbox.setChecked(True)
            self.shp_table.setCellWidget(row, 3, checkbox)
            self._shp_checkboxes[key] = checkbox

            select_btn = QPushButton("选择文件")
            select_btn.setFixedWidth(80)
            select_btn.clicked.connect(lambda checked, k=key: self._select_shp_file(k))
            self.shp_table.setCellWidget(row, 4, select_btn)
            self._shp_select_btns[key] = select_btn

        self.shp_table.resizeRowsToContents()
        self._adjust_shp_table_height()

    def _adjust_shp_table_height(self):
        total_height = self.shp_table.horizontalHeader().height()
        for row in range(self.shp_table.rowCount()):
            total_height += self.shp_table.rowHeight(row)
        total_height += self.shp_table.frameWidth() * 2
        if self.shp_table.rowCount() > 0:
            max_visible_rows = 10
            row_height = self.shp_table.verticalHeader().defaultSectionSize()
            max_height = self.shp_table.horizontalHeader().height() + (row_height * max_visible_rows) + self.shp_table.frameWidth() * 2
            total_height = min(total_height, max_height)
        self.shp_table.setFixedHeight(total_height)

    def _toggle_shp_section(self):
        self._shp_collapsed = not self._shp_collapsed
        if self._shp_collapsed:
            self.shp_table.setVisible(False)
            self._shp_collapse_btn.setText("▶ 已识别的SHP图层")
        else:
            self.shp_table.setVisible(True)
            self._shp_collapse_btn.setText("▼ 已识别的SHP图层")

    def _update_shp_table(self, scan_result: dict):
        layers = scan_result.get('layers', {})
        spatial_folder = scan_result.get('spatial_folder')

        fangzhi_info = layers.get('fangzhi', {})
        yinhuan_info = layers.get('yinhuan', {})

        self._fangzhi_path = fangzhi_info.get('path', '') if fangzhi_info.get('matched') else ''
        self._yinhuan_path = yinhuan_info.get('path', '') if yinhuan_info.get('matched') else ''
        self._spatial_folder = spatial_folder or self._spatial_folder

        for key, info in [('fangzhi', fangzhi_info), ('yinhuan', yinhuan_info)]:
            path = info.get('path', '')
            matched = info.get('matched', False)

            if key in self._shp_path_edits:
                self._shp_path_edits[key].setText(path if matched else '')
                self._shp_path_edits[key].setPlaceholderText(
                    path if matched else '未找到匹配文件...'
                )

            if key in self._shp_checkboxes:
                self._shp_checkboxes[key].setChecked(matched)

            row = 0 if key == 'fangzhi' else 1
            status_label = self.shp_table.cellWidget(row, 1)
            if status_label:
                if matched:
                    status_label.setText("已匹配")
                    status_label.setObjectName("layerBadgePass")
                else:
                    status_label.setText("未匹配")
                    status_label.setObjectName("layerBadgeFail")
                status_label.style().unpolish(status_label)
                status_label.style().polish(status_label)

        self.shp_table.resizeRowsToContents()
        self._adjust_shp_table_height()

    def _select_shp_file(self, layer_key: str):
        file_path, _ = QFileDialog.getOpenFileName(
            self, f"选择{'防治对象' if layer_key == 'fangzhi' else '隐患要素'}SHP文件",
            "", "Shapefiles (*.shp)"
        )
        if file_path:
            if layer_key == 'fangzhi':
                self._fangzhi_path = file_path
            else:
                self._yinhuan_path = file_path

            if layer_key in self._shp_path_edits:
                self._shp_path_edits[layer_key].setText(file_path)

            if layer_key in self._shp_checkboxes:
                self._shp_checkboxes[layer_key].setChecked(True)

            row = 0 if layer_key == 'fangzhi' else 1
            status_label = self.shp_table.cellWidget(row, 1)
            if status_label:
                status_label.setText("已选择")
                status_label.setObjectName("layerBadgePass")
                status_label.style().unpolish(status_label)
                status_label.style().polish(status_label)

            if not self.validate_btn.isEnabled():
                self.validate_btn.setEnabled(True)

    def select_shp_folder(self):
        folder = QFileDialog.getExistingDirectory(self, "选择空间数据文件夹")
        if folder:
            self.shp_folder_edit.setText(folder)
            self._spatial_folder = folder
            scan_result = scan_shp_files(folder)
            self._update_shp_table(scan_result)
            self.validate_btn.setEnabled(True)

    def toggle_log(self):
        if self.log_visible:
            self.main_splitter.setSizes([1, 0])
            self.log_toggle_btn.setText("显示日志")
            self.log_visible = False
        else:
            self.log_panel.show()
            total_width = self.main_splitter.width()
            self.main_splitter.setSizes([int(total_width * 0.7), int(total_width * 0.3)])
            self.log_toggle_btn.setText("隐藏日志")
            self.log_visible = True

    def clear_results(self):
        self.shp_folder_edit.clear()
        self.shp_folder_edit.setPlaceholderText("加载附表后自动搜索...")
        self._fangzhi_path = ''
        self._yinhuan_path = ''
        self._spatial_folder = ''

        empty_scan = {'layers': {'fangzhi': {'path': '', 'matched': False}, 'yinhuan': {'path': '', 'matched': False}}, 'spatial_folder': ''}
        self._update_shp_table(empty_scan)

        self.log_text.clear()
        self.validation_results = None
        self.fubiao1_detail_data = []
        self.fubiao23_detail_data = []
        self.fubiao1_original_data = []
        self.fubiao23_original_data = []

        self.fubiao1_result_table.setRowCount(0)
        self.fubiao1_detail_table.setColumnCount(0)
        self.fubiao23_result_table.setRowCount(0)
        self.fubiao23_detail_table.setColumnCount(0)

        self.fubiao1_status_combo.setCurrentText("全部")
        self.fubiao1_name_edit.clear()
        self.fubiao1_match_combo.setCurrentText("全部")
        self.fubiao23_status_combo.setCurrentText("全部")
        self.fubiao23_name_edit.clear()
        self.fubiao23_match_combo.setCurrentText("全部")

        self.validate_btn.setEnabled(False)
        self.export_btn.setEnabled(False)
        self.fubiao1_stats_label.setText("等待校验...")
        self.fubiao23_stats_label.setText("等待校验...")

    def append_log(self, msg):
        self.log_text.append(msg)
        scrollbar = self.log_text.verticalScrollBar()
        scrollbar.setValue(scrollbar.maximum())

    def start_validation(self):
        fangzhi_enabled = self._shp_checkboxes.get('fangzhi', QCheckBox()).isChecked()
        yinhuan_enabled = self._shp_checkboxes.get('yinhuan', QCheckBox()).isChecked()

        if not fangzhi_enabled and not yinhuan_enabled:
            QMessageBox.warning(self, "警告", "请至少启用一个SHP图层")
            return

        has_manual_paths = False
        load_fangzhi_path = None
        load_yinhuan_path = None

        if fangzhi_enabled and self._fangzhi_path:
            load_fangzhi_path = self._fangzhi_path
            has_manual_paths = True
        if yinhuan_enabled and self._yinhuan_path:
            load_yinhuan_path = self._yinhuan_path
            has_manual_paths = True

        shp_folder = self._spatial_folder or self.shp_folder_edit.text()
        if not has_manual_paths and not shp_folder:
            QMessageBox.warning(self, "警告", "请先选择空间数据文件夹或手动指定SHP文件")
            return

        if shp_folder and not os.path.exists(shp_folder):
            QMessageBox.warning(self, "警告", "空间数据文件夹不存在")
            return

        self.log_text.clear()
        self.validation_results = None

        try:
            self.validator = DataValidator()
            self.validator.progress_callback = self.append_log
            self.validator.load_fubiao(self.report_data)

            self.append_log("正在加载空间数据...")
            if has_manual_paths:
                if not self.validator.load_shp_files(
                    fangzhi_path=load_fangzhi_path if fangzhi_enabled else None,
                    yinhuan_path=load_yinhuan_path if yinhuan_enabled else None
                ):
                    QMessageBox.warning(self, "警告", "空间数据加载失败")
                    return
            else:
                if not self.validator.load_shp_data(shp_folder):
                    QMessageBox.warning(self, "警告", "空间数据加载失败")
                    return
        except Exception as e:
            QMessageBox.critical(self, "错误", f"数据加载失败:\n{e}")
            return

        self.field_mapping = copy.deepcopy(self._validation_mapping_cfg.mapping)
        self.validator.set_field_mapping(self.field_mapping)
        self.append_log("已加载系统设置中的字段映射配置，开始校验...")

        try:
            self.validation_results = self.validator.validate_all()
        except Exception as e:
            QMessageBox.critical(self, "错误", f"校验执行失败:\n{e}")
            return

        self.update_fubiao1_result()
        self.update_fubiao23_result()

        self.export_btn.setEnabled(True)

    def update_fubiao1_result(self):
        r = self.validation_results.get('fubiao1_vs_fangzhi', {})

        stats_text = f"✓ 匹配成功: {r.get('match_count', 0)} 条   |   " \
                     f"⚠ 仅附表有: {r.get('fubiao_only_count', 0)} 条   |   " \
                     f"⚠ 仅shp有: {r.get('shp_only_count', 0)} 条"
        self.fubiao1_stats_label.setText(stats_text)

        mapping = self.field_mapping.get('fubiao1_vs_fangzhi', {})
        match_fields = mapping.get('match_fields', {})
        detail_fields = mapping.get('detail_fields', {})

        headers = ['状态', '是否完全匹配', '序号', '名称', '代码', '来源', '错误描述']
        self.fubiao1_detail_data = []

        for item in r.get('matched', []):
            fubiao_rec = item.get('fubiao_record', {})
            shp_recs = item.get('shp_records', [])
            shp_rec = shp_recs[0] if shp_recs else None

            orig_seq = fubiao_rec.get('序号', '')

            diff_fields = []
            for shp_field, fb_field in detail_fields.items():
                if not fb_field or not shp_field:
                    continue
                fb_val = str(fubiao_rec.get(fb_field, '')).strip()
                shp_val = str(shp_rec.get(shp_field, '')).strip() if shp_rec else ''
                if fb_val and shp_val and fb_val != shp_val:
                    diff_fields.append(f"{shp_field}不一致")

            is_match = '否' if diff_fields else '是'
            error_desc = '，'.join(diff_fields) if diff_fields else ''

            self.fubiao1_detail_data.append({
                '状态': '✓ 匹配',
                '是否完全匹配': is_match,
                '序号': orig_seq,
                '名称': fubiao_rec.get(match_fields.get('名称', '5.名称'), ''),
                '代码': fubiao_rec.get(match_fields.get('代码', '6.代码'), ''),
                '来源': '附表+shp',
                '错误描述': error_desc,
                '_fubiao_rec': fubiao_rec,
                '_shp_rec': shp_rec
            })

        for rec in r.get('fubiao_only', []):
            orig_seq = rec.get('序号', '')
            self.fubiao1_detail_data.append({
                '状态': '⚠ 仅附表',
                '是否完全匹配': '否',
                '序号': orig_seq,
                '名称': rec.get(match_fields.get('名称', '5.名称'), ''),
                '代码': rec.get(match_fields.get('代码', '6.代码'), ''),
                '来源': '附表',
                '错误描述': 'shp中未找到',
                '_fubiao_rec': rec,
                '_shp_rec': None
            })

        for rec in r.get('shp_only', []):
            self.fubiao1_detail_data.append({
                '状态': '⚠ 仅shp',
                '是否完全匹配': '否',
                '序号': '',
                '名称': rec.get('名称', ''),
                '代码': rec.get('代码', ''),
                '来源': f"shp ({rec.get('_source_folder', '')})",
                '错误描述': '附表中未找到',
                '_fubiao_rec': None,
                '_shp_rec': rec
            })

        self.fubiao1_result_table.clear()
        self.fubiao1_result_table.setColumnCount(len(headers))
        self.fubiao1_result_table.setHorizontalHeaderLabels(headers)
        self.fubiao1_result_table.setRowCount(len(self.fubiao1_detail_data))

        for row, rec in enumerate(self.fubiao1_detail_data):
            for col, header in enumerate(headers):
                value = rec.get(header, '')
                if header == '状态':
                    badge = QLabel(str(value))
                    if '✓' in value:
                        badge.setObjectName("badgePass")
                    elif '⚠' in value:
                        badge.setObjectName("badgeFail")
                    else:
                        badge.setObjectName("badgePass")
                    badge.setAlignment(Qt.AlignCenter)
                    badge.setFixedHeight(22)
                    self.fubiao1_result_table.setCellWidget(row, col, badge)
                else:
                    item = QTableWidgetItem(str(value))
                    self.fubiao1_result_table.setItem(row, col, item)

        self.fubiao1_result_table.resizeColumnsToContents()

        self.fubiao1_original_data = self.fubiao1_detail_data.copy()
        self.fubiao1_status_combo.setCurrentText("全部")
        self.fubiao1_name_edit.clear()
        self.fubiao1_match_combo.setCurrentText("全部")

        if len(self.fubiao1_detail_data) > 0:
            self.fubiao1_result_table.selectRow(0)

    def apply_fubiao1_result_filter(self):
        status_filter = self.fubiao1_status_combo.currentText()
        name_filter = self.fubiao1_name_edit.text().strip().lower()
        match_filter = self.fubiao1_match_combo.currentText()

        filtered_data = []
        for rec in self.fubiao1_original_data:
            status = rec.get('状态', '')
            name = str(rec.get('名称', '')).lower()
            is_match = rec.get('是否完全匹配', '')

            if status_filter != "全部" and status != status_filter:
                continue
            if name_filter and name_filter not in name:
                continue
            if match_filter != "全部" and is_match != match_filter:
                continue

            filtered_data.append(rec)

        self.fubiao1_detail_data = filtered_data
        headers = ['状态', '是否完全匹配', '序号', '名称', '代码', '来源', '错误描述']
        self.fubiao1_result_table.clear()
        self.fubiao1_result_table.setColumnCount(len(headers))
        self.fubiao1_result_table.setHorizontalHeaderLabels(headers)
        self.fubiao1_result_table.setRowCount(len(filtered_data))

        for row, rec in enumerate(filtered_data):
            for col, header in enumerate(headers):
                value = rec.get(header, '')
                if header == '状态':
                    badge = QLabel(str(value))
                    if '✓' in value:
                        badge.setObjectName("badgePass")
                    elif '⚠' in value:
                        badge.setObjectName("badgeFail")
                    else:
                        badge.setObjectName("badgePass")
                    badge.setAlignment(Qt.AlignCenter)
                    badge.setFixedHeight(22)
                    self.fubiao1_result_table.setCellWidget(row, col, badge)
                else:
                    item = QTableWidgetItem(str(value))
                    self.fubiao1_result_table.setItem(row, col, item)

        self.fubiao1_result_table.resizeColumnsToContents()

        if len(filtered_data) > 0:
            self.fubiao1_result_table.selectRow(0)

    def clear_fubiao1_result_filter(self):
        self.fubiao1_status_combo.setCurrentText("全部")
        self.fubiao1_name_edit.clear()
        self.fubiao1_match_combo.setCurrentText("全部")
        self.apply_fubiao1_result_filter()

    def on_fubiao1_selection_changed(self):
        selected_rows = self.fubiao1_result_table.selectedItems()
        if not selected_rows:
            return

        row = selected_rows[0].row()
        if row >= len(self.fubiao1_detail_data):
            return

        data = self.fubiao1_detail_data[row]
        fubiao_rec = data.get('_fubiao_rec', {})
        shp_rec = data.get('_shp_rec')

        mapping = self.field_mapping.get('fubiao1_vs_fangzhi', {})
        detail_fields = mapping.get('detail_fields', {})
        match_fields = mapping.get('match_fields', {})
        all_fields = {**match_fields, **detail_fields}

        field_mapping = []
        for shp_field, fb_field in all_fields.items():
            field_mapping.append((shp_field, fb_field, shp_field))

        self.fubiao1_detail_table.setColumnCount(len(field_mapping))
        self.fubiao1_detail_table.setHorizontalHeaderLabels([f[0] for f in field_mapping])
        self.fubiao1_detail_table.setRowCount(2)
        self.fubiao1_detail_table.setVerticalHeaderLabels(['附表', 'shp'])

        for col, (label, fubiao_field, shp_field) in enumerate(field_mapping):
            fubiao_val = str(fubiao_rec.get(fubiao_field, '')) if fubiao_rec else '-'
            shp_val = str(shp_rec.get(shp_field, '')) if shp_rec and shp_field else '-'

            fubiao_item = QTableWidgetItem(fubiao_val if fubiao_val else '-')
            shp_item = QTableWidgetItem(shp_val if shp_val else '-')

            if fubiao_rec and shp_rec and shp_field and fubiao_val != shp_val and fubiao_val and shp_val:
                fubiao_item.setBackground(QColor('#fff3cd'))
                shp_item.setBackground(QColor('#fff3cd'))

            self.fubiao1_detail_table.setItem(0, col, fubiao_item)
            self.fubiao1_detail_table.setItem(1, col, shp_item)

        self.fubiao1_detail_table.resizeColumnsToContents()

    def update_fubiao23_result(self):
        r = self.validation_results.get('fubiao23_vs_yinhuan', {})

        stats_text = f"附表2: ✓ {r.get('fubiao2_match_count', 0)} 条, ⚠ {r.get('fubiao2_only_count', 0)} 条   |   " \
                     f"附表3: ✓ {r.get('fubiao3_match_count', 0)} 条, ⚠ {r.get('fubiao3_only_count', 0)} 条   |   " \
                     f"仅shp: {r.get('shp_only_count', 0)} 条"
        self.fubiao23_stats_label.setText(stats_text)

        headers = ['状态', '是否完全匹配', '序号', '名称', '编码', '来源表', '来源', '错误描述']
        self.fubiao23_detail_data = []

        mapping2 = self.field_mapping.get('fubiao2_vs_yinhuan', {})
        mapping3 = self.field_mapping.get('fubiao3_vs_yinhuan', {})
        match_fields2 = mapping2.get('match_fields', {})
        match_fields3 = mapping3.get('match_fields', {})
        detail_fields2 = mapping2.get('detail_fields', {})
        detail_fields3 = mapping3.get('detail_fields', {})

        name_field2 = match_fields2.get('名称', '名称')
        code_field2 = match_fields2.get('编号', '编码')
        name_field3 = match_fields3.get('名称', '名称')
        code_field3 = match_fields3.get('编号', '编码')

        for item in r.get('fubiao2_matched', []):
            fubiao_rec = item.get('fubiao_record', {})
            shp_recs = item.get('shp_records', [])
            shp_rec = shp_recs[0] if shp_recs else None

            orig_seq = fubiao_rec.get('序号', '')

            diff_fields = []
            for shp_field, fb_field in detail_fields2.items():
                if not fb_field or not shp_field:
                    continue
                fb_val = str(fubiao_rec.get(fb_field, '')).strip()
                shp_val = str(shp_rec.get(shp_field, '')).strip() if shp_rec else ''
                if fb_val and shp_val and fb_val != shp_val:
                    diff_fields.append(f"{shp_field}不一致")

            is_match = '否' if diff_fields else '是'
            error_desc = '，'.join(diff_fields) if diff_fields else ''

            self.fubiao23_detail_data.append({
                '状态': '✓ 匹配',
                '是否完全匹配': is_match,
                '序号': orig_seq,
                '名称': fubiao_rec.get(name_field2, ''),
                '编码': fubiao_rec.get(code_field2, ''),
                '来源表': '附表2',
                '来源': '附表2+shp',
                '错误描述': error_desc,
                '_fubiao_rec': fubiao_rec,
                '_shp_rec': shp_rec
            })

        for rec in r.get('fubiao2_only', []):
            orig_seq = rec.get('序号', '')
            self.fubiao23_detail_data.append({
                '状态': '⚠ 仅附表2',
                '是否完全匹配': '否',
                '序号': orig_seq,
                '名称': rec.get(name_field2, ''),
                '编码': rec.get(code_field2, ''),
                '来源表': '附表2',
                '来源': '附表2',
                '错误描述': 'shp中未找到',
                '_fubiao_rec': rec,
                '_shp_rec': None
            })

        for item in r.get('fubiao3_matched', []):
            fubiao_rec = item.get('fubiao_record', {})
            shp_recs = item.get('shp_records', [])
            shp_rec = shp_recs[0] if shp_recs else None

            orig_seq = fubiao_rec.get('序号', '')

            diff_fields = []
            for shp_field, fb_field in detail_fields3.items():
                if not fb_field or not shp_field:
                    continue
                fb_val = str(fubiao_rec.get(fb_field, '')).strip()
                shp_val = str(shp_rec.get(shp_field, '')).strip() if shp_rec else ''
                if fb_val and shp_val and fb_val != shp_val:
                    diff_fields.append(f"{shp_field}不一致")

            is_match = '否' if diff_fields else '是'
            error_desc = '，'.join(diff_fields) if diff_fields else ''

            self.fubiao23_detail_data.append({
                '状态': '✓ 匹配',
                '是否完全匹配': is_match,
                '序号': orig_seq,
                '名称': fubiao_rec.get(name_field3, ''),
                '编码': fubiao_rec.get(code_field3, ''),
                '来源表': '附表3',
                '来源': '附表3+shp',
                '错误描述': error_desc,
                '_fubiao_rec': fubiao_rec,
                '_shp_rec': shp_rec
            })

        for rec in r.get('fubiao3_only', []):
            orig_seq = rec.get('序号', '')
            self.fubiao23_detail_data.append({
                '状态': '⚠ 仅附表3',
                '是否完全匹配': '否',
                '序号': orig_seq,
                '名称': rec.get(name_field3, ''),
                '编码': rec.get(code_field3, ''),
                '来源表': '附表3',
                '来源': '附表3',
                '错误描述': 'shp中未找到',
                '_fubiao_rec': rec,
                '_shp_rec': None
            })

        for rec in r.get('shp_only', []):
            self.fubiao23_detail_data.append({
                '状态': '⚠ 仅shp',
                '是否完全匹配': '否',
                '序号': '',
                '名称': rec.get('名称', ''),
                '编码': rec.get('编号', ''),
                '来源表': '',
                '来源': f"shp ({rec.get('_source_folder', '')})",
                '错误描述': '附表中未找到',
                '_fubiao_rec': None,
                '_shp_rec': rec
            })

        self.fubiao23_result_table.clear()
        self.fubiao23_result_table.setColumnCount(len(headers))
        self.fubiao23_result_table.setHorizontalHeaderLabels(headers)
        self.fubiao23_result_table.setRowCount(len(self.fubiao23_detail_data))

        for row, rec in enumerate(self.fubiao23_detail_data):
            for col, header in enumerate(headers):
                value = rec.get(header, '')
                if header == '状态':
                    badge = QLabel(str(value))
                    if '✓' in value:
                        badge.setObjectName("badgePass")
                    elif '⚠' in value:
                        badge.setObjectName("badgeFail")
                    else:
                        badge.setObjectName("badgePass")
                    badge.setAlignment(Qt.AlignCenter)
                    badge.setFixedHeight(22)
                    self.fubiao23_result_table.setCellWidget(row, col, badge)
                else:
                    item = QTableWidgetItem(str(value))
                    self.fubiao23_result_table.setItem(row, col, item)

        self.fubiao23_result_table.resizeColumnsToContents()

        self.fubiao23_original_data = self.fubiao23_detail_data.copy()

        self.fubiao23_status_combo.setCurrentText("全部")
        self.fubiao23_name_edit.clear()
        self.fubiao23_match_combo.setCurrentText("全部")

        if len(self.fubiao23_detail_data) > 0:
            self.fubiao23_result_table.selectRow(0)

    def apply_fubiao23_result_filter(self):
        status_filter = self.fubiao23_status_combo.currentText()
        name_filter = self.fubiao23_name_edit.text().strip().lower()
        match_filter = self.fubiao23_match_combo.currentText()

        filtered_data = []
        for rec in self.fubiao23_original_data:
            status = rec.get('状态', '')
            name = str(rec.get('名称', '')).lower()
            is_match = rec.get('是否完全匹配', '')

            if status_filter != "全部" and status != status_filter:
                continue
            if name_filter and name_filter not in name:
                continue
            if match_filter != "全部" and is_match != match_filter:
                continue

            filtered_data.append(rec)

        self.fubiao23_detail_data = filtered_data
        headers = ['状态', '是否完全匹配', '序号', '名称', '编码', '来源表', '来源', '错误描述']
        self.fubiao23_result_table.clear()
        self.fubiao23_result_table.setColumnCount(len(headers))
        self.fubiao23_result_table.setHorizontalHeaderLabels(headers)
        self.fubiao23_result_table.setRowCount(len(filtered_data))

        for row, rec in enumerate(filtered_data):
            for col, header in enumerate(headers):
                value = rec.get(header, '')
                if header == '状态':
                    badge = QLabel(str(value))
                    if '✓' in value:
                        badge.setObjectName("badgePass")
                    elif '⚠' in value:
                        badge.setObjectName("badgeFail")
                    else:
                        badge.setObjectName("badgePass")
                    badge.setAlignment(Qt.AlignCenter)
                    badge.setFixedHeight(22)
                    self.fubiao23_result_table.setCellWidget(row, col, badge)
                else:
                    item = QTableWidgetItem(str(value))
                    self.fubiao23_result_table.setItem(row, col, item)

        self.fubiao23_result_table.resizeColumnsToContents()

        if len(filtered_data) > 0:
            self.fubiao23_result_table.selectRow(0)

    def clear_fubiao23_result_filter(self):
        self.fubiao23_status_combo.setCurrentText("全部")
        self.fubiao23_name_edit.clear()
        self.fubiao23_match_combo.setCurrentText("全部")
        self.apply_fubiao23_result_filter()

    def on_fubiao23_selection_changed(self):
        selected_rows = self.fubiao23_result_table.selectedItems()
        if not selected_rows:
            return

        row = selected_rows[0].row()
        if row >= len(self.fubiao23_detail_data):
            return

        data = self.fubiao23_detail_data[row]
        fubiao_rec = data.get('_fubiao_rec', {})
        shp_rec = data.get('_shp_rec')
        source_table = data.get('来源表', '')

        if source_table == '附表2':
            mapping = self.field_mapping.get('fubiao2_vs_yinhuan', {})
        elif source_table == '附表3':
            mapping = self.field_mapping.get('fubiao3_vs_yinhuan', {})
        else:
            mapping = {}

        detail_fields = mapping.get('detail_fields', {})
        match_fields = mapping.get('match_fields', {})
        all_fields = {**match_fields, **detail_fields}

        field_mapping = []
        for shp_field, fb_field in all_fields.items():
            field_mapping.append((shp_field, fb_field, shp_field))

        self.fubiao23_detail_table.setColumnCount(len(field_mapping))
        self.fubiao23_detail_table.setHorizontalHeaderLabels([f[0] for f in field_mapping])
        self.fubiao23_detail_table.setRowCount(2)
        self.fubiao23_detail_table.setVerticalHeaderLabels(['附表', 'shp'])

        for col, (label, fubiao_field, shp_field) in enumerate(field_mapping):
            fubiao_val = ''
            if fubiao_rec:
                if fubiao_field in fubiao_rec:
                    fubiao_val = str(fubiao_rec.get(fubiao_field, ''))
                else:
                    for k in fubiao_rec.keys():
                        if fubiao_field in k or k in fubiao_field:
                            fubiao_val = str(fubiao_rec.get(k, ''))
                            break

            shp_val = str(shp_rec.get(shp_field, '')) if shp_rec and shp_field else '-'

            fubiao_item = QTableWidgetItem(fubiao_val if fubiao_val else '-')
            shp_item = QTableWidgetItem(shp_val if shp_val else '-')

            if fubiao_rec and shp_rec and shp_field and fubiao_val != shp_val and fubiao_val and shp_val:
                fubiao_item.setBackground(QColor('#fff3cd'))
                shp_item.setBackground(QColor('#fff3cd'))

            self.fubiao23_detail_table.setItem(0, col, fubiao_item)
            self.fubiao23_detail_table.setItem(1, col, shp_item)

        self.fubiao23_detail_table.resizeColumnsToContents()

    def export_report(self):
        if not self.validation_results:
            QMessageBox.warning(self, "警告", "没有校验结果可导出")
            return

        file, _ = QFileDialog.getSaveFileName(
            self, "导出校验报告", "附表校验报告.xlsx", "Excel Files (*.xlsx)"
        )

        if file:
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

                wb = Workbook()
                header_font = Font(bold=True)
                header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
                header_alignment = Alignment(horizontal='center', vertical='center')
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

                ws1 = wb.active
                ws1.title = "附表1对比结果"
                r1 = self.validation_results.get('fubiao1_vs_fangzhi', {})

                mapping1 = self.field_mapping.get('fubiao1_vs_fangzhi', {})
                match_fields1 = mapping1.get('match_fields', {})
                detail_fields1 = mapping1.get('detail_fields', {})

                headers1 = ['状态', '是否完全匹配', '序号', '名称', '代码', '来源', '错误描述']
                ws1.append(headers1)
                for cell in ws1[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = thin_border
                    cell.alignment = header_alignment

                for item in r1.get('matched', []):
                    fubiao_rec = item.get('fubiao_record', {})
                    shp_recs = item.get('shp_records', [])
                    shp_rec = shp_recs[0] if shp_recs else None
                    orig_seq = fubiao_rec.get('序号', '')

                    diff_fields = []
                    for shp_field, fb_field in detail_fields1.items():
                        if not fb_field or not shp_field:
                            continue
                        fb_val = str(fubiao_rec.get(fb_field, '')).strip()
                        shp_val = str(shp_rec.get(shp_field, '')).strip() if shp_rec else ''
                        if fb_val and shp_val and fb_val != shp_val:
                            diff_fields.append(f"{shp_field}不一致")

                    is_match = '否' if diff_fields else '是'
                    error_desc = '，'.join(diff_fields) if diff_fields else ''

                    ws1.append([
                        '✓ 匹配',
                        is_match,
                        orig_seq,
                        fubiao_rec.get(match_fields1.get('名称', '5.名称'), ''),
                        fubiao_rec.get(match_fields1.get('代码', '6.代码'), ''),
                        '附表+shp',
                        error_desc
                    ])

                for rec in r1.get('fubiao_only', []):
                    ws1.append([
                        '⚠ 仅附表',
                        '否',
                        rec.get('序号', ''),
                        rec.get(match_fields1.get('名称', '5.名称'), ''),
                        rec.get(match_fields1.get('代码', '6.代码'), ''),
                        '附表',
                        'shp中未找到'
                    ])

                for rec in r1.get('shp_only', []):
                    ws1.append([
                        '⚠ 仅shp',
                        '否',
                        '',
                        rec.get('名称', ''),
                        rec.get('代码', ''),
                        f"shp ({rec.get('_source_folder', '')})",
                        '附表中未找到'
                    ])

                for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row, max_col=len(headers1)):
                    for cell in row:
                        cell.border = thin_border

                ws2 = wb.create_sheet("附表2-3对比结果")
                r2 = self.validation_results.get('fubiao23_vs_yinhuan', {})

                mapping2 = self.field_mapping.get('fubiao2_vs_yinhuan', {})
                mapping3 = self.field_mapping.get('fubiao3_vs_yinhuan', {})
                match_fields2 = mapping2.get('match_fields', {})
                match_fields3 = mapping3.get('match_fields', {})
                detail_fields2 = mapping2.get('detail_fields', {})
                detail_fields3 = mapping3.get('detail_fields', {})

                headers2 = ['状态', '是否完全匹配', '序号', '名称', '编码', '来源表', '来源', '错误描述']
                ws2.append(headers2)
                for cell in ws2[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = thin_border
                    cell.alignment = header_alignment

                for item in r2.get('fubiao2_matched', []):
                    fubiao_rec = item.get('fubiao_record', {})
                    shp_recs = item.get('shp_records', [])
                    shp_rec = shp_recs[0] if shp_recs else None
                    orig_seq = fubiao_rec.get('序号', '')

                    diff_fields = []
                    for shp_field, fb_field in detail_fields2.items():
                        if not fb_field or not shp_field:
                            continue
                        fb_val = str(fubiao_rec.get(fb_field, '')).strip()
                        shp_val = str(shp_rec.get(shp_field, '')).strip() if shp_rec else ''
                        if fb_val and shp_val and fb_val != shp_val:
                            diff_fields.append(f"{shp_field}不一致")

                    is_match = '否' if diff_fields else '是'
                    error_desc = '，'.join(diff_fields) if diff_fields else ''

                    ws2.append([
                        '✓ 匹配',
                        is_match,
                        orig_seq,
                        fubiao_rec.get(match_fields2.get('名称', '名称'), ''),
                        fubiao_rec.get(match_fields2.get('编号', '编码'), ''),
                        '附表2',
                        '附表2+shp',
                        error_desc
                    ])

                for rec in r2.get('fubiao2_only', []):
                    ws2.append([
                        '⚠ 仅附表2',
                        '否',
                        rec.get('序号', ''),
                        rec.get(match_fields2.get('名称', '名称'), ''),
                        rec.get(match_fields2.get('编号', '编码'), ''),
                        '附表2',
                        '附表2',
                        'shp中未找到'
                    ])

                for item in r2.get('fubiao3_matched', []):
                    fubiao_rec = item.get('fubiao_record', {})
                    shp_recs = item.get('shp_records', [])
                    shp_rec = shp_recs[0] if shp_recs else None
                    orig_seq = fubiao_rec.get('序号', '')

                    diff_fields = []
                    for shp_field, fb_field in detail_fields3.items():
                        if not fb_field or not shp_field:
                            continue
                        fb_val = str(fubiao_rec.get(fb_field, '')).strip()
                        shp_val = str(shp_rec.get(shp_field, '')).strip() if shp_rec else ''
                        if fb_val and shp_val and fb_val != shp_val:
                            diff_fields.append(f"{shp_field}不一致")

                    is_match = '否' if diff_fields else '是'
                    error_desc = '，'.join(diff_fields) if diff_fields else ''

                    ws2.append([
                        '✓ 匹配',
                        is_match,
                        orig_seq,
                        fubiao_rec.get(match_fields3.get('名称', '名称'), ''),
                        fubiao_rec.get(match_fields3.get('编号', '编码'), ''),
                        '附表3',
                        '附表3+shp',
                        error_desc
                    ])

                for rec in r2.get('fubiao3_only', []):
                    ws2.append([
                        '⚠ 仅附表3',
                        '否',
                        rec.get('序号', ''),
                        rec.get(match_fields3.get('名称', '名称'), ''),
                        rec.get(match_fields3.get('编号', '编码'), ''),
                        '附表3',
                        '附表3',
                        'shp中未找到'
                    ])

                for rec in r2.get('shp_only', []):
                    ws2.append([
                        '⚠ 仅shp',
                        '否',
                        '',
                        rec.get('名称', ''),
                        rec.get('编号', ''),
                        '',
                        f"shp ({rec.get('_source_folder', '')})",
                        '附表中未找到'
                    ])

                for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row, max_col=len(headers2)):
                    for cell in row:
                        cell.border = thin_border

                wb.save(file)
                QMessageBox.information(self, "完成", f"已导出到:\n{file}")

            except Exception as e:
                QMessageBox.critical(self, "错误", f"导出失败:\n{str(e)}")

    def showEvent(self, event):
        super().showEvent(event)
        if not hasattr(self, '_animated'):
            self._animated = True
            self._animate_entrance()

    def _animate_entrance(self):
        effect = QGraphicsOpacityEffect(self)
        self.setGraphicsEffect(effect)
        effect.setOpacity(0.0)

        anim = QPropertyAnimation(effect, b"opacity")
        anim.setDuration(200)
        anim.setStartValue(0.0)
        anim.setEndValue(1.0)
        anim.setEasingCurve(QEasingCurve.OutCubic)
        anim.finished.connect(lambda: self.setGraphicsEffect(None))
        anim.start()
        self._entrance_anim = anim
