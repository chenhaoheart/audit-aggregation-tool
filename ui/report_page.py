# -*- coding: utf-8 -*-
"""
成果报表展示页面
"""

import os
from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QFrame,
    QLabel, QLineEdit, QFileDialog, QTableWidget, QTableWidgetItem,
    QHeaderView, QTabWidget, QMessageBox, QAbstractItemView,
    QGroupBox, QTextEdit, QScrollArea, QFrame, QSplitter, QDialog,
    QComboBox, QFormLayout
)
from PySide6.QtCore import Qt
from PySide6.QtGui import QColor
from core.report_reader import load_all_reports
from core.data_validator import DataValidator


class ReportPage(QWidget):
    """成果报表展示页面"""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.report_data = None
        self.validator = None
        self.validation_results = None
        self.result_dialog = None  # 校验结果弹窗
        self.check_errors = []  # 检查结果数据
        self.check_original_errors = []  # 原始数据用于筛选
        self.original_data = {
            'fubiao1': [],
            'fubiao2': [],
            'fubiao3': []
        }
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout()
        layout.setSpacing(15)
        layout.setContentsMargins(0, 0, 0, 0)

        # 文件选择区域
        file_group = QGroupBox("数据加载")
        group_layout = QVBoxLayout()
        group_layout.setSpacing(10)

        # 第一行：文件夹选择
        file_row = QHBoxLayout()
        file_row.setSpacing(10)

        folder_label = QLabel("成果报表文件夹:")
        folder_label.setStyleSheet("font-weight: bold; color: #2c3e50;")
        self.folder_edit = QLineEdit()
        self.folder_edit.setPlaceholderText("选择包含附表Excel文件的文件夹...")
        self.folder_edit.setReadOnly(True)

        self.browse_btn = QPushButton("浏览...")
        self.browse_btn.setFixedWidth(80)
        self.browse_btn.clicked.connect(self.select_folder)

        file_row.addWidget(folder_label)
        file_row.addWidget(self.folder_edit, 1)
        file_row.addWidget(self.browse_btn)
        group_layout.addLayout(file_row)

        # 第二行：按钮
        btn_row = QHBoxLayout()
        btn_row.setSpacing(15)
        btn_row.addStretch()

        self.load_btn = QPushButton("加载数据")
        self.load_btn.setFixedWidth(100)
        self.load_btn.clicked.connect(self.load_data)
        self.load_btn.setEnabled(False)

        self.check_btn = QPushButton("报表成果检查")
        self.check_btn.setObjectName("checkBtn")
        self.check_btn.setFixedWidth(120)
        self.check_btn.clicked.connect(self.run_data_check)
        self.check_btn.setEnabled(False)

        self.export_btn = QPushButton("导出Excel")
        self.export_btn.setObjectName("exportBtn")
        self.export_btn.setFixedWidth(100)
        self.export_btn.clicked.connect(self.export_excel)
        self.export_btn.setEnabled(False)

        self.validate_btn = QPushButton("空间数据校验")
        self.validate_btn.setObjectName("validateBtn")
        self.validate_btn.setFixedWidth(120)
        self.validate_btn.clicked.connect(self.show_validation_dialog)
        self.validate_btn.setEnabled(False)

        self.clear_btn = QPushButton("清空")
        self.clear_btn.setObjectName("clearBtn")
        self.clear_btn.setFixedWidth(80)
        self.clear_btn.clicked.connect(self.clear_data)

        # 第一组：加载数据、清空
        btn_row.addWidget(self.load_btn)
        btn_row.addWidget(self.clear_btn)
        btn_row.addSpacing(10)

        # 分隔符
        separator1 = QFrame()
        separator1.setFrameShape(QFrame.VLine)
        separator1.setFrameShadow(QFrame.Sunken)
        btn_row.addWidget(separator1)
        btn_row.addSpacing(10)

        # 第二组：报表成果检查、导出Excel
        btn_row.addWidget(self.check_btn)
        btn_row.addWidget(self.export_btn)
        btn_row.addSpacing(10)

        # 分隔符
        separator2 = QFrame()
        separator2.setFrameShape(QFrame.VLine)
        separator2.setFrameShadow(QFrame.Sunken)
        btn_row.addWidget(separator2)
        btn_row.addSpacing(10)

        # 第三组：空间数据校验
        btn_row.addWidget(self.validate_btn)
        btn_row.addStretch()
        group_layout.addLayout(btn_row)

        file_group.setLayout(group_layout)
        layout.addWidget(file_group)

        # Tab区域
        self.tabs = QTabWidget()

        # 附表1 Tab
        self.fubiao1_tab = QWidget()
        self.setup_fubiao1_tab()
        self.tabs.addTab(self.fubiao1_tab, "附表1-防治对象名录")

        # 附表2 Tab
        self.fubiao2_tab = QWidget()
        self.setup_fubiao2_tab()
        self.tabs.addTab(self.fubiao2_tab, "附表2-跨沟道路桥涵")

        # 附表3 Tab
        self.fubiao3_tab = QWidget()
        self.setup_fubiao3_tab()
        self.tabs.addTab(self.fubiao3_tab, "附表3-沟滩占地")

        # 检查结果 Tab
        self.check_result_tab = QWidget()
        self.setup_check_result_tab()
        self.tabs.addTab(self.check_result_tab, "检查结果")

        layout.addWidget(self.tabs, 1)

        # 底部状态
        self.status_label = QLabel("")
        self.status_label.setStyleSheet("color: #7f8c8d; font-size: 12px;")
        layout.addWidget(self.status_label)

        self.setLayout(layout)

    def setup_fubiao1_tab(self):
        """设置附表1 Tab"""
        layout = QVBoxLayout()

        # 筛选区域
        filter_layout = QHBoxLayout()
        filter_layout.setSpacing(10)

        filter_label = QLabel("筛选:")
        filter_label.setStyleSheet("font-weight: bold; color: #2c3e50;")

        self.fubiao1_search_edit = QLineEdit()
        self.fubiao1_search_edit.setPlaceholderText("搜索名称...")
        self.fubiao1_search_edit.setFixedWidth(200)
        self.fubiao1_search_edit.textChanged.connect(self.apply_fubiao1_filter)

        clear_btn = QPushButton("清除筛选")
        clear_btn.setFixedWidth(80)
        clear_btn.clicked.connect(self.clear_fubiao1_filter)

        filter_layout.addWidget(filter_label)
        filter_layout.addWidget(self.fubiao1_search_edit)
        filter_layout.addWidget(clear_btn)
        filter_layout.addStretch()
        layout.addLayout(filter_layout)

        # 表格
        self.fubiao1_table = QTableWidget()
        self.fubiao1_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.fubiao1_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.fubiao1_table.setAlternatingRowColors(True)
        self.fubiao1_table.horizontalHeader().setStretchLastSection(False)
        self.fubiao1_table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        layout.addWidget(self.fubiao1_table)

        self.fubiao1_tab.setLayout(layout)

    def setup_fubiao2_tab(self):
        """设置附表2 Tab"""
        layout = QVBoxLayout()

        # 筛选区域
        filter_layout = QHBoxLayout()
        filter_layout.setSpacing(10)

        filter_label = QLabel("筛选:")
        filter_label.setStyleSheet("font-weight: bold; color: #2c3e50;")

        self.fubiao2_search_edit = QLineEdit()
        self.fubiao2_search_edit.setPlaceholderText("搜索名称...")
        self.fubiao2_search_edit.setFixedWidth(200)
        self.fubiao2_search_edit.textChanged.connect(self.apply_fubiao2_filter)

        clear_btn = QPushButton("清除筛选")
        clear_btn.setFixedWidth(80)
        clear_btn.clicked.connect(self.clear_fubiao2_filter)

        filter_layout.addWidget(filter_label)
        filter_layout.addWidget(self.fubiao2_search_edit)
        filter_layout.addWidget(clear_btn)
        filter_layout.addStretch()
        layout.addLayout(filter_layout)

        # 表格
        self.fubiao2_table = QTableWidget()
        self.fubiao2_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.fubiao2_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.fubiao2_table.setAlternatingRowColors(True)
        self.fubiao2_table.horizontalHeader().setStretchLastSection(False)
        self.fubiao2_table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        layout.addWidget(self.fubiao2_table)

        self.fubiao2_tab.setLayout(layout)

    def setup_fubiao3_tab(self):
        """设置附表3 Tab"""
        layout = QVBoxLayout()

        # 筛选区域
        filter_layout = QHBoxLayout()
        filter_layout.setSpacing(10)

        filter_label = QLabel("筛选:")
        filter_label.setStyleSheet("font-weight: bold; color: #2c3e50;")

        self.fubiao3_search_edit = QLineEdit()
        self.fubiao3_search_edit.setPlaceholderText("搜索名称...")
        self.fubiao3_search_edit.setFixedWidth(200)
        self.fubiao3_search_edit.textChanged.connect(self.apply_fubiao3_filter)

        clear_btn = QPushButton("清除筛选")
        clear_btn.setFixedWidth(80)
        clear_btn.clicked.connect(self.clear_fubiao3_filter)

        filter_layout.addWidget(filter_label)
        filter_layout.addWidget(self.fubiao3_search_edit)
        filter_layout.addWidget(clear_btn)
        filter_layout.addStretch()
        layout.addLayout(filter_layout)

        # 表格
        self.fubiao3_table = QTableWidget()
        self.fubiao3_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.fubiao3_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.fubiao3_table.setAlternatingRowColors(True)
        self.fubiao3_table.horizontalHeader().setStretchLastSection(False)
        self.fubiao3_table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        layout.addWidget(self.fubiao3_table)

        self.fubiao3_tab.setLayout(layout)

    def setup_check_result_tab(self):
        """设置检查结果Tab"""
        layout = QVBoxLayout()

        # 统计信息
        self.check_stats_label = QLabel('点击"报表成果检查"开始检查...')
        self.check_stats_label.setStyleSheet("font-weight: bold; color: #2c3e50; padding: 10px;")
        layout.addWidget(self.check_stats_label)

        # 筛选区域
        filter_layout = QHBoxLayout()
        filter_layout.setSpacing(10)

        filter_label = QLabel("筛选:")
        filter_label.setStyleSheet("font-weight: bold; color: #2c3e50;")

        self.check_table_combo = QComboBox()
        self.check_table_combo.addItems(["全部", "附表1", "附表2", "附表3"])
        self.check_table_combo.setFixedWidth(80)
        self.check_table_combo.currentTextChanged.connect(self.apply_check_filter)

        self.check_type_combo = QComboBox()
        self.check_type_combo.addItems(["全部", "格式错误", "一致性错误"])
        self.check_type_combo.setFixedWidth(100)
        self.check_type_combo.currentTextChanged.connect(self.apply_check_filter)

        filter_layout.addWidget(filter_label)
        filter_layout.addWidget(QLabel("表名:"))
        filter_layout.addWidget(self.check_table_combo)
        filter_layout.addWidget(QLabel("错误类型:"))
        filter_layout.addWidget(self.check_type_combo)
        filter_layout.addStretch()
        layout.addLayout(filter_layout)

        # 结果表格
        self.check_result_table = QTableWidget()
        self.check_result_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.check_result_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.check_result_table.setAlternatingRowColors(True)
        self.check_result_table.horizontalHeader().setStretchLastSection(True)
        self.check_result_table.verticalHeader().setVisible(False)
        layout.addWidget(self.check_result_table)

        # 导出按钮
        export_btn = QPushButton("导出检查报告")
        export_btn.setFixedWidth(120)
        export_btn.clicked.connect(self.export_check_report)
        layout.addWidget(export_btn, alignment=Qt.AlignRight)

        self.check_result_tab.setLayout(layout)
        self.check_errors = []  # 检查结果数据
        self.check_original_errors = []  # 原始数据用于筛选

    def select_folder(self):
        """选择文件夹"""
        folder = QFileDialog.getExistingDirectory(self, "选择成果报表文件夹")
        if folder:
            self.folder_edit.setText(folder)
            self.load_btn.setEnabled(True)

    def clear_data(self):
        """清空数据"""
        self.folder_edit.clear()
        self.report_data = None
        self.original_data = {'fubiao1': [], 'fubiao2': [], 'fubiao3': []}
        self.validation_results = None
        self.check_errors = []
        self.check_original_errors = []

        # 清空表格
        self.fubiao1_table.clear()
        self.fubiao1_table.setRowCount(0)
        self.fubiao2_table.clear()
        self.fubiao2_table.setRowCount(0)
        self.fubiao3_table.clear()
        self.fubiao3_table.setRowCount(0)
        self.check_result_table.clear()
        self.check_result_table.setRowCount(0)

        # 重置按钮状态
        self.load_btn.setEnabled(False)
        self.check_btn.setEnabled(False)
        self.export_btn.setEnabled(False)
        self.validate_btn.setEnabled(False)
        self.check_stats_label.setText('点击"报表成果检查"开始检查...')
        self.status_label.setText("")

    def load_data(self):
        """加载数据"""
        folder = self.folder_edit.text()
        if not folder:
            QMessageBox.warning(self, "警告", "请先选择文件夹")
            return

        if not os.path.exists(folder):
            QMessageBox.warning(self, "警告", "文件夹不存在")
            return

        # 加载数据
        self.report_data = load_all_reports(folder)

        # 检查是否有数据
        if self.report_data['missing']:
            QMessageBox.warning(
                self, "警告",
                f"以下文件未找到:\n" + "\n".join(self.report_data['missing'])
            )

        # 更新表格
        self.update_fubiao1_table()
        self.update_fubiao2_table()
        self.update_fubiao3_table()

        # 保存原始数据用于筛选
        self.original_data['fubiao1'] = self.report_data['fubiao1']['records']
        self.original_data['fubiao2'] = self.report_data['fubiao2']['records']
        self.original_data['fubiao3'] = self.report_data['fubiao3']['records']

        # 更新状态
        total = (
            len(self.report_data['fubiao1']['records']) +
            len(self.report_data['fubiao2']['records']) +
            len(self.report_data['fubiao3']['records'])
        )
        self.status_label.setText(f"共 {total} 条记录")
        self.load_btn.setEnabled(False)
        self.check_btn.setEnabled(total > 0)
        self.export_btn.setEnabled(total > 0)
        self.validate_btn.setEnabled(total > 0)

        QMessageBox.information(self, "完成", f"数据加载完成，共 {total} 条记录")

    def update_fubiao1_table(self, records=None):
        """更新附表1表格 - 支持3行合并表头和数据行合并"""
        if records is None:
            records = self.report_data['fubiao1']['records']

        # 获取数据行合并信息
        merge_cells = self.report_data['fubiao1'].get('merge_info', {}).get('merge_cells', [])

        # 总列数 29列
        col_count = 29

        # 第一行表头（按原始Excel结构）
        header_row1 = [
            '序号', '1.县（区、市、旗）名称', '2.县（区、市、旗）代码',
            '3.乡镇名称', '4.乡镇代码', '5.名称', '6.代码',
            '7.类型', '8.人口', '9.河流名称', '10.河流代码',
            '风险隐患要素类别', '', '', '', '', '', '', '', '', '', '', '',  # 列11-22，共12列
            '风险隐患影响类型', '', '', '', '',  # 列23-27，共5列
            '28.备注'  # 列28
        ]

        # 第二行表头
        header_row2 = [
            '', '', '', '', '', '', '', '', '', '', '',
            '跨沟道路、桥涵', '',
            '塘（堰）坝', '',
            '多支齐汇', '',
            '局地河势与微地形', '', '', '', '',
            '22.沟滩占地',
            '23.溃决', '24.壅水', '25.顶托', '26.改道', '27.漫流',
            ''
        ]

        # 第三行表头
        header_row3 = [
            '', '', '', '', '', '', '', '', '', '', '',
            '11.名称', '12.编码',
            '13.名称', '14.编码',
            '15.河流名称', '16.河流代码',
            '17.束窄', '18.急弯', '19.低洼地', '20.临河滑坡', '21.泥石流',
            '',
            '', '', '', '', '',
            ''
        ]

        # 数据列名
        data_headers = [
            '序号', '1.县（区、市、旗）名称', '2.县（区、市、旗）代码',
            '3.乡镇名称', '4.乡镇代码', '5.名称', '6.代码',
            '7.类型', '8.人口', '9.河流名称', '10.河流代码',
            '11.名称', '12.编码', '13.名称', '14.编码',
            '15.河流名称', '16.河流代码',
            '17.束窄', '18.急弯', '19.低洼地', '20.临河滑坡', '21.泥石流',
            '22.沟滩占地', '23.溃决', '24.壅水', '25.顶托', '26.改道', '27.漫流',
            '28.备注'
        ]

        self.fubiao1_table.clear()
        self.fubiao1_table.setColumnCount(col_count)
        self.fubiao1_table.setRowCount(len(records) + 3)

        # 设置三行表头内容
        for col in range(col_count):
            item1 = QTableWidgetItem(header_row1[col] if col < len(header_row1) else '')
            item1.setBackground(QColor(240, 240, 240))
            item1.setTextAlignment(Qt.AlignCenter)
            item1.setFlags(item1.flags() & ~Qt.ItemIsEditable & ~Qt.ItemIsSelectable)
            self.fubiao1_table.setItem(0, col, item1)

            item2 = QTableWidgetItem(header_row2[col] if col < len(header_row2) else '')
            item2.setBackground(QColor(240, 240, 240))
            item2.setTextAlignment(Qt.AlignCenter)
            item2.setFlags(item2.flags() & ~Qt.ItemIsEditable & ~Qt.ItemIsSelectable)
            self.fubiao1_table.setItem(1, col, item2)

            item3 = QTableWidgetItem(header_row3[col] if col < len(header_row3) else '')
            item3.setBackground(QColor(240, 240, 240))
            item3.setTextAlignment(Qt.AlignCenter)
            item3.setFlags(item3.flags() & ~Qt.ItemIsEditable & ~Qt.ItemIsSelectable)
            self.fubiao1_table.setItem(2, col, item3)

        # ========== 合并第一行表头 ==========
        for col in range(11):
            self.fubiao1_table.setSpan(0, col, 3, 1)
        self.fubiao1_table.setSpan(0, 11, 1, 12)
        self.fubiao1_table.setSpan(0, 23, 1, 5)
        self.fubiao1_table.setSpan(0, 28, 3, 1)

        # ========== 合并第二行表头 ==========
        self.fubiao1_table.setSpan(1, 11, 1, 2)
        self.fubiao1_table.setSpan(1, 13, 1, 2)
        self.fubiao1_table.setSpan(1, 15, 1, 2)
        self.fubiao1_table.setSpan(1, 17, 1, 5)
        for col in range(22, 28):
            self.fubiao1_table.setSpan(1, col, 2, 1)

        # 设置数据行
        for row, record in enumerate(records):
            for col, header in enumerate(data_headers):
                value = record.get(header, '')
                item = QTableWidgetItem(str(value) if value else '')

                # 勾选列居中显示
                if value == '☑' or value == '☒':
                    item.setTextAlignment(Qt.AlignCenter)

                self.fubiao1_table.setItem(row + 3, col, item)

        # ========== 合并数据行单元格 ==========
        for merge_info in merge_cells:
            start_row, start_col, row_span, col_span = merge_info
            # 数据行从第4行开始（索引3）
            self.fubiao1_table.setSpan(start_row + 3, start_col, row_span, col_span)

        # 隐藏默认表头
        self.fubiao1_table.horizontalHeader().hide()
        self.fubiao1_table.verticalHeader().hide()

        # 设置表头行高度
        self.fubiao1_table.setRowHeight(0, 30)
        self.fubiao1_table.setRowHeight(1, 30)
        self.fubiao1_table.setRowHeight(2, 30)

    def update_fubiao2_table(self, records=None):
        """更新附表2表格"""
        if records is None:
            records = self.report_data['fubiao2']['records']
        headers = self.report_data['fubiao2']['headers']

        self.fubiao2_table.clear()
        self.fubiao2_table.setColumnCount(len(headers))
        self.fubiao2_table.setHorizontalHeaderLabels(headers)
        self.fubiao2_table.setRowCount(len(records))

        for row, record in enumerate(records):
            for col, header in enumerate(headers):
                value = record.get(header, '')
                item = QTableWidgetItem(str(value) if value else '')
                self.fubiao2_table.setItem(row, col, item)

        self.fubiao2_table.resizeColumnsToContents()

    def update_fubiao3_table(self, records=None):
        """更新附表3表格"""
        if records is None:
            records = self.report_data['fubiao3']['records']
        headers = self.report_data['fubiao3']['headers']

        self.fubiao3_table.clear()
        self.fubiao3_table.setColumnCount(len(headers))
        self.fubiao3_table.setHorizontalHeaderLabels(headers)
        self.fubiao3_table.setRowCount(len(records))

        for row, record in enumerate(records):
            for col, header in enumerate(headers):
                value = record.get(header, '')
                item = QTableWidgetItem(str(value) if value else '')
                self.fubiao3_table.setItem(row, col, item)

        self.fubiao3_table.resizeColumnsToContents()

    def apply_fubiao1_filter(self):
        """应用附表1筛选"""
        search_text = self.fubiao1_search_edit.text().strip().lower()
        if not search_text:
            self.update_fubiao1_table(self.original_data['fubiao1'])
            return

        filtered = []
        for record in self.original_data['fubiao1']:
            # 搜索名称列
            name = str(record.get('5.名称', '')).lower()
            if search_text in name:
                filtered.append(record)

        self.update_fubiao1_table(filtered)

    def clear_fubiao1_filter(self):
        """清除附表1筛选"""
        self.fubiao1_search_edit.clear()
        self.update_fubiao1_table(self.original_data['fubiao1'])

    def apply_fubiao2_filter(self):
        """应用附表2筛选"""
        search_text = self.fubiao2_search_edit.text().strip().lower()
        if not search_text:
            self.update_fubiao2_table(self.original_data['fubiao2'])
            return

        filtered = []
        for record in self.original_data['fubiao2']:
            name = str(record.get('5.名称', '')).lower()
            if search_text in name:
                filtered.append(record)

        self.update_fubiao2_table(filtered)

    def clear_fubiao2_filter(self):
        """清除附表2筛选"""
        self.fubiao2_search_edit.clear()
        self.update_fubiao2_table(self.original_data['fubiao2'])

    def apply_fubiao3_filter(self):
        """应用附表3筛选"""
        search_text = self.fubiao3_search_edit.text().strip().lower()
        if not search_text:
            self.update_fubiao3_table(self.original_data['fubiao3'])
            return

        filtered = []
        for record in self.original_data['fubiao3']:
            name = str(record.get('5.名称', '')).lower()
            if search_text in name:
                filtered.append(record)

        self.update_fubiao3_table(filtered)

    def clear_fubiao3_filter(self):
        """清除附表3筛选"""
        self.fubiao3_search_edit.clear()
        self.update_fubiao3_table(self.original_data['fubiao3'])

    def export_excel(self):
        """导出Excel - 保持原始格式（3行表头、合并单元格）"""
        if not self.report_data:
            QMessageBox.warning(self, "警告", "没有可导出的数据")
            return

        file, _ = QFileDialog.getSaveFileName(
            self, "导出Excel", "成果报表.xlsx", "Excel Files (*.xlsx)"
        )

        if file:
            try:
                import pandas as pd
                from openpyxl import Workbook
                from openpyxl.styles import Alignment, Font, Border, Side
                from openpyxl.utils import get_column_letter

                wb = Workbook()

                # ========== 导出附表1 ==========
                ws1 = wb.active
                ws1.title = '附表1-防治对象名录'

                records = self.report_data['fubiao1']['records']
                merge_cells = self.report_data['fubiao1'].get('merge_info', {}).get('merge_cells', [])

                # 第一行表头
                header_row1 = [
                    '序号', '1.县（区、市、旗）名称', '2.县（区、市、旗）代码',
                    '3.乡镇名称', '4.乡镇代码', '5.名称', '6.代码',
                    '7.类型', '8.人口', '9.河流名称', '10.河流代码',
                    '风险隐患要素类别', '', '', '', '', '', '', '', '', '', '', '',
                    '风险隐患影响类型', '', '', '', '',
                    '28.备注'
                ]

                # 第二行表头
                header_row2 = [
                    '', '', '', '', '', '', '', '', '', '', '',
                    '跨沟道路、桥涵', '',
                    '塘（堰）坝', '',
                    '多支齐汇', '',
                    '局地河势与微地形', '', '', '', '',
                    '22.沟滩占地',
                    '23.溃决', '24.壅水', '25.顶托', '26.改道', '27.漫流',
                    ''
                ]

                # 第三行表头
                header_row3 = [
                    '', '', '', '', '', '', '', '', '', '', '',
                    '11.名称', '12.编码',
                    '13.名称', '14.编码',
                    '15.河流名称', '16.河流代码',
                    '17.束窄', '18.急弯', '19.低洼地', '20.临河滑坡', '21.泥石流',
                    '',
                    '', '', '', '', '',
                    ''
                ]

                # 数据列名
                data_headers = [
                    '序号', '1.县（区、市、旗）名称', '2.县（区、市、旗）代码',
                    '3.乡镇名称', '4.乡镇代码', '5.名称', '6.代码',
                    '7.类型', '8.人口', '9.河流名称', '10.河流代码',
                    '11.名称', '12.编码', '13.名称', '14.编码',
                    '15.河流名称', '16.河流代码',
                    '17.束窄', '18.急弯', '19.低洼地', '20.临河滑坡', '21.泥石流',
                    '22.沟滩占地', '23.溃决', '24.壅水', '25.顶托', '26.改道', '27.漫流',
                    '28.备注'
                ]

                # 写入表头
                ws1.append(header_row1)
                ws1.append(header_row2)
                ws1.append(header_row3)

                # 写入数据
                for record in records:
                    row_data = []
                    for header in data_headers:
                        val = record.get(header, '')
                        # 转换为Wingdings 2字体的字符
                        if val == '☑':
                            val = 'R'  # Wingdings 2 的 R 是带方框对号
                        elif val == '☒':
                            val = 'S'  # Wingdings 2 的 S 是带方框叉号
                        row_data.append(val)
                    ws1.append(row_data)

                # 合并第一行表头
                for col in range(1, 12):  # A1-K1 (列1-11)
                    ws1.merge_cells(start_row=1, start_column=col, end_row=3, end_column=col)
                ws1.merge_cells(start_row=1, start_column=12, end_row=1, end_column=23)  # L1-W1
                ws1.merge_cells(start_row=1, start_column=24, end_row=1, end_column=28)  # X1-AB1
                ws1.merge_cells(start_row=1, start_column=29, end_row=3, end_column=29)  # AC1-AC3

                # 合并第二行表头
                ws1.merge_cells(start_row=2, start_column=12, end_row=2, end_column=13)  # L2-M2
                ws1.merge_cells(start_row=2, start_column=14, end_row=2, end_column=15)  # N2-O2
                ws1.merge_cells(start_row=2, start_column=16, end_row=2, end_column=17)  # P2-Q2
                ws1.merge_cells(start_row=2, start_column=18, end_row=2, end_column=22)  # R2-V2
                for col in range(23, 29):  # W2-AB2
                    ws1.merge_cells(start_row=2, start_column=col, end_row=3, end_column=col)

                # 合并数据行单元格
                for merge_info in merge_cells:
                    start_row, start_col, row_span, col_span = merge_info
                    # 数据行从第4行开始（Excel行号为start_row + 4）
                    ws1.merge_cells(
                        start_row=start_row + 4,
                        start_column=start_col + 1,
                        end_row=start_row + 4 + row_span - 1,
                        end_column=start_col + col_span
                    )

                # 设置表头样式
                header_font = Font(bold=True)
                header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

                for row in ws1.iter_rows(min_row=1, max_row=3, max_col=29):
                    for cell in row:
                        cell.font = header_font
                        cell.alignment = header_alignment
                        cell.border = thin_border

                # 设置数据行边框和勾选列字体
                wingdings2_font = Font(name='Wingdings 2')
                # 勾选列 (Excel列号17-27): 束窄、急弯、低洼地、临河滑坡、泥石流、沟滩占地、溃决、壅水、顶托、改道、漫流
                checkbox_cols = list(range(17, 28))  # Excel 1-based 列号

                for row in ws1.iter_rows(min_row=4, max_row=3 + len(records), max_col=29):
                    for cell in row:
                        cell.border = thin_border
                        cell.alignment = Alignment(vertical='center')
                        # 勾选列设置Wingdings 2字体
                        if cell.column in checkbox_cols:
                            cell.font = wingdings2_font
                            cell.alignment = Alignment(horizontal='center', vertical='center')

                # ========== 导出附表2 ==========
                ws2 = wb.create_sheet('附表2-跨沟道路桥涵')
                records2 = self.report_data['fubiao2']['records']
                headers2 = self.report_data['fubiao2']['headers']

                ws2.append(headers2)
                for record in records2:
                    row_data = [record.get(h, '') for h in headers2]
                    ws2.append(row_data)

                # 设置边框
                for row in ws2.iter_rows(min_row=1, max_row=1 + len(records2), max_col=len(headers2)):
                    for cell in row:
                        cell.border = thin_border
                        if cell.row == 1:
                            cell.font = header_font
                            cell.alignment = header_alignment

                # ========== 导出附表3 ==========
                ws3 = wb.create_sheet('附表3-沟滩占地')
                records3 = self.report_data['fubiao3']['records']
                headers3 = self.report_data['fubiao3']['headers']

                ws3.append(headers3)
                for record in records3:
                    row_data = [record.get(h, '') for h in headers3]
                    ws3.append(row_data)

                # 设置边框
                for row in ws3.iter_rows(min_row=1, max_row=1 + len(records3), max_col=len(headers3)):
                    for cell in row:
                        cell.border = thin_border
                        if cell.row == 1:
                            cell.font = header_font
                            cell.alignment = header_alignment

                wb.save(file)
                QMessageBox.information(self, "完成", f"已导出到:\n{file}")

            except Exception as e:
                QMessageBox.critical(self, "错误", f"导出失败:\n{str(e)}")

    def run_data_check(self):
        """执行数据检查"""
        if not self.report_data:
            QMessageBox.warning(self, "警告", "请先加载数据")
            return

        from utils.data_checker import DataChecker
        checker = DataChecker(self.report_data)
        self.check_errors = checker.check_all()
        self.check_original_errors = self.check_errors.copy()

        # 更新统计
        format_count = sum(1 for e in self.check_errors if e['错误类型'] == '格式错误')
        consist_count = sum(1 for e in self.check_errors if e['错误类型'] == '一致性错误')
        self.check_stats_label.setText(
            f"共发现 {len(self.check_errors)} 个问题 | 格式错误: {format_count} | 一致性错误: {consist_count}"
        )

        # 显示结果
        self.display_check_results(self.check_errors)

    def display_check_results(self, errors):
        """显示检查结果"""
        headers = ['序号', '表名', '字段名', '错误类型', '错误描述', '当前值']
        self.check_result_table.clear()
        self.check_result_table.setColumnCount(len(headers))
        self.check_result_table.setHorizontalHeaderLabels(headers)
        self.check_result_table.setRowCount(len(errors))

        for row, err in enumerate(errors):
            for col, header in enumerate(headers):
                value = err.get(header, '')
                item = QTableWidgetItem(str(value))
                if header == '错误类型':
                    if value == '格式错误':
                        item.setForeground(QColor('#e67e22'))
                    elif value == '一致性错误':
                        item.setForeground(QColor('#e74c3c'))
                self.check_result_table.setItem(row, col, item)

        self.check_result_table.resizeColumnsToContents()

    def apply_check_filter(self):
        """应用检查结果筛选"""
        table_filter = self.check_table_combo.currentText()
        type_filter = self.check_type_combo.currentText()

        filtered = []
        for err in self.check_original_errors:
            if table_filter != "全部" and err['表名'] != table_filter:
                continue
            if type_filter != "全部" and err['错误类型'] != type_filter:
                continue
            filtered.append(err)

        self.check_errors = filtered
        self.display_check_results(filtered)

    def export_check_report(self):
        """导出检查报告 - 包含附表原始数据和检查结果，错误字段标红"""
        if not self.report_data:
            QMessageBox.warning(self, "警告", "没有数据可导出")
            return

        file, _ = QFileDialog.getSaveFileName(
            self, "导出检查报告", "数据检查报告.xlsx", "Excel Files (*.xlsx)"
        )

        if file:
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
                from openpyxl.utils import get_column_letter

                wb = Workbook()

                header_font = Font(bold=True)
                header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
                error_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")  # 浅红色
                header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

                # 构建检查结果索引 {表名: {行号: [(字段名, 错误描述), ...]}}
                error_index = {'附表1': {}, '附表2': {}, '附表3': {}}
                for err in self.check_errors:
                    table = err['表名']
                    row_num = err['序号']
                    field = err['字段名']
                    desc = err['错误描述']
                    if table not in error_index:
                        error_index[table] = {}
                    if row_num not in error_index[table]:
                        error_index[table][row_num] = []
                    error_index[table][row_num].append((field, desc))

                # ========== 导出附表1 ==========
                ws1 = wb.active
                ws1.title = '附表1-防治对象名录'

                records = self.report_data['fubiao1']['records']
                merge_cells = self.report_data['fubiao1'].get('merge_info', {}).get('merge_cells', [])

                # 附表1字段映射：字段名 -> 列号（用于标红）
                fubiao1_field_col_map = {
                    '县代码': 3, '乡镇代码': 5, '跨沟道路桥涵': [12, 13]
                }

                # 表头定义（增加"检查结果"列）
                header_row1 = [
                    '序号', '1.县（区、市、旗）名称', '2.县（区、市、旗）代码',
                    '3.乡镇名称', '4.乡镇代码', '5.名称', '6.代码',
                    '7.类型', '8.人口', '9.河流名称', '10.河流代码',
                    '风险隐患要素类别', '', '', '', '', '', '', '', '', '', '', '',
                    '风险隐患影响类型', '', '', '', '',
                    '28.备注', '检查结果'
                ]
                header_row2 = [
                    '', '', '', '', '', '', '', '', '', '', '',
                    '跨沟道路、桥涵', '',
                    '塘（堰）坝', '',
                    '多支齐汇', '',
                    '局地河势与微地形', '', '', '', '',
                    '22.沟滩占地',
                    '23.溃决', '24.壅水', '25.顶托', '26.改道', '27.漫流',
                    '', ''
                ]
                header_row3 = [
                    '', '', '', '', '', '', '', '', '', '', '',
                    '11.名称', '12.编码',
                    '13.名称', '14.编码',
                    '15.河流名称', '16.河流代码',
                    '17.束窄', '18.急弯', '19.低洼地', '20.临河滑坡', '21.泥石流',
                    '',
                    '', '', '', '', '',
                    '', ''
                ]
                data_headers = [
                    '序号', '1.县（区、市、旗）名称', '2.县（区、市、旗）代码',
                    '3.乡镇名称', '4.乡镇代码', '5.名称', '6.代码',
                    '7.类型', '8.人口', '9.河流名称', '10.河流代码',
                    '11.名称', '12.编码', '13.名称', '14.编码',
                    '15.河流名称', '16.河流代码',
                    '17.束窄', '18.急弯', '19.低洼地', '20.临河滑坡', '21.泥石流',
                    '22.沟滩占地', '23.溃决', '24.壅水', '25.顶托', '26.改道', '27.漫流',
                    '28.备注'
                ]

                ws1.append(header_row1)
                ws1.append(header_row2)
                ws1.append(header_row3)

                # 写入数据并标红错误字段
                for i, record in enumerate(records):
                    row_num = i + 1  # 行号从1开始
                    row_data = []
                    for header in data_headers:
                        val = record.get(header, '')
                        if val == '☑':
                            val = 'R'
                        elif val == '☒':
                            val = 'S'
                        row_data.append(val)

                    # 添加检查结果列
                    errors = error_index.get('附表1', {}).get(row_num, [])
                    check_result = '；'.join([f"{f}:{d}" for f, d in errors]) if errors else ''
                    row_data.append(check_result)

                    ws1.append(row_data)

                    # 标红错误字段
                    if errors:
                        excel_row = i + 4  # Excel行号从4开始（3行表头）
                        for field, desc in errors:
                            if field == '县代码':
                                ws1.cell(row=excel_row, column=3).fill = error_fill
                            elif field == '乡镇代码':
                                ws1.cell(row=excel_row, column=5).fill = error_fill
                            elif field == '跨沟道路桥涵':
                                ws1.cell(row=excel_row, column=12).fill = error_fill
                                ws1.cell(row=excel_row, column=13).fill = error_fill
                        # 标红检查结果列
                        if check_result:
                            ws1.cell(row=excel_row, column=30).fill = error_fill

                # 合并表头（更新列数）
                for col in range(1, 12):
                    ws1.merge_cells(start_row=1, start_column=col, end_row=3, end_column=col)
                ws1.merge_cells(start_row=1, start_column=12, end_row=1, end_column=23)
                ws1.merge_cells(start_row=1, start_column=24, end_row=1, end_column=28)
                ws1.merge_cells(start_row=1, start_column=29, end_row=3, end_column=29)
                ws1.merge_cells(start_row=1, start_column=30, end_row=3, end_column=30)  # 检查结果列
                ws1.merge_cells(start_row=2, start_column=12, end_row=2, end_column=13)
                ws1.merge_cells(start_row=2, start_column=14, end_row=2, end_column=15)
                ws1.merge_cells(start_row=2, start_column=16, end_row=2, end_column=17)
                ws1.merge_cells(start_row=2, start_column=18, end_row=2, end_column=22)
                for col in range(23, 29):
                    ws1.merge_cells(start_row=2, start_column=col, end_row=3, end_column=col)

                # 合并数据行
                for merge_info in merge_cells:
                    start_row, start_col, row_span, col_span = merge_info
                    ws1.merge_cells(
                        start_row=start_row + 4,
                        start_column=start_col + 1,
                        end_row=start_row + 4 + row_span - 1,
                        end_column=start_col + col_span
                    )

                for row in ws1.iter_rows(min_row=1, max_row=3, max_col=30):
                    for cell in row:
                        cell.font = header_font
                        cell.alignment = header_alignment
                        cell.border = thin_border

                wingdings2_font = Font(name='Wingdings 2')
                checkbox_cols = list(range(17, 28))
                for row in ws1.iter_rows(min_row=4, max_row=3 + len(records), max_col=30):
                    for cell in row:
                        cell.border = thin_border
                        cell.alignment = Alignment(vertical='center')
                        if cell.column in checkbox_cols:
                            cell.font = wingdings2_font
                            cell.alignment = Alignment(horizontal='center', vertical='center')

                # ========== 导出附表2 ==========
                ws2 = wb.create_sheet('附表2-跨沟道路桥涵')
                records2 = self.report_data['fubiao2']['records']
                headers2 = self.report_data['fubiao2']['headers'] + ['检查结果']

                ws2.append(headers2)
                for i, record in enumerate(records2):
                    row_num = i + 1
                    row_data = [record.get(h, '') for h in self.report_data['fubiao2']['headers']]

                    # 添加检查结果
                    errors = error_index.get('附表2', {}).get(row_num, [])
                    check_result = '；'.join([f"{f}:{d}" for f, d in errors]) if errors else ''
                    row_data.append(check_result)

                    ws2.append(row_data)

                    # 标红错误字段
                    if errors:
                        excel_row = i + 2
                        error_fields = [f for f, d in errors]
                        for col_idx, h in enumerate(self.report_data['fubiao2']['headers'], 1):
                            # 检查该字段是否在错误中
                            for ef in error_fields:
                                if ef in h or h in ef:
                                    ws2.cell(row=excel_row, column=col_idx).fill = error_fill
                                    break
                        # 标红检查结果列
                        if check_result:
                            ws2.cell(row=excel_row, column=len(headers2)).fill = error_fill

                for row in ws2.iter_rows(min_row=1, max_row=1 + len(records2), max_col=len(headers2)):
                    for cell in row:
                        cell.border = thin_border
                        if cell.row == 1:
                            cell.font = header_font
                            cell.alignment = header_alignment

                # ========== 导出附表3 ==========
                ws3 = wb.create_sheet('附表3-沟滩占地')
                records3 = self.report_data['fubiao3']['records']
                headers3 = self.report_data['fubiao3']['headers'] + ['检查结果']

                ws3.append(headers3)
                for i, record in enumerate(records3):
                    row_num = i + 1
                    row_data = [record.get(h, '') for h in self.report_data['fubiao3']['headers']]

                    # 添加检查结果
                    errors = error_index.get('附表3', {}).get(row_num, [])
                    check_result = '；'.join([f"{f}:{d}" for f, d in errors]) if errors else ''
                    row_data.append(check_result)

                    ws3.append(row_data)

                    # 标红错误字段
                    if errors:
                        excel_row = i + 2
                        error_fields = [f for f, d in errors]
                        for col_idx, h in enumerate(self.report_data['fubiao3']['headers'], 1):
                            for ef in error_fields:
                                if ef in h or h in ef:
                                    ws3.cell(row=excel_row, column=col_idx).fill = error_fill
                                    break
                        if check_result:
                            ws3.cell(row=excel_row, column=len(headers3)).fill = error_fill

                for row in ws3.iter_rows(min_row=1, max_row=1 + len(records3), max_col=len(headers3)):
                    for cell in row:
                        cell.border = thin_border
                        if cell.row == 1:
                            cell.font = header_font
                            cell.alignment = header_alignment

                # ========== 导出检查结果汇总 ==========
                ws_check = wb.create_sheet("检查结果汇总")
                check_headers = ['序号', '表名', '字段名', '错误类型', '错误描述', '当前值']
                ws_check.append(check_headers)
                for cell in ws_check[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = thin_border
                    cell.alignment = header_alignment

                for err in self.check_errors:
                    ws_check.append([
                        err['序号'],
                        err['表名'],
                        err['字段名'],
                        err['错误类型'],
                        err['错误描述'],
                        err['当前值']
                    ])

                for row in ws_check.iter_rows(min_row=2, max_row=ws_check.max_row, max_col=len(check_headers)):
                    for cell in row:
                        cell.border = thin_border

                wb.save(file)
                QMessageBox.information(self, "完成", f"已导出到:\n{file}")

            except Exception as e:
                import traceback
                traceback.print_exc()
                QMessageBox.critical(self, "错误", f"导出失败:\n{str(e)}")

    def show_validation_dialog(self):
        """显示空间数据校验弹窗"""
        if not self.report_data:
            QMessageBox.warning(self, "警告", "请先加载附表数据")
            return

        # 如果已有弹窗则激活，否则创建新的
        if self.result_dialog is not None and self.result_dialog.isVisible():
            self.result_dialog.raise_()
            self.result_dialog.activateWindow()
            return

        # 创建弹窗（非模态，使用show而非exec）
        self.result_dialog = ValidationDialog(self)
        self.result_dialog.set_report_data(self.report_data)
        self.result_dialog.show()


class ValidationDialog(QDialog):
    """空间数据校验弹窗"""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("空间数据校验")
        self.setMinimumSize(900, 600)
        # 设置为非模态窗口，允许底层交互
        self.setWindowFlags(Qt.Window | Qt.WindowMinMaxButtonsHint | Qt.WindowCloseButtonHint)
        self.setWindowModality(Qt.NonModal)  # 非模态，允许底层窗口交互
        self.report_data = None
        self.validator = None
        self.validation_results = None
        self.fubiao1_detail_data = []  # 附表1详情数据
        self.fubiao23_detail_data = []  # 附表2/3详情数据
        self.fubiao1_original_data = []  # 附表1原始数据用于筛选
        self.fubiao23_original_data = []  # 附表2/3原始数据用于筛选
        import copy
        self.field_mapping = copy.deepcopy(DataValidator.DEFAULT_FIELD_MAPPING)  # 字段映射配置（深拷贝）
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout()

        # 数据加载区域（白色背景容器）
        load_group = QGroupBox("数据加载")
        load_group.setStyleSheet("""
            QGroupBox {
                background-color: white;
                border: 1px solid #e0e4e8;
                border-radius: 10px;
                margin-top: 12px;
                padding: 15px;
                font-weight: bold;
                color: #34495e;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                subcontrol-position: top left;
                padding: 0 10px;
                color: #6b7d9e;
            }
        """)
        group_layout = QVBoxLayout()
        group_layout.setSpacing(10)

        # 第一行：文件夹选择 + 浏览按钮
        file_row = QHBoxLayout()
        shp_label = QLabel("空间数据文件夹:")
        shp_label.setStyleSheet("font-weight: bold; color: #2c3e50;")
        self.shp_folder_edit = QLineEdit()
        self.shp_folder_edit.setPlaceholderText("选择包含shp文件的文件夹...")
        self.shp_folder_edit.setReadOnly(True)
        self.browse_btn = QPushButton("浏览...")
        self.browse_btn.setFixedWidth(80)
        self.browse_btn.clicked.connect(self.select_shp_folder)

        file_row.addWidget(shp_label)
        file_row.addWidget(self.shp_folder_edit, 1)
        file_row.addWidget(self.browse_btn)
        group_layout.addLayout(file_row)

        # 第二行：按钮行
        btn_row = QHBoxLayout()
        btn_row.addStretch()

        self.mapping_btn = QPushButton("字段映射")
        self.mapping_btn.setFixedWidth(100)
        self.mapping_btn.setEnabled(False)  # 默认禁用，选择文件夹后启用
        self.mapping_btn.clicked.connect(self.show_field_mapping)

        self.validate_btn = QPushButton("开始校验")
        self.validate_btn.setFixedWidth(100)
        self.validate_btn.clicked.connect(self.start_validation)
        self.validate_btn.setEnabled(False)

        self.clear_btn = QPushButton("清空")
        self.clear_btn.setObjectName("clearBtn")
        self.clear_btn.setFixedWidth(80)
        self.clear_btn.clicked.connect(self.clear_results)

        self.export_btn = QPushButton("导出报告")
        self.export_btn.setObjectName("exportBtn")
        self.export_btn.setFixedWidth(100)
        self.export_btn.clicked.connect(self.export_report)
        self.export_btn.setEnabled(False)

        self.log_toggle_btn = QPushButton("显示日志")
        self.log_toggle_btn.setObjectName("logToggleBtn")
        self.log_toggle_btn.setFixedWidth(80)
        self.log_toggle_btn.clicked.connect(self.toggle_log)

        # 第一组：开始校验、清空
        btn_row.addWidget(self.validate_btn)
        btn_row.addWidget(self.clear_btn)
        btn_row.addSpacing(10)

        # 分隔符
        separator1 = QFrame()
        separator1.setFrameShape(QFrame.VLine)
        separator1.setFrameShadow(QFrame.Sunken)
        btn_row.addWidget(separator1)
        btn_row.addSpacing(10)

        # 第二组：字段映射
        btn_row.addWidget(self.mapping_btn)
        btn_row.addSpacing(10)

        # 分隔符
        separator2 = QFrame()
        separator2.setFrameShape(QFrame.VLine)
        separator2.setFrameShadow(QFrame.Sunken)
        btn_row.addWidget(separator2)
        btn_row.addSpacing(10)

        # 第三组：导出报告
        btn_row.addWidget(self.export_btn)
        btn_row.addSpacing(10)

        # 分隔符
        separator3 = QFrame()
        separator3.setFrameShape(QFrame.VLine)
        separator3.setFrameShadow(QFrame.Sunken)
        btn_row.addWidget(separator3)
        btn_row.addSpacing(10)

        # 第四组：显示日志
        btn_row.addWidget(self.log_toggle_btn)
        btn_row.addStretch()
        group_layout.addLayout(btn_row)

        load_group.setLayout(group_layout)
        layout.addWidget(load_group)

        # 主内容区域（结果 + 日志）
        main_splitter = QSplitter(Qt.Horizontal)

        # 结果Tab区域
        self.result_tabs = QTabWidget()

        # 附表1对比结果Tab
        self.fubiao1_result_tab = QWidget()
        self.setup_fubiao1_result_tab()
        self.result_tabs.addTab(self.fubiao1_result_tab, "附表1 ↔ 防治对象分布P")

        # 附表2/3对比结果Tab
        self.fubiao23_result_tab = QWidget()
        self.setup_fubiao23_result_tab()
        self.result_tabs.addTab(self.fubiao23_result_tab, "附表2/3 ↔ 隐患要素分布L")

        main_splitter.addWidget(self.result_tabs)

        # 日志区域（右侧滑入）
        log_widget = QWidget()
        log_layout = QVBoxLayout()
        log_layout.setContentsMargins(5, 5, 5, 5)

        log_title = QLabel("校验日志")
        log_title.setStyleSheet("font-weight: bold; color: #2c3e50;")
        log_layout.addWidget(log_title)

        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)
        self.log_text.setStyleSheet("""
            QTextEdit {
                background-color: #1e1e1e;
                color: #d4d4d4;
                font-family: Consolas, Monaco, monospace;
                font-size: 12px;
                border: 1px solid #3c3c3c;
                border-radius: 5px;
                padding: 5px;
            }
        """)
        log_layout.addWidget(self.log_text)
        log_widget.setLayout(log_layout)

        self.log_panel = log_widget
        main_splitter.addWidget(self.log_panel)

        # 保存splitter引用
        self.main_splitter = main_splitter

        # 默认隐藏日志
        self.log_panel.hide()
        self.log_visible = False

        layout.addWidget(main_splitter, 1)

        self.setLayout(layout)

    def setup_fubiao1_result_tab(self):
        """设置附表1对比结果Tab"""
        layout = QVBoxLayout()

        self.fubiao1_stats_label = QLabel("等待校验...")
        self.fubiao1_stats_label.setStyleSheet("font-weight: bold; color: #2c3e50; padding: 10px;")
        layout.addWidget(self.fubiao1_stats_label)

        # 筛选区域
        filter_layout = QHBoxLayout()
        filter_layout.setSpacing(10)

        filter_label = QLabel("筛选:")
        filter_label.setStyleSheet("font-weight: bold; color: #2c3e50;")

        self.fubiao1_status_combo = QComboBox()
        self.fubiao1_status_combo.addItems(["全部", "✓ 匹配", "⚠ 仅附表", "⚠ 仅shp"])
        self.fubiao1_status_combo.setFixedWidth(100)
        self.fubiao1_status_combo.currentTextChanged.connect(self.apply_fubiao1_result_filter)

        self.fubiao1_name_edit = QLineEdit()
        self.fubiao1_name_edit.setPlaceholderText("名称搜索...")
        self.fubiao1_name_edit.setFixedWidth(150)
        self.fubiao1_name_edit.textChanged.connect(self.apply_fubiao1_result_filter)

        # 是否完全匹配筛选
        match_label = QLabel("是否完全匹配:")
        match_label.setStyleSheet("font-weight: bold; color: #2c3e50;")
        self.fubiao1_match_combo = QComboBox()
        self.fubiao1_match_combo.addItems(["全部", "是", "否"])
        self.fubiao1_match_combo.setFixedWidth(80)
        self.fubiao1_match_combo.currentTextChanged.connect(self.apply_fubiao1_result_filter)

        clear_filter_btn = QPushButton("清除筛选")
        clear_filter_btn.setFixedWidth(80)
        clear_filter_btn.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #95a5a6, stop:1 #7f8c8d);
                padding: 5px 10px;
                font-size: 12px;
                min-width: 60px;
            }
        """)
        clear_filter_btn.clicked.connect(self.clear_fubiao1_result_filter)

        filter_layout.addWidget(filter_label)
        filter_layout.addWidget(self.fubiao1_status_combo)
        filter_layout.addWidget(match_label)
        filter_layout.addWidget(self.fubiao1_match_combo)
        filter_layout.addWidget(QLabel("名称:"))
        filter_layout.addWidget(self.fubiao1_name_edit)
        filter_layout.addWidget(clear_filter_btn)
        filter_layout.addStretch()
        layout.addLayout(filter_layout)

        # 使用Splitter分割结果表格和详情区域
        splitter = QSplitter(Qt.Vertical)

        # 结果表格
        self.fubiao1_result_table = QTableWidget()
        self.fubiao1_result_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.fubiao1_result_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.fubiao1_result_table.setSelectionMode(QAbstractItemView.SingleSelection)
        self.fubiao1_result_table.setAlternatingRowColors(True)
        self.fubiao1_result_table.horizontalHeader().setStretchLastSection(True)
        self.fubiao1_result_table.verticalHeader().setVisible(False)  # 隐藏行号
        self.fubiao1_result_table.itemSelectionChanged.connect(self.on_fubiao1_selection_changed)
        splitter.addWidget(self.fubiao1_result_table)

        # 详情区域（横向对比）
        detail_widget = QWidget()
        detail_layout = QVBoxLayout()
        detail_layout.setContentsMargins(5, 5, 5, 5)

        detail_title = QLabel("字段详情对比")
        detail_title.setStyleSheet("font-weight: bold; color: #2c3e50;")
        detail_layout.addWidget(detail_title)

        # 横向对比表格：两行（附表、shp），多列（各字段）
        self.fubiao1_detail_table = QTableWidget()
        self.fubiao1_detail_table.setRowCount(2)
        self.fubiao1_detail_table.setVerticalHeaderLabels(['附表', 'shp'])
        self.fubiao1_detail_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.fubiao1_detail_table.setAlternatingRowColors(True)
        self.fubiao1_detail_table.horizontalHeader().setStretchLastSection(True)
        self.fubiao1_detail_table.verticalHeader().setDefaultAlignment(Qt.AlignCenter)
        detail_layout.addWidget(self.fubiao1_detail_table)

        detail_widget.setLayout(detail_layout)
        splitter.addWidget(detail_widget)

        splitter.setSizes([300, 150])
        layout.addWidget(splitter)

        self.fubiao1_result_tab.setLayout(layout)

    def setup_fubiao23_result_tab(self):
        """设置附表2/3对比结果Tab"""
        layout = QVBoxLayout()

        self.fubiao23_stats_label = QLabel("等待校验...")
        self.fubiao23_stats_label.setStyleSheet("font-weight: bold; color: #2c3e50; padding: 10px;")
        layout.addWidget(self.fubiao23_stats_label)

        # 筛选区域
        filter_layout = QHBoxLayout()
        filter_layout.setSpacing(10)

        filter_label = QLabel("筛选:")
        filter_label.setStyleSheet("font-weight: bold; color: #2c3e50;")

        self.fubiao23_status_combo = QComboBox()
        self.fubiao23_status_combo.addItems(["全部", "✓ 匹配", "⚠ 仅附表", "⚠ 仅shp"])
        self.fubiao23_status_combo.setFixedWidth(100)
        self.fubiao23_status_combo.currentTextChanged.connect(self.apply_fubiao23_result_filter)

        self.fubiao23_name_edit = QLineEdit()
        self.fubiao23_name_edit.setPlaceholderText("名称搜索...")
        self.fubiao23_name_edit.setFixedWidth(150)
        self.fubiao23_name_edit.textChanged.connect(self.apply_fubiao23_result_filter)

        # 是否完全匹配筛选
        match_label = QLabel("是否完全匹配:")
        match_label.setStyleSheet("font-weight: bold; color: #2c3e50;")
        self.fubiao23_match_combo = QComboBox()
        self.fubiao23_match_combo.addItems(["全部", "是", "否"])
        self.fubiao23_match_combo.setFixedWidth(80)
        self.fubiao23_match_combo.currentTextChanged.connect(self.apply_fubiao23_result_filter)

        clear_filter_btn = QPushButton("清除筛选")
        clear_filter_btn.setFixedWidth(80)
        clear_filter_btn.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #95a5a6, stop:1 #7f8c8d);
                padding: 5px 10px;
                font-size: 12px;
                min-width: 60px;
            }
        """)
        clear_filter_btn.clicked.connect(self.clear_fubiao23_result_filter)

        filter_layout.addWidget(filter_label)
        filter_layout.addWidget(self.fubiao23_status_combo)
        filter_layout.addWidget(match_label)
        filter_layout.addWidget(self.fubiao23_match_combo)
        filter_layout.addWidget(QLabel("名称:"))
        filter_layout.addWidget(self.fubiao23_name_edit)
        filter_layout.addWidget(clear_filter_btn)
        filter_layout.addStretch()
        layout.addLayout(filter_layout)

        # 使用Splitter分割结果表格和详情区域
        splitter = QSplitter(Qt.Vertical)

        # 结果表格
        self.fubiao23_result_table = QTableWidget()
        self.fubiao23_result_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.fubiao23_result_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.fubiao23_result_table.setSelectionMode(QAbstractItemView.SingleSelection)
        self.fubiao23_result_table.setAlternatingRowColors(True)
        self.fubiao23_result_table.horizontalHeader().setStretchLastSection(True)
        self.fubiao23_result_table.verticalHeader().setVisible(False)  # 隐藏行号
        self.fubiao23_result_table.itemSelectionChanged.connect(self.on_fubiao23_selection_changed)
        splitter.addWidget(self.fubiao23_result_table)

        # 详情区域（横向对比）
        detail_widget = QWidget()
        detail_layout = QVBoxLayout()
        detail_layout.setContentsMargins(5, 5, 5, 5)

        detail_title = QLabel("字段详情对比")
        detail_title.setStyleSheet("font-weight: bold; color: #2c3e50;")
        detail_layout.addWidget(detail_title)

        # 横向对比表格：两行（附表、shp），多列（各字段）
        self.fubiao23_detail_table = QTableWidget()
        self.fubiao23_detail_table.setRowCount(2)
        self.fubiao23_detail_table.setVerticalHeaderLabels(['附表', 'shp'])
        self.fubiao23_detail_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.fubiao23_detail_table.setAlternatingRowColors(True)
        self.fubiao23_detail_table.horizontalHeader().setStretchLastSection(True)
        self.fubiao23_detail_table.verticalHeader().setDefaultAlignment(Qt.AlignCenter)
        detail_layout.addWidget(self.fubiao23_detail_table)

        detail_widget.setLayout(detail_layout)
        splitter.addWidget(detail_widget)

        splitter.setSizes([300, 150])
        layout.addWidget(splitter)

        self.fubiao23_result_tab.setLayout(layout)

    def set_report_data(self, report_data):
        """设置附表数据"""
        self.report_data = report_data

    def select_shp_folder(self):
        """选择空间数据文件夹"""
        folder = QFileDialog.getExistingDirectory(self, "选择空间数据文件夹")
        if folder:
            self.shp_folder_edit.setText(folder)
            self.validate_btn.setEnabled(True)
            self.mapping_btn.setEnabled(True)  # 选择文件夹后启用

    def show_field_mapping(self):
        """显示字段映射配置对话框"""
        shp_folder = self.shp_folder_edit.text()
        if not shp_folder:
            QMessageBox.warning(self, "提示", "请先选择空间数据文件夹")
            return

        # 如果还没加载shp数据，先加载
        if not self.validator or not self.validator.shp_data.get('fangzhi'):
            self.validator = DataValidator()
            self.validator.progress_callback = self.append_log
            self.validator.load_fubiao(self.report_data)
            if not self.validator.load_shp_data(shp_folder):
                QMessageBox.warning(self, "警告", "空间数据加载失败")
                return

        # 获取当前字段列表
        fubiao1_fields = []
        fubiao2_fields = []
        fubiao3_fields = []
        fangzhi_fields = self.validator.get_shp_fields('fangzhi')
        yinhuan_fields = self.validator.get_shp_fields('yinhuan')

        if self.report_data:
            fubiao1_records = self.report_data.get('fubiao1', {}).get('records', [])
            fubiao2_records = self.report_data.get('fubiao2', {}).get('records', [])
            fubiao3_records = self.report_data.get('fubiao3', {}).get('records', [])

            if fubiao1_records:
                fubiao1_fields = [k for k in fubiao1_records[0].keys() if not k.startswith('_')]
            if fubiao2_records:
                fubiao2_fields = [k for k in fubiao2_records[0].keys() if not k.startswith('_')]
            if fubiao3_records:
                fubiao3_fields = [k for k in fubiao3_records[0].keys() if not k.startswith('_')]

        # 打开映射配置对话框 - 使用新参数
        dialog = FieldMappingDialog(
            self,
            fubiao1_fields, fubiao2_fields, fubiao3_fields,
            fangzhi_fields, yinhuan_fields,
            self.field_mapping
        )
        if dialog.exec() == QDialog.Accepted:
            self.field_mapping = dialog.get_mapping()
            self.validator.set_field_mapping(self.field_mapping)
            self.append_log("字段映射已更新")

    def toggle_log(self):
        """切换日志显示"""
        if self.log_visible:
            # 隐藏日志
            self.main_splitter.setSizes([1, 0])
            self.log_toggle_btn.setText("显示日志")
            self.log_visible = False
        else:
            # 显示日志
            self.log_panel.show()
            # 设置分割比例：70% 结果，30% 日志
            total_width = self.main_splitter.width()
            self.main_splitter.setSizes([int(total_width * 0.7), int(total_width * 0.3)])
            self.log_toggle_btn.setText("隐藏日志")
            self.log_visible = True

    def clear_results(self):
        """清空结果"""
        self.shp_folder_edit.clear()
        self.log_text.clear()
        self.validation_results = None
        self.fubiao1_detail_data = []
        self.fubiao23_detail_data = []
        self.fubiao1_original_data = []
        self.fubiao23_original_data = []

        # 清空表格
        self.fubiao1_result_table.setRowCount(0)
        self.fubiao1_detail_table.setColumnCount(0)
        self.fubiao23_result_table.setRowCount(0)
        self.fubiao23_detail_table.setColumnCount(0)

        # 重置筛选
        self.fubiao1_status_combo.setCurrentText("全部")
        self.fubiao1_name_edit.clear()
        self.fubiao1_match_combo.setCurrentText("全部")
        self.fubiao23_status_combo.setCurrentText("全部")
        self.fubiao23_name_edit.clear()
        self.fubiao23_match_combo.setCurrentText("全部")

        # 重置状态
        self.validate_btn.setEnabled(False)
        self.export_btn.setEnabled(False)
        self.fubiao1_stats_label.setText("等待校验...")
        self.fubiao23_stats_label.setText("等待校验...")

    def append_log(self, msg):
        """追加日志"""
        self.log_text.append(msg)
        scrollbar = self.log_text.verticalScrollBar()
        scrollbar.setValue(scrollbar.maximum())

    def start_validation(self):
        """开始校验 - 先加载shp数据，再配置字段映射，最后执行校验"""
        shp_folder = self.shp_folder_edit.text()
        if not shp_folder:
            QMessageBox.warning(self, "警告", "请先选择空间数据文件夹")
            return

        if not os.path.exists(shp_folder):
            QMessageBox.warning(self, "警告", "空间数据文件夹不存在")
            return

        # 清空
        self.log_text.clear()
        self.validation_results = None

        # 创建校验器并加载数据
        self.validator = DataValidator()
        self.validator.progress_callback = self.append_log
        self.validator.load_fubiao(self.report_data)

        # 加载空间数据
        self.append_log("正在加载空间数据...")
        if not self.validator.load_shp_data(shp_folder):
            QMessageBox.warning(self, "警告", "空间数据加载失败")
            return

        # 获取字段列表
        fubiao1_fields = self.validator.get_fubiao_fields('fubiao1')
        fubiao2_fields = self.validator.get_fubiao_fields('fubiao2')
        fubiao3_fields = self.validator.get_fubiao_fields('fubiao3')
        fangzhi_fields = self.validator.get_shp_fields('fangzhi')
        yinhuan_fields = self.validator.get_shp_fields('yinhuan')

        # 弹出字段映射配置对话框
        dialog = FieldMappingDialog(
            self,
            fubiao1_fields, fubiao2_fields, fubiao3_fields,
            fangzhi_fields, yinhuan_fields,
            self.field_mapping
        )
        if dialog.exec() != QDialog.Accepted:
            self.append_log("用户取消校验")
            return

        # 应用字段映射
        self.field_mapping = dialog.get_mapping()
        self.validator.set_field_mapping(self.field_mapping)
        self.append_log("字段映射配置完成，开始校验...")

        # 执行校验
        self.validation_results = self.validator.validate_all()

        # 更新结果显示
        self.update_fubiao1_result()
        self.update_fubiao23_result()

        # 启用导出按钮
        self.export_btn.setEnabled(True)

    def update_fubiao1_result(self):
        """更新附表1对比结果"""
        r = self.validation_results.get('fubiao1_vs_fangzhi', {})

        stats_text = f"✓ 匹配成功: {r.get('match_count', 0)} 条   |   " \
                     f"⚠ 仅附表有: {r.get('fubiao_only_count', 0)} 条   |   " \
                     f"⚠ 仅shp有: {r.get('shp_only_count', 0)} 条"
        self.fubiao1_stats_label.setText(stats_text)

        # 获取映射配置
        mapping = self.field_mapping.get('fubiao1_vs_fangzhi', {})
        match_fields = mapping.get('match_fields', {})
        detail_fields = mapping.get('detail_fields', {})

        headers = ['状态', '是否完全匹配', '序号', '名称', '代码', '来源', '错误描述']
        self.fubiao1_detail_data = []

        # 匹配成功的记录
        for item in r.get('matched', []):
            fubiao_rec = item.get('fubiao_record', {})
            shp_recs = item.get('shp_records', [])
            shp_rec = shp_recs[0] if shp_recs else None

            # 获取附表原始序号
            orig_seq = fubiao_rec.get('序号', '')

            # 对比详情字段，找出不一致的
            diff_fields = []
            for shp_field, fb_field in detail_fields.items():
                if not fb_field or not shp_field:
                    continue
                fb_val = str(fubiao_rec.get(fb_field, '')).strip()
                shp_val = str(shp_rec.get(shp_field, '')).strip() if shp_rec else ''
                if fb_val and shp_val and fb_val != shp_val:
                    diff_fields.append(f"{shp_field}不一致")

            is_match = '否' if diff_fields else '是'
            error_desc = '，'.join(diff_fields) if diff_fields else ''

            self.fubiao1_detail_data.append({
                '状态': '✓ 匹配',
                '是否完全匹配': is_match,
                '序号': orig_seq,
                '名称': fubiao_rec.get(match_fields.get('名称', '5.名称'), ''),
                '代码': fubiao_rec.get(match_fields.get('代码', '6.代码'), ''),
                '来源': '附表+shp',
                '错误描述': error_desc,
                '_fubiao_rec': fubiao_rec,
                '_shp_rec': shp_rec
            })

        # 仅附表有的记录
        for rec in r.get('fubiao_only', []):
            orig_seq = rec.get('序号', '')
            self.fubiao1_detail_data.append({
                '状态': '⚠ 仅附表',
                '是否完全匹配': '否',
                '序号': orig_seq,
                '名称': rec.get(match_fields.get('名称', '5.名称'), ''),
                '代码': rec.get(match_fields.get('代码', '6.代码'), ''),
                '来源': '附表',
                '错误描述': 'shp中未找到',
                '_fubiao_rec': rec,
                '_shp_rec': None
            })

        # 仅shp有的记录
        for rec in r.get('shp_only', []):
            self.fubiao1_detail_data.append({
                '状态': '⚠ 仅shp',
                '是否完全匹配': '否',
                '序号': '',
                '名称': rec.get('名称', ''),
                '代码': rec.get('代码', ''),
                '来源': f"shp ({rec.get('_source_folder', '')})",
                '错误描述': '附表中未找到',
                '_fubiao_rec': None,
                '_shp_rec': rec
            })

        self.fubiao1_result_table.clear()
        self.fubiao1_result_table.setColumnCount(len(headers))
        self.fubiao1_result_table.setHorizontalHeaderLabels(headers)
        self.fubiao1_result_table.setRowCount(len(self.fubiao1_detail_data))

        for row, rec in enumerate(self.fubiao1_detail_data):
            for col, header in enumerate(headers):
                value = rec.get(header, '')
                item = QTableWidgetItem(str(value))
                if header == '状态':
                    if '✓' in value:
                        item.setForeground(QColor('#27ae60'))
                    elif '⚠' in value:
                        item.setForeground(QColor('#e74c3c'))
                self.fubiao1_result_table.setItem(row, col, item)

        self.fubiao1_result_table.resizeColumnsToContents()

        self.fubiao1_original_data = self.fubiao1_detail_data.copy()
        self.fubiao1_status_combo.setCurrentText("全部")
        self.fubiao1_name_edit.clear()
        self.fubiao1_match_combo.setCurrentText("全部")

        if len(self.fubiao1_detail_data) > 0:
            self.fubiao1_result_table.selectRow(0)

    def apply_fubiao1_result_filter(self):
        """应用附表1结果筛选"""
        status_filter = self.fubiao1_status_combo.currentText()
        name_filter = self.fubiao1_name_edit.text().strip().lower()
        match_filter = self.fubiao1_match_combo.currentText()

        filtered_data = []
        for rec in self.fubiao1_original_data:
            status = rec.get('状态', '')
            name = str(rec.get('名称', '')).lower()
            is_match = rec.get('是否完全匹配', '')

            if status_filter != "全部" and status != status_filter:
                continue
            if name_filter and name_filter not in name:
                continue
            if match_filter != "全部" and is_match != match_filter:
                continue

            filtered_data.append(rec)

        # 更新表格
        self.fubiao1_detail_data = filtered_data
        headers = ['状态', '是否完全匹配', '序号', '名称', '代码', '来源', '错误描述']
        self.fubiao1_result_table.clear()
        self.fubiao1_result_table.setColumnCount(len(headers))
        self.fubiao1_result_table.setHorizontalHeaderLabels(headers)
        self.fubiao1_result_table.setRowCount(len(filtered_data))

        for row, rec in enumerate(filtered_data):
            for col, header in enumerate(headers):
                value = rec.get(header, '')
                item = QTableWidgetItem(str(value))
                if header == '状态':
                    if '✓' in value:
                        item.setForeground(QColor('#27ae60'))
                    elif '⚠' in value:
                        item.setForeground(QColor('#e74c3c'))
                self.fubiao1_result_table.setItem(row, col, item)

        self.fubiao1_result_table.resizeColumnsToContents()

        # 选中第一行
        if len(filtered_data) > 0:
            self.fubiao1_result_table.selectRow(0)

    def clear_fubiao1_result_filter(self):
        """清除附表1结果筛选"""
        self.fubiao1_status_combo.setCurrentText("全部")
        self.fubiao1_name_edit.clear()
        self.fubiao1_match_combo.setCurrentText("全部")
        self.apply_fubiao1_result_filter()

    def on_fubiao1_selection_changed(self):
        """附表1结果选择变化时显示详情（横向对比）"""
        selected_rows = self.fubiao1_result_table.selectedItems()
        if not selected_rows:
            return

        row = selected_rows[0].row()
        if row >= len(self.fubiao1_detail_data):
            return

        data = self.fubiao1_detail_data[row]
        fubiao_rec = data.get('_fubiao_rec', {})
        shp_rec = data.get('_shp_rec')

        # 从字段映射配置获取详情字段
        mapping = self.field_mapping.get('fubiao1_vs_fangzhi', {})
        detail_fields = mapping.get('detail_fields', {})

        # 合并匹配字段和详情字段
        match_fields = mapping.get('match_fields', {})
        all_fields = {**match_fields, **detail_fields}

        # 构建字段映射列表：(显示名称, 附表字段, shp字段)
        # 新格式是 {shp字段: 附表字段}
        field_mapping = []
        for shp_field, fb_field in all_fields.items():
            field_mapping.append((shp_field, fb_field, shp_field))

        # 横向对比：列是字段，行是附表/shp
        self.fubiao1_detail_table.setColumnCount(len(field_mapping))
        self.fubiao1_detail_table.setHorizontalHeaderLabels([f[0] for f in field_mapping])
        self.fubiao1_detail_table.setRowCount(2)
        self.fubiao1_detail_table.setVerticalHeaderLabels(['附表', 'shp'])

        for col, (label, fubiao_field, shp_field) in enumerate(field_mapping):
            fubiao_val = str(fubiao_rec.get(fubiao_field, '')) if fubiao_rec else '-'
            shp_val = str(shp_rec.get(shp_field, '')) if shp_rec and shp_field else '-'

            fubiao_item = QTableWidgetItem(fubiao_val if fubiao_val else '-')
            shp_item = QTableWidgetItem(shp_val if shp_val else '-')

            # 高亮不一致的字段
            if fubiao_rec and shp_rec and shp_field and fubiao_val != shp_val and fubiao_val and shp_val:
                fubiao_item.setBackground(QColor('#fff3cd'))
                shp_item.setBackground(QColor('#fff3cd'))

            self.fubiao1_detail_table.setItem(0, col, fubiao_item)
            self.fubiao1_detail_table.setItem(1, col, shp_item)

        self.fubiao1_detail_table.resizeColumnsToContents()

    def update_fubiao23_result(self):
        """更新附表2/3对比结果"""
        r = self.validation_results.get('fubiao23_vs_yinhuan', {})

        stats_text = f"附表2: ✓ {r.get('fubiao2_match_count', 0)} 条, ⚠ {r.get('fubiao2_only_count', 0)} 条   |   " \
                     f"附表3: ✓ {r.get('fubiao3_match_count', 0)} 条, ⚠ {r.get('fubiao3_only_count', 0)} 条   |   " \
                     f"仅shp: {r.get('shp_only_count', 0)} 条"
        self.fubiao23_stats_label.setText(stats_text)

        headers = ['状态', '是否完全匹配', '序号', '名称', '编码', '来源表', '来源', '错误描述']
        self.fubiao23_detail_data = []

        # 获取映射配置
        mapping2 = self.field_mapping.get('fubiao2_vs_yinhuan', {})
        mapping3 = self.field_mapping.get('fubiao3_vs_yinhuan', {})
        match_fields2 = mapping2.get('match_fields', {})
        match_fields3 = mapping3.get('match_fields', {})
        detail_fields2 = mapping2.get('detail_fields', {})
        detail_fields3 = mapping3.get('detail_fields', {})

        # 获取名称和编码字段名
        name_field2 = match_fields2.get('名称', '名称')
        code_field2 = match_fields2.get('编号', '编码')
        name_field3 = match_fields3.get('名称', '名称')
        code_field3 = match_fields3.get('编号', '编码')

        # 附表2匹配成功的记录
        for item in r.get('fubiao2_matched', []):
            fubiao_rec = item.get('fubiao_record', {})
            shp_recs = item.get('shp_records', [])
            shp_rec = shp_recs[0] if shp_recs else None

            # 获取附表原始序号
            orig_seq = fubiao_rec.get('序号', '')

            # 对比详情字段
            diff_fields = []
            for shp_field, fb_field in detail_fields2.items():
                if not fb_field or not shp_field:
                    continue
                fb_val = str(fubiao_rec.get(fb_field, '')).strip()
                shp_val = str(shp_rec.get(shp_field, '')).strip() if shp_rec else ''
                if fb_val and shp_val and fb_val != shp_val:
                    diff_fields.append(f"{shp_field}不一致")

            is_match = '否' if diff_fields else '是'
            error_desc = '，'.join(diff_fields) if diff_fields else ''

            self.fubiao23_detail_data.append({
                '状态': '✓ 匹配',
                '是否完全匹配': is_match,
                '序号': orig_seq,
                '名称': fubiao_rec.get(name_field2, ''),
                '编码': fubiao_rec.get(code_field2, ''),
                '来源表': '附表2',
                '来源': '附表2+shp',
                '错误描述': error_desc,
                '_fubiao_rec': fubiao_rec,
                '_shp_rec': shp_rec
            })

        # 附表2仅附表有的记录
        for rec in r.get('fubiao2_only', []):
            orig_seq = rec.get('序号', '')
            self.fubiao23_detail_data.append({
                '状态': '⚠ 仅附表2',
                '是否完全匹配': '否',
                '序号': orig_seq,
                '名称': rec.get(name_field2, ''),
                '编码': rec.get(code_field2, ''),
                '来源表': '附表2',
                '来源': '附表2',
                '错误描述': 'shp中未找到',
                '_fubiao_rec': rec,
                '_shp_rec': None
            })

        # 附表3匹配成功的记录
        for item in r.get('fubiao3_matched', []):
            fubiao_rec = item.get('fubiao_record', {})
            shp_recs = item.get('shp_records', [])
            shp_rec = shp_recs[0] if shp_recs else None

            # 获取附表原始序号
            orig_seq = fubiao_rec.get('序号', '')

            # 对比详情字段
            diff_fields = []
            for shp_field, fb_field in detail_fields3.items():
                if not fb_field or not shp_field:
                    continue
                fb_val = str(fubiao_rec.get(fb_field, '')).strip()
                shp_val = str(shp_rec.get(shp_field, '')).strip() if shp_rec else ''
                if fb_val and shp_val and fb_val != shp_val:
                    diff_fields.append(f"{shp_field}不一致")

            is_match = '否' if diff_fields else '是'
            error_desc = '，'.join(diff_fields) if diff_fields else ''

            self.fubiao23_detail_data.append({
                '状态': '✓ 匹配',
                '是否完全匹配': is_match,
                '序号': orig_seq,
                '名称': fubiao_rec.get(name_field3, ''),
                '编码': fubiao_rec.get(code_field3, ''),
                '来源表': '附表3',
                '来源': '附表3+shp',
                '错误描述': error_desc,
                '_fubiao_rec': fubiao_rec,
                '_shp_rec': shp_rec
            })

        # 附表3仅附表有的记录
        for rec in r.get('fubiao3_only', []):
            orig_seq = rec.get('序号', '')
            self.fubiao23_detail_data.append({
                '状态': '⚠ 仅附表3',
                '是否完全匹配': '否',
                '序号': orig_seq,
                '名称': rec.get(name_field3, ''),
                '编码': rec.get(code_field3, ''),
                '来源表': '附表3',
                '来源': '附表3',
                '错误描述': 'shp中未找到',
                '_fubiao_rec': rec,
                '_shp_rec': None
            })

        # 仅shp有的记录
        for rec in r.get('shp_only', []):
            self.fubiao23_detail_data.append({
                '状态': '⚠ 仅shp',
                '是否完全匹配': '否',
                '序号': '',
                '名称': rec.get('名称', ''),
                '编码': rec.get('编号', ''),
                '来源表': '',
                '来源': f"shp ({rec.get('_source_folder', '')})",
                '错误描述': '附表中未找到',
                '_fubiao_rec': None,
                '_shp_rec': rec
            })

        self.fubiao23_result_table.clear()
        self.fubiao23_result_table.setColumnCount(len(headers))
        self.fubiao23_result_table.setHorizontalHeaderLabels(headers)
        self.fubiao23_result_table.setRowCount(len(self.fubiao23_detail_data))

        for row, rec in enumerate(self.fubiao23_detail_data):
            for col, header in enumerate(headers):
                value = rec.get(header, '')
                item = QTableWidgetItem(str(value))
                if header == '状态':
                    if '✓' in value:
                        item.setForeground(QColor('#27ae60'))
                    elif '⚠' in value:
                        item.setForeground(QColor('#e74c3c'))
                self.fubiao23_result_table.setItem(row, col, item)

        self.fubiao23_result_table.resizeColumnsToContents()

        self.fubiao23_original_data = self.fubiao23_detail_data.copy()

        self.fubiao23_status_combo.setCurrentText("全部")
        self.fubiao23_name_edit.clear()
        self.fubiao23_match_combo.setCurrentText("全部")

        if len(self.fubiao23_detail_data) > 0:
            self.fubiao23_result_table.selectRow(0)

    def apply_fubiao23_result_filter(self):
        """应用附表2/3结果筛选"""
        status_filter = self.fubiao23_status_combo.currentText()
        name_filter = self.fubiao23_name_edit.text().strip().lower()
        match_filter = self.fubiao23_match_combo.currentText()

        filtered_data = []
        for rec in self.fubiao23_original_data:
            status = rec.get('状态', '')
            name = str(rec.get('名称', '')).lower()
            is_match = rec.get('是否完全匹配', '')

            if status_filter != "全部" and status != status_filter:
                continue
            if name_filter and name_filter not in name:
                continue
            if match_filter != "全部" and is_match != match_filter:
                continue

            filtered_data.append(rec)

        # 更新表格
        self.fubiao23_detail_data = filtered_data
        headers = ['状态', '是否完全匹配', '序号', '名称', '编码', '来源表', '来源', '错误描述']
        self.fubiao23_result_table.clear()
        self.fubiao23_result_table.setColumnCount(len(headers))
        self.fubiao23_result_table.setHorizontalHeaderLabels(headers)
        self.fubiao23_result_table.setRowCount(len(filtered_data))

        for row, rec in enumerate(filtered_data):
            for col, header in enumerate(headers):
                value = rec.get(header, '')
                item = QTableWidgetItem(str(value))
                if header == '状态':
                    if '✓' in value:
                        item.setForeground(QColor('#27ae60'))
                    elif '⚠' in value:
                        item.setForeground(QColor('#e74c3c'))
                self.fubiao23_result_table.setItem(row, col, item)

        self.fubiao23_result_table.resizeColumnsToContents()

        # 选中第一行
        if len(filtered_data) > 0:
            self.fubiao23_result_table.selectRow(0)

    def clear_fubiao23_result_filter(self):
        """清除附表2/3结果筛选"""
        self.fubiao23_status_combo.setCurrentText("全部")
        self.fubiao23_name_edit.clear()
        self.fubiao23_match_combo.setCurrentText("全部")
        self.apply_fubiao23_result_filter()

    def on_fubiao23_selection_changed(self):
        """附表2/3结果选择变化时显示详情（横向对比）"""
        selected_rows = self.fubiao23_result_table.selectedItems()
        if not selected_rows:
            return

        row = selected_rows[0].row()
        if row >= len(self.fubiao23_detail_data):
            return

        data = self.fubiao23_detail_data[row]
        fubiao_rec = data.get('_fubiao_rec', {})
        shp_rec = data.get('_shp_rec')
        source_table = data.get('来源表', '')

        # 根据来源表选择映射配置
        if source_table == '附表2':
            mapping = self.field_mapping.get('fubiao2_vs_yinhuan', {})
        elif source_table == '附表3':
            mapping = self.field_mapping.get('fubiao3_vs_yinhuan', {})
        else:
            mapping = {}

        # 从字段映射配置获取详情字段
        detail_fields = mapping.get('detail_fields', {})

        # 合并匹配字段和详情字段
        match_fields = mapping.get('match_fields', {})
        all_fields = {**match_fields, **detail_fields}

        # 构建字段映射列表：(显示名称, 附表字段, shp字段)
        # 新格式是 {shp字段: 附表字段}
        field_mapping = []
        for shp_field, fb_field in all_fields.items():
            field_mapping.append((shp_field, fb_field, shp_field))

        # 横向对比：列是字段，行是附表/shp
        self.fubiao23_detail_table.setColumnCount(len(field_mapping))
        self.fubiao23_detail_table.setHorizontalHeaderLabels([f[0] for f in field_mapping])
        self.fubiao23_detail_table.setRowCount(2)
        self.fubiao23_detail_table.setVerticalHeaderLabels(['附表', 'shp'])

        for col, (label, fubiao_field, shp_field) in enumerate(field_mapping):
            # 尝试多个可能的字段名
            fubiao_val = ''
            if fubiao_rec:
                if fubiao_field in fubiao_rec:
                    fubiao_val = str(fubiao_rec.get(fubiao_field, ''))
                else:
                    # 尝试相似字段
                    for k in fubiao_rec.keys():
                        if fubiao_field in k or k in fubiao_field:
                            fubiao_val = str(fubiao_rec.get(k, ''))
                            break

            shp_val = str(shp_rec.get(shp_field, '')) if shp_rec and shp_field else '-'

            fubiao_item = QTableWidgetItem(fubiao_val if fubiao_val else '-')
            shp_item = QTableWidgetItem(shp_val if shp_val else '-')

            # 高亮不一致的字段
            if fubiao_rec and shp_rec and shp_field and fubiao_val != shp_val and fubiao_val and shp_val:
                fubiao_item.setBackground(QColor('#fff3cd'))
                shp_item.setBackground(QColor('#fff3cd'))

            self.fubiao23_detail_table.setItem(0, col, fubiao_item)
            self.fubiao23_detail_table.setItem(1, col, shp_item)

        self.fubiao23_detail_table.resizeColumnsToContents()

    def export_report(self):
        """导出校验报告 - 与tab页显示内容一致"""
        if not self.validation_results:
            QMessageBox.warning(self, "警告", "没有校验结果可导出")
            return

        file, _ = QFileDialog.getSaveFileName(
            self, "导出校验报告", "附表校验报告.xlsx", "Excel Files (*.xlsx)"
        )

        if file:
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

                wb = Workbook()
                header_font = Font(bold=True)
                header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
                header_alignment = Alignment(horizontal='center', vertical='center')
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

                # ========== 附表1结果 ==========
                ws1 = wb.active
                ws1.title = "附表1对比结果"
                r1 = self.validation_results.get('fubiao1_vs_fangzhi', {})

                # 获取映射配置
                mapping1 = self.field_mapping.get('fubiao1_vs_fangzhi', {})
                match_fields1 = mapping1.get('match_fields', {})
                detail_fields1 = mapping1.get('detail_fields', {})

                headers1 = ['状态', '是否完全匹配', '序号', '名称', '代码', '来源', '错误描述']
                ws1.append(headers1)
                for cell in ws1[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = thin_border
                    cell.alignment = header_alignment

                # 匹配成功的记录
                for item in r1.get('matched', []):
                    fubiao_rec = item.get('fubiao_record', {})
                    shp_recs = item.get('shp_records', [])
                    shp_rec = shp_recs[0] if shp_recs else None
                    orig_seq = fubiao_rec.get('序号', '')

                    # 对比详情字段
                    diff_fields = []
                    for shp_field, fb_field in detail_fields1.items():
                        if not fb_field or not shp_field:
                            continue
                        fb_val = str(fubiao_rec.get(fb_field, '')).strip()
                        shp_val = str(shp_rec.get(shp_field, '')).strip() if shp_rec else ''
                        if fb_val and shp_val and fb_val != shp_val:
                            diff_fields.append(f"{shp_field}不一致")

                    is_match = '否' if diff_fields else '是'
                    error_desc = '，'.join(diff_fields) if diff_fields else ''

                    ws1.append([
                        '✓ 匹配',
                        is_match,
                        orig_seq,
                        fubiao_rec.get(match_fields1.get('名称', '5.名称'), ''),
                        fubiao_rec.get(match_fields1.get('代码', '6.代码'), ''),
                        '附表+shp',
                        error_desc
                    ])

                # 仅附表有的记录
                for rec in r1.get('fubiao_only', []):
                    ws1.append([
                        '⚠ 仅附表',
                        '否',
                        rec.get('序号', ''),
                        rec.get(match_fields1.get('名称', '5.名称'), ''),
                        rec.get(match_fields1.get('代码', '6.代码'), ''),
                        '附表',
                        'shp中未找到'
                    ])

                # 仅shp有的记录
                for rec in r1.get('shp_only', []):
                    ws1.append([
                        '⚠ 仅shp',
                        '否',
                        '',
                        rec.get('名称', ''),
                        rec.get('代码', ''),
                        f"shp ({rec.get('_source_folder', '')})",
                        '附表中未找到'
                    ])

                # 设置数据行边框
                for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row, max_col=len(headers1)):
                    for cell in row:
                        cell.border = thin_border

                # ========== 附表2/3结果 ==========
                ws2 = wb.create_sheet("附表2-3对比结果")
                r2 = self.validation_results.get('fubiao23_vs_yinhuan', {})

                # 获取映射配置
                mapping2 = self.field_mapping.get('fubiao2_vs_yinhuan', {})
                mapping3 = self.field_mapping.get('fubiao3_vs_yinhuan', {})
                match_fields2 = mapping2.get('match_fields', {})
                match_fields3 = mapping3.get('match_fields', {})
                detail_fields2 = mapping2.get('detail_fields', {})
                detail_fields3 = mapping3.get('detail_fields', {})

                headers2 = ['状态', '是否完全匹配', '序号', '名称', '编码', '来源表', '来源', '错误描述']
                ws2.append(headers2)
                for cell in ws2[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = thin_border
                    cell.alignment = header_alignment

                # 附表2匹配成功的记录
                for item in r2.get('fubiao2_matched', []):
                    fubiao_rec = item.get('fubiao_record', {})
                    shp_recs = item.get('shp_records', [])
                    shp_rec = shp_recs[0] if shp_recs else None
                    orig_seq = fubiao_rec.get('序号', '')

                    diff_fields = []
                    for shp_field, fb_field in detail_fields2.items():
                        if not fb_field or not shp_field:
                            continue
                        fb_val = str(fubiao_rec.get(fb_field, '')).strip()
                        shp_val = str(shp_rec.get(shp_field, '')).strip() if shp_rec else ''
                        if fb_val and shp_val and fb_val != shp_val:
                            diff_fields.append(f"{shp_field}不一致")

                    is_match = '否' if diff_fields else '是'
                    error_desc = '，'.join(diff_fields) if diff_fields else ''

                    ws2.append([
                        '✓ 匹配',
                        is_match,
                        orig_seq,
                        fubiao_rec.get(match_fields2.get('名称', '名称'), ''),
                        fubiao_rec.get(match_fields2.get('编号', '编码'), ''),
                        '附表2',
                        '附表2+shp',
                        error_desc
                    ])

                # 附表2仅附表有的记录
                for rec in r2.get('fubiao2_only', []):
                    ws2.append([
                        '⚠ 仅附表2',
                        '否',
                        rec.get('序号', ''),
                        rec.get(match_fields2.get('名称', '名称'), ''),
                        rec.get(match_fields2.get('编号', '编码'), ''),
                        '附表2',
                        '附表2',
                        'shp中未找到'
                    ])

                # 附表3匹配成功的记录
                for item in r2.get('fubiao3_matched', []):
                    fubiao_rec = item.get('fubiao_record', {})
                    shp_recs = item.get('shp_records', [])
                    shp_rec = shp_recs[0] if shp_recs else None
                    orig_seq = fubiao_rec.get('序号', '')

                    diff_fields = []
                    for shp_field, fb_field in detail_fields3.items():
                        if not fb_field or not shp_field:
                            continue
                        fb_val = str(fubiao_rec.get(fb_field, '')).strip()
                        shp_val = str(shp_rec.get(shp_field, '')).strip() if shp_rec else ''
                        if fb_val and shp_val and fb_val != shp_val:
                            diff_fields.append(f"{shp_field}不一致")

                    is_match = '否' if diff_fields else '是'
                    error_desc = '，'.join(diff_fields) if diff_fields else ''

                    ws2.append([
                        '✓ 匹配',
                        is_match,
                        orig_seq,
                        fubiao_rec.get(match_fields3.get('名称', '名称'), ''),
                        fubiao_rec.get(match_fields3.get('编号', '编码'), ''),
                        '附表3',
                        '附表3+shp',
                        error_desc
                    ])

                # 附表3仅附表有的记录
                for rec in r2.get('fubiao3_only', []):
                    ws2.append([
                        '⚠ 仅附表3',
                        '否',
                        rec.get('序号', ''),
                        rec.get(match_fields3.get('名称', '名称'), ''),
                        rec.get(match_fields3.get('编号', '编码'), ''),
                        '附表3',
                        '附表3',
                        'shp中未找到'
                    ])

                # 仅shp有的记录
                for rec in r2.get('shp_only', []):
                    ws2.append([
                        '⚠ 仅shp',
                        '否',
                        '',
                        rec.get('名称', ''),
                        rec.get('编号', ''),
                        '',
                        f"shp ({rec.get('_source_folder', '')})",
                        '附表中未找到'
                    ])

                # 设置数据行边框
                for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row, max_col=len(headers2)):
                    for cell in row:
                        cell.border = thin_border

                wb.save(file)
                QMessageBox.information(self, "完成", f"已导出到:\n{file}")

            except Exception as e:
                QMessageBox.critical(self, "错误", f"导出失败:\n{str(e)}")


class FieldMappingDialog(QDialog):
    """字段映射配置对话框 - 以shp字段为主，映射到附表字段"""

    def __init__(self, parent, fubiao1_fields, fubiao2_fields, fubiao3_fields, fangzhi_fields, yinhuan_fields, current_mapping):
        super().__init__(parent)
        self.setWindowTitle("字段映射配置")
        self.setMinimumSize(700, 600)
        self.fubiao1_fields = fubiao1_fields or []  # 附表1字段
        self.fubiao2_fields = fubiao2_fields or []  # 附表2字段
        self.fubiao3_fields = fubiao3_fields or []  # 附表3字段
        self.fangzhi_fields = fangzhi_fields or []  # 防治对象shp字段
        self.yinhuan_fields = yinhuan_fields or []  # 隐患要素shp字段
        self.current_mapping = current_mapping.copy()
        self.combos = {}  # 保存所有下拉框引用
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout()

        # 说明
        info_label = QLabel("配置shp字段与附表字段的对应关系。左侧为shp字段，右侧选择对应的附表字段。")
        info_label.setStyleSheet("color: #7f8c8d; padding: 10px;")
        layout.addWidget(info_label)

        # Tab区域
        tabs = QTabWidget()

        # === 防治对象 ↔ 附表1 映射Tab ===
        tab1 = QWidget()
        tab1_layout = QVBoxLayout()

        match_group1 = QGroupBox("匹配字段（用于记录匹配）")
        match_layout1 = QFormLayout()
        self.combos['fangzhi_名称'] = self.create_combo(self.fubiao1_fields, '')
        self.combos['fangzhi_代码'] = self.create_combo(self.fubiao1_fields, '')
        match_layout1.addRow("shp[名称] →", self.combos['fangzhi_名称'])
        match_layout1.addRow("shp[代码] →", self.combos['fangzhi_代码'])
        match_group1.setLayout(match_layout1)
        tab1_layout.addWidget(match_group1)

        detail_group1 = QGroupBox("详情字段（用于对比展示）")
        detail_layout1 = QFormLayout()
        match_shp_fields = ['名称', '代码']
        for shp_field in self.fangzhi_fields:
            if shp_field in match_shp_fields or shp_field.startswith('_'):
                continue
            key = f'fangzhi_{shp_field}'
            self.combos[key] = self.create_combo(self.fubiao1_fields, '')
            detail_layout1.addRow(f"shp[{shp_field}] →", self.combos[key])
        detail_group1.setLayout(detail_layout1)
        tab1_layout.addWidget(detail_group1)

        auto_btn1 = QPushButton("自动匹配")
        auto_btn1.clicked.connect(self.auto_match_fubiao1)
        tab1_layout.addWidget(auto_btn1)
        tab1_layout.addStretch()
        tab1.setLayout(tab1_layout)
        tabs.addTab(tab1, "防治对象 ↔ 附表1")

        # === 隐患要素 ↔ 附表2 映射Tab ===
        tab2 = QWidget()
        tab2_layout = QVBoxLayout()

        match_group2 = QGroupBox("匹配字段（用于记录匹配）")
        match_layout2 = QFormLayout()
        self.combos['yinhuan2_名称'] = self.create_combo(self.fubiao2_fields, '')
        self.combos['yinhuan2_编号'] = self.create_combo(self.fubiao2_fields, '')
        match_layout2.addRow("shp[名称] →", self.combos['yinhuan2_名称'])
        match_layout2.addRow("shp[编号] →", self.combos['yinhuan2_编号'])
        match_group2.setLayout(match_layout2)
        tab2_layout.addWidget(match_group2)

        detail_group2 = QGroupBox("详情字段（用于对比展示）")
        detail_layout2 = QFormLayout()
        match_shp_fields2 = ['名称', '编号']
        for shp_field in self.yinhuan_fields:
            if shp_field in match_shp_fields2 or shp_field.startswith('_'):
                continue
            key = f'yinhuan2_{shp_field}'
            self.combos[key] = self.create_combo(self.fubiao2_fields, '')
            detail_layout2.addRow(f"shp[{shp_field}] →", self.combos[key])
        detail_group2.setLayout(detail_layout2)
        tab2_layout.addWidget(detail_group2)

        auto_btn2 = QPushButton("自动匹配")
        auto_btn2.clicked.connect(self.auto_match_fubiao2)
        tab2_layout.addWidget(auto_btn2)
        tab2_layout.addStretch()
        tab2.setLayout(tab2_layout)
        tabs.addTab(tab2, "隐患要素 ↔ 附表2")

        # === 隐患要素 ↔ 附表3 映射Tab ===
        tab3 = QWidget()
        tab3_layout = QVBoxLayout()

        match_group3 = QGroupBox("匹配字段（用于记录匹配）")
        match_layout3 = QFormLayout()
        self.combos['yinhuan3_名称'] = self.create_combo(self.fubiao3_fields, '')
        self.combos['yinhuan3_编号'] = self.create_combo(self.fubiao3_fields, '')
        match_layout3.addRow("shp[名称] →", self.combos['yinhuan3_名称'])
        match_layout3.addRow("shp[编号] →", self.combos['yinhuan3_编号'])
        match_group3.setLayout(match_layout3)
        tab3_layout.addWidget(match_group3)

        detail_group3 = QGroupBox("详情字段（用于对比展示）")
        detail_layout3 = QFormLayout()
        for shp_field in self.yinhuan_fields:
            if shp_field in match_shp_fields2 or shp_field.startswith('_'):
                continue
            key = f'yinhuan3_{shp_field}'
            self.combos[key] = self.create_combo(self.fubiao3_fields, '')
            detail_layout3.addRow(f"shp[{shp_field}] →", self.combos[key])
        detail_group3.setLayout(detail_layout3)
        tab3_layout.addWidget(detail_group3)

        auto_btn3 = QPushButton("自动匹配")
        auto_btn3.clicked.connect(self.auto_match_fubiao3)
        tab3_layout.addWidget(auto_btn3)
        tab3_layout.addStretch()
        tab3.setLayout(tab3_layout)
        tabs.addTab(tab3, "隐患要素 ↔ 附表3")

        layout.addWidget(tabs)

        # 按钮
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        ok_btn = QPushButton("确定")
        ok_btn.clicked.connect(self.accept)
        cancel_btn = QPushButton("取消")
        cancel_btn.clicked.connect(self.reject)
        btn_layout.addWidget(ok_btn)
        btn_layout.addWidget(cancel_btn)
        layout.addLayout(btn_layout)

        self.setLayout(layout)
        self.load_current_mapping()

    def create_combo(self, fields, default_value):
        """创建下拉框 - 从附表字段列表创建"""
        combo = QComboBox()
        combo.addItem("-- 不映射 --", "")
        for field in fields:
            combo.addItem(field, field)
        # 设置默认值
        if default_value:
            idx = combo.findData(default_value)
            if idx >= 0:
                combo.setCurrentIndex(idx)
        return combo

    def load_current_mapping(self):
        """加载当前映射配置"""
        # 新格式：{shp字段: 附表字段}
        mapping1 = self.current_mapping.get('fubiao1_vs_fangzhi', {})
        match_fields1 = mapping1.get('match_fields', {})
        detail_fields1 = mapping1.get('detail_fields', {})

        mapping23 = self.current_mapping.get('fubiao23_vs_yinhuan', {})
        match_fields23 = mapping23.get('match_fields', {})
        detail_fields23 = mapping23.get('detail_fields', {})

        # 设置防治对象下拉框
        self.set_combo_value('fangzhi_名称', match_fields1.get('名称', ''))
        self.set_combo_value('fangzhi_代码', match_fields1.get('代码', ''))
        for shp_field, fubiao_field in detail_fields1.items():
            key = f'fangzhi_{shp_field}'
            self.set_combo_value(key, fubiao_field)

        # 设置附表2隐患要素下拉框
        mapping2 = self.current_mapping.get('fubiao2_vs_yinhuan', {})
        match_fields2 = mapping2.get('match_fields', {})
        detail_fields2 = mapping2.get('detail_fields', {})
        self.set_combo_value('yinhuan2_名称', match_fields2.get('名称', ''))
        self.set_combo_value('yinhuan2_编号', match_fields2.get('编号', ''))
        for shp_field, fubiao_field in detail_fields2.items():
            key = f'yinhuan2_{shp_field}'
            self.set_combo_value(key, fubiao_field)

        # 设置附表3隐患要素下拉框
        mapping3 = self.current_mapping.get('fubiao3_vs_yinhuan', {})
        match_fields3 = mapping3.get('match_fields', {})
        detail_fields3 = mapping3.get('detail_fields', {})
        self.set_combo_value('yinhuan3_名称', match_fields3.get('名称', ''))
        self.set_combo_value('yinhuan3_编号', match_fields3.get('编号', ''))
        for shp_field, fubiao_field in detail_fields3.items():
            key = f'yinhuan3_{shp_field}'
            self.set_combo_value(key, fubiao_field)

    def set_combo_value(self, key, value):
        """设置下拉框值"""
        if key in self.combos and value:
            combo = self.combos[key]
            idx = combo.findData(value)
            if idx >= 0:
                combo.setCurrentIndex(idx)

    def auto_match_fubiao1(self):
        """自动匹配防治对象字段 - 根据shp字段名模糊匹配附表字段"""
        if not self.fubiao1_fields:
            QMessageBox.warning(self, "提示", "附表字段列表为空")
            return

        auto_map = {
            '名称': self._find_field(self.fubiao1_fields, '名称'),
            '代码': self._find_field(self.fubiao1_fields, '代码'),
            '类型': self._find_field(self.fubiao1_fields, '类型'),
            '人口': self._find_field(self.fubiao1_fields, '人口'),
            '河流名称': self._find_field(self.fubiao1_fields, '河流名称'),
            '河流代码': self._find_field(self.fubiao1_fields, '河流代码'),
        }

        for shp_field, fubiao_field in auto_map.items():
            if fubiao_field:
                key = f'fangzhi_{shp_field}'
                self.set_combo_value(key, fubiao_field)

        QMessageBox.information(self, "完成", "已自动匹配字段")

    def auto_match_fubiao2(self):
        """自动匹配附表2字段"""
        if not self.fubiao2_fields:
            QMessageBox.warning(self, "提示", "附表2字段列表为空")
            return

        auto_map = {
            '名称': self._find_field(self.fubiao2_fields, '名称'),
            '编号': self._find_field(self.fubiao2_fields, '编码') or self._find_field(self.fubiao2_fields, '编号'),
            '类型': self._find_field(self.fubiao2_fields, '类型'),
            '河流名称': self._find_field(self.fubiao2_fields, '河流名称'),
            '河流代码': self._find_field(self.fubiao2_fields, '河流代码'),
            '经度': self._find_field(self.fubiao2_fields, '经度'),
            '纬度': self._find_field(self.fubiao2_fields, '纬度'),
        }

        for shp_field, fubiao_field in auto_map.items():
            if fubiao_field:
                key = f'yinhuan2_{shp_field}'
                self.set_combo_value(key, fubiao_field)

        QMessageBox.information(self, "完成", "已自动匹配字段")

    def auto_match_fubiao3(self):
        """自动匹配附表3字段"""
        if not self.fubiao3_fields:
            QMessageBox.warning(self, "提示", "附表3字段列表为空")
            return

        auto_map = {
            '名称': self._find_field(self.fubiao3_fields, '名称'),
            '编号': self._find_field(self.fubiao3_fields, '编码') or self._find_field(self.fubiao3_fields, '编号'),
            '类型': self._find_field(self.fubiao3_fields, '类型'),
            '河流名称': self._find_field(self.fubiao3_fields, '河流名称'),
            '河流代码': self._find_field(self.fubiao3_fields, '河流代码'),
            '经度': self._find_field(self.fubiao3_fields, '经度'),
            '纬度': self._find_field(self.fubiao3_fields, '纬度'),
        }

        for shp_field, fubiao_field in auto_map.items():
            if fubiao_field:
                key = f'yinhuan3_{shp_field}'
                self.set_combo_value(key, fubiao_field)

        QMessageBox.information(self, "完成", "已自动匹配字段")

    def _find_field(self, field_list, keyword):
        """在字段列表中模糊查找包含关键词的字段"""
        for field in field_list:
            if keyword in field:
                return field
        return None

    def get_mapping(self):
        """获取映射配置 - 返回 {shp字段: 附表字段} 格式"""
        # 构建防治对象映射
        match_fields1 = {}
        detail_fields1 = {}

        if 'fangzhi_名称' in self.combos:
            val = self.combos['fangzhi_名称'].currentData()
            if val:
                match_fields1['名称'] = val
        if 'fangzhi_代码' in self.combos:
            val = self.combos['fangzhi_代码'].currentData()
            if val:
                match_fields1['代码'] = val

        match_shp_fields = ['名称', '代码']
        for shp_field in self.fangzhi_fields:
            if shp_field in match_shp_fields or shp_field.startswith('_'):
                continue
            key = f'fangzhi_{shp_field}'
            if key in self.combos:
                val = self.combos[key].currentData()
                if val:
                    detail_fields1[shp_field] = val

        # 构建附表2隐患要素映射
        match_fields2 = {}
        detail_fields2 = {}

        if 'yinhuan2_名称' in self.combos:
            val = self.combos['yinhuan2_名称'].currentData()
            if val:
                match_fields2['名称'] = val
        if 'yinhuan2_编号' in self.combos:
            val = self.combos['yinhuan2_编号'].currentData()
            if val:
                match_fields2['编号'] = val

        match_shp_fields2 = ['名称', '编号']
        for shp_field in self.yinhuan_fields:
            if shp_field in match_shp_fields2 or shp_field.startswith('_'):
                continue
            key = f'yinhuan2_{shp_field}'
            if key in self.combos:
                val = self.combos[key].currentData()
                if val:
                    detail_fields2[shp_field] = val

        # 构建附表3隐患要素映射
        match_fields3 = {}
        detail_fields3 = {}

        if 'yinhuan3_名称' in self.combos:
            val = self.combos['yinhuan3_名称'].currentData()
            if val:
                match_fields3['名称'] = val
        if 'yinhuan3_编号' in self.combos:
            val = self.combos['yinhuan3_编号'].currentData()
            if val:
                match_fields3['编号'] = val

        for shp_field in self.yinhuan_fields:
            if shp_field in match_shp_fields2 or shp_field.startswith('_'):
                continue
            key = f'yinhuan3_{shp_field}'
            if key in self.combos:
                val = self.combos[key].currentData()
                if val:
                    detail_fields3[shp_field] = val

        return {
            'fubiao1_vs_fangzhi': {
                'match_fields': match_fields1,
                'detail_fields': detail_fields1
            },
            'fubiao2_vs_yinhuan': {
                'match_fields': match_fields2,
                'detail_fields': detail_fields2
            },
            'fubiao3_vs_yinhuan': {
                'match_fields': match_fields3,
                'detail_fields': detail_fields3
            }
        }