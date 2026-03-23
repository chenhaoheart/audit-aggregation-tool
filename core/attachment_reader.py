# -*- coding: utf-8 -*-
"""
附表数据读取模块
处理Excel附表文件的读取，支持合并单元格
"""

import os
import pandas as pd
from openpyxl import load_workbook


class AttachmentReader:
    def __init__(self):
        self.attachment_files = {}
        self.attachment_data = {}

    def set_file(self, key, file_path):
        if file_path and os.path.exists(file_path):
            self.attachment_files[key] = file_path
        else:
            self.attachment_files[key] = None

    def load_excel_with_merged_cells(self, file_path):
        if not file_path or not os.path.exists(file_path):
            return pd.DataFrame()

        try:
            wb = load_workbook(file_path, data_only=True)
            ws = wb.active

            merged_cells_ranges = list(ws.merged_cells.ranges)
            merged_value_map = {}

            for merged_range in merged_cells_ranges:
                top_left_cell = ws.cell(merged_range.min_row, merged_range.min_col)
                top_left_value = top_left_cell.value
                for row in range(merged_range.min_row, merged_range.max_row + 1):
                    for col in range(merged_range.min_col, merged_range.max_col + 1):
                        merged_value_map[(row, col)] = top_left_value

            data = []
            headers = None
            for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
                row_data = []
                for col_idx, cell_value in enumerate(row, 1):
                    if (row_idx, col_idx) in merged_value_map:
                        row_data.append(merged_value_map[(row_idx, col_idx)])
                    else:
                        row_data.append(cell_value)

                if row_idx == 1:
                    headers = [str(h) if h is not None else f'列{col_idx}' for col_idx, h in enumerate(row_data, 1)]
                else:
                    if any(v is not None for v in row_data):
                        record = {}
                        for col_idx, value in enumerate(row_data):
                            if col_idx < len(headers):
                                record[headers[col_idx]] = value
                        data.append(record)

            wb.close()
            df = pd.DataFrame(data)
            return df
        except Exception as e:
            print(f"加载Excel失败 {file_path}: {str(e)}")
            return pd.DataFrame()

    def load_all(self):
        self.attachment_data = {}

        for key, file_path in self.attachment_files.items():
            if file_path:
                df = self.load_excel_with_merged_cells(file_path)
                self.attachment_data[key] = {
                    'file_path': file_path,
                    'file_name': os.path.basename(file_path),
                    'data': df,
                    'record_count': len(df),
                    'columns': list(df.columns) if not df.empty else []
                }
            else:
                self.attachment_data[key] = {
                    'file_path': '',
                    'file_name': '',
                    'data': pd.DataFrame(),
                    'record_count': 0,
                    'columns': []
                }

        return self.attachment_data

    def get_data(self, key):
        return self.attachment_data.get(key, {})

    def get_all_data(self):
        return self.attachment_data
