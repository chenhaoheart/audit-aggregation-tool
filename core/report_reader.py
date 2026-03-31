# -*- coding: utf-8 -*-
"""
成果报表数据读取模块
"""

import os
import pandas as pd
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from openpyxl import load_workbook


def parse_checkbox(value) -> str:
    """
    解析勾选符号
    Wingdings 2 字体中：
    - 'R' 或 'þ' 显示为带框对号 -> ☑
    - 其他非空值显示为带框叉号 -> ☒
    """
    if pd.isna(value):
        return ''
    val = str(value).strip()
    # Wingdings 2 字体中 'R' 和 'þ' 都显示为带框对号
    if val in ['R', 'r', 'þ']:
        return '☑'
    elif val:
        return '☒'
    return ''


def find_report_files(folder_path: str) -> Dict[str, Optional[str]]:
    """
    在文件夹及其子目录中查找3个附表文件

    Returns:
        dict: {'fubiao1': path, 'fubiao2': path, 'fubiao3': path}
    """
    result = {
        'fubiao1': None,
        'fubiao2': None,
        'fubiao3': None
    }

    if not folder_path or not os.path.exists(folder_path):
        return result

    folder = Path(folder_path)

    # 递归查找所有xlsx文件
    for file in folder.rglob('*.xlsx'):
        name = file.name
        if '附表1' in name and '防治对象名录' in name:
            result['fubiao1'] = str(file)
        elif '附表2' in name and ('跨沟道路' in name or '桥涵' in name):
            result['fubiao2'] = str(file)
        elif '附表3' in name and '沟滩占地' in name:
            result['fubiao3'] = str(file)

    return result


def get_merge_cells_info(file_path: str, data_start_row: int = 3) -> List[Tuple[int, int, int, int]]:
    """
    获取数据行的合并单元格信息

    Returns:
        List of (起始行, 起始列, 行数, 列数)，行和列都是0-indexed
    """
    if not file_path or not os.path.exists(file_path):
        return []

    wb = load_workbook(file_path)
    ws = wb.active

    merge_info = []
    for merged_range in ws.merged_cells.ranges:
        min_row = merged_range.min_row - 1  # 转为0-indexed
        max_row = merged_range.max_row - 1
        min_col = merged_range.min_col - 1
        max_col = merged_range.max_col - 1

        # 只处理数据行的合并（从第4行开始，索引为3）
        if min_row >= data_start_row:
            row_span = max_row - min_row + 1
            col_span = max_col - min_col + 1
            # 转换为相对于数据行的索引
            merge_info.append((min_row - data_start_row, min_col, row_span, col_span))

    wb.close()
    return merge_info


def load_fubiao1(file_path: str) -> Tuple[List[Dict], List[str], Dict]:
    """
    读取附表1 - 山洪灾害防治对象名录

    该表有复杂的多级合并表头，前3行为表头行
    对县名、乡镇名等基础信息列使用ffill填充合并单元格

    Returns:
        (records, headers, merge_info): 记录列表、表头列表、合并信息
    """
    if not file_path or not os.path.exists(file_path):
        return [], [], {}

    # 读取 Excel 文件，不自动转换表头
    df = pd.read_excel(file_path, header=None)

    # 去掉所有字符串的前后空白
    df = df.map(lambda x: x.strip() if isinstance(x, str) else x)

    # 数据从第4行开始（索引3），前3行是表头
    data_start_row = 3

    # 分离表头和数据
    header_df = df.iloc[:data_start_row]
    data_df = df.iloc[data_start_row:].copy()

    # 对列0-10（序号、县名、县代码、乡镇名、乡镇代码、名称、代码、类型、人口、河流名称、河流代码）填充合并单元格
    # 这些是防治对象的基础信息和唯一标识字段
    for col in [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10]:
        if col < len(data_df.columns):
            data_df.iloc[:, col] = data_df.iloc[:, col].ffill()

    # 获取合并单元格信息
    merge_cells = get_merge_cells_info(file_path, data_start_row)

    # 定义表头（29列）
    headers = [
        '序号', '1.县（区、市、旗）名称', '2.县（区、市、旗）代码',
        '3.乡镇名称', '4.乡镇代码', '5.名称', '6.代码',
        '7.类型', '8.人口', '9.河流名称', '10.河流代码',
        '11.名称', '12.编码', '13.名称', '14.编码',
        '15.河流名称', '16.河流代码',
        '17.束窄', '18.急弯', '19.低洼地', '20.临河滑坡', '21.泥石流',
        '22.沟滩占地', '23.溃决', '24.壅水', '25.顶托', '26.改道', '27.漫流',
        '28.备注'
    ]

    records = []

    # 勾选列的索引（0-based）：列17-27
    checkbox_col_indices = list(range(16, 27))  # 索引16-26对应列17-27

    # 遍历数据行
    for i in range(len(data_df)):
        row = data_df.iloc[i]
        record = {}

        for col_idx in range(min(29, len(row))):
            val = row[col_idx]
            header = headers[col_idx]

            # 处理勾选符号（列17-27）
            if col_idx in checkbox_col_indices:
                record[header] = parse_checkbox(val)
            elif pd.notna(val):
                record[header] = val
            else:
                record[header] = ''

        records.append(record)

    return records, headers, {'merge_cells': merge_cells}


def load_fubiao2(file_path: str) -> Tuple[List[Dict], List[str]]:
    """
    读取附表2 - 跨沟道路、桥涵、塘（堰）坝调查成果表

    Returns:
        (records, headers): 记录列表和表头列表
    """
    if not file_path or not os.path.exists(file_path):
        return [], []

    # 读取原始表头
    df_header = pd.read_excel(file_path, header=None, nrows=1)
    headers = [str(df_header.iloc[0, i]) if pd.notna(df_header.iloc[0, i]) else f'列{i}'
               for i in range(len(df_header.columns))]

    # 读取数据
    df = pd.read_excel(file_path)

    records = []
    for _, row in df.iterrows():
        record = {}
        for i, col in enumerate(df.columns):
            val = row[col]
            if pd.notna(val):
                record[headers[i]] = val
            else:
                record[headers[i]] = ''
        records.append(record)

    return records, headers


def load_fubiao3(file_path: str) -> Tuple[List[Dict], List[str]]:
    """
    读取附表3 - 沟滩占地情况调查成果表

    Returns:
        (records, headers): 记录列表和表头列表
    """
    if not file_path or not os.path.exists(file_path):
        return [], []

    # 读取原始表头
    df_header = pd.read_excel(file_path, header=None, nrows=1)
    headers = [str(df_header.iloc[0, i]) if pd.notna(df_header.iloc[0, i]) else f'列{i}'
               for i in range(len(df_header.columns))]

    # 读取数据
    df = pd.read_excel(file_path)

    records = []
    for _, row in df.iterrows():
        record = {}
        for i, col in enumerate(df.columns):
            val = row[col]
            if pd.notna(val):
                record[headers[i]] = val
            else:
                record[headers[i]] = ''
        records.append(record)

    return records, headers


def load_all_reports(folder_path: str) -> Dict:
    """
    从文件夹加载所有成果报表

    Returns:
        {
            'fubiao1': {'records': [], 'headers': [], 'merge_info': [], 'file': ''},
            'fubiao2': {'records': [], 'headers': [], 'file': ''},
            'fubiao3': {'records': [], 'headers': [], 'file': ''},
            'missing': []  # 未找到的文件列表
        }
    """
    result = {
        'fubiao1': {'records': [], 'headers': [], 'merge_info': [], 'file': None},
        'fubiao2': {'records': [], 'headers': [], 'file': None},
        'fubiao3': {'records': [], 'headers': [], 'file': None},
        'missing': []
    }

    files = find_report_files(folder_path)

    # 加载附表1
    if files['fubiao1']:
        records, headers, merge_info = load_fubiao1(files['fubiao1'])
        result['fubiao1'] = {'records': records, 'headers': headers, 'merge_info': merge_info, 'file': files['fubiao1']}
    else:
        result['missing'].append('附表1_山洪灾害防治对象名录.xlsx')

    # 加载附表2
    if files['fubiao2']:
        records, headers = load_fubiao2(files['fubiao2'])
        result['fubiao2'] = {'records': records, 'headers': headers, 'file': files['fubiao2']}
    else:
        result['missing'].append('附表2_跨沟道路、桥涵、塘（堰）坝调查成果表.xlsx')

    # 加载附表3
    if files['fubiao3']:
        records, headers = load_fubiao3(files['fubiao3'])
        result['fubiao3'] = {'records': records, 'headers': headers, 'file': files['fubiao3']}
    else:
        result['missing'].append('附表3_沟滩占地情况调查成果表.xlsx')

    return result